#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_ai.py —— excel-ai-analyst 技能配套工具脚本（单文件，四个子命令）

把含公式的业务 Excel 当作"没有文档的遗留代码"来逆向工程，本脚本负责其中的机械工作：

    tomd     Step 1  Excel → 结构化 MD（多行表头 / 合并单元格 / 真实公式 / 列画像）
    verify   Step 4  按 spec.json 用全量真实数据逐行回算，与 Excel 现值比对（核心）
    output   Step 5  按验证明细产出带配色标注的结果 xlsx（含血缘 / 本体 / 校验汇总）
    analyze  Step 5  按验证明细产出业务分析报告（分组 / 分布 / TopN / 离群 / 规则 / What-If）

设计纪律（每一条都是踩过的坑）：
  * 绝不修改用户原表，所有产物只写到 -o 指定的输出目录
  * 表达式一律走 ast 白名单**递归求值器**，不使用 eval/exec/compile（安全边界，见 SafeEval）
  * 非数值单元格（"/"、"-"、"不适用"、空）一律按 0.0 参与运算
  * 金额比较用容差，不用 ==
  * 所有 CSV 一律 utf-8-sig，否则用户用 Excel 打开是乱码

仅依赖 pandas + openpyxl + 标准库。
"""

import argparse
import ast
import csv
import datetime as _dt
import difflib
import json
import keyword as _kw
import math
import os
import re
import sys
from collections import OrderedDict

# ---------------------------------------------------------------------------
# 依赖检查：缺依赖时给出中文安装提示，而不是抛 traceback。
# 注意这里只做"软导入"，保证 `--help` 在没装依赖时也能正常显示。
# ---------------------------------------------------------------------------
try:
    import pandas as pd
    import openpyxl
    from openpyxl.utils import range_boundaries, get_column_letter
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _DEP_ERROR = None
except Exception as _e:  # pragma: no cover - 仅在缺依赖环境触发
    pd = None
    openpyxl = None
    range_boundaries = get_column_letter = None
    Alignment = Border = Font = PatternFill = Side = None
    _DEP_ERROR = _e


def require_deps():
    """在真正干活前调用；缺依赖时打印中文提示并退出。"""
    if _DEP_ERROR is None:
        return
    sys.stderr.write(
        "\n[依赖缺失] 本脚本需要 pandas 与 openpyxl，当前环境导入失败：\n"
        "    %s\n\n"
        "请先安装依赖后重试：\n"
        "    pip3 install pandas openpyxl\n"
        "若系统 Python 受保护（externally-managed-environment），加参数：\n"
        "    pip3 install pandas openpyxl --break-system-packages\n\n" % _DEP_ERROR
    )
    sys.exit(2)


VERSION = "1.0.0"

# 配色（SKILL.md 规定，勿改）
FILL_INPUT = "DAEEF3"   # 蓝：原始输入（来自 Excel 原表）
FILL_AI = "E2EFDA"      # 绿：AI 计算（脚本按 spec 回算）
FILL_DIFF = "FCE4D6"    # 橙：差异非零（超出容差）

# 汇总行/小计行的默认识别词
DEFAULT_SUMMARY_WORDS = ["汇总", "合计", "总计", "小计", "总额", "共计", "Total", "TOTAL", "total"]

# 表示"不参加/无此项"的常见占位符（按 0 处理，但业务上要与真正的 0 区分）
PLACEHOLDER_TOKENS = {"/", "\\", "-", "--", "—", "－", "N/A", "n/a", "NA", "null", "NULL",
                      "无", "不适用", "不参加", "空", "None", "#N/A", "#DIV/0!", "#VALUE!", "#REF!"}

SEP = "·"  # 校验项三列的分隔符：名·AI / 名·Excel / 名·差异

# SKILL.md：金额容差 0.01、比率容差 0.0001，**不要为了让报告好看去放大容差**。
# 超过 LOOSE_TOL 的容差一律点名；差异大于 STRICT_TOL 却因为容差被判通过的，逐项计数。
STRICT_TOL = 0.01
LOOSE_TOL = 1.0

# verification_detail.csv 的保留列名，变量名不许占用（否则 CSV 出现同名列，
# DictReader 静默丢列，output/analyze 拿到的是串了的数据）
RESERVED_DETAIL_COLS = {"Excel行号"}

# 全角/特殊符号 → 半角（中文表里 －500、−1、（300）、１２３４ 都很常见）
FULLWIDTH_MAP = {
    "－": "-", "−": "-", "–": "-", "—": "-", "﹣": "-",   # 各种减号/连字符
    "＋": "+", "﹢": "+",
    "．": ".", "。": ".",
    "（": "(", "）": ")",
    "％": "%",
}
for _i in range(10):                       # 全角数字 ０-９
    FULLWIDTH_MAP[chr(0xFF10 + _i)] = str(_i)
_FULLWIDTH_RE = re.compile("[%s]" % re.escape("".join(FULLWIDTH_MAP)))


# ===========================================================================
# 一、通用小工具
# ===========================================================================

def is_blank(v):
    """单元格是否为空（None / NaN / 纯空白字符串）。"""
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    # pandas 的 NA / NaT
    try:
        if pd is not None and v is not None and not isinstance(v, (str, bytes)) and pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    return False


def is_datelike(v):
    """是否是日期/时间类单元格（pandas Timestamp 也是 datetime 的子类）。"""
    if isinstance(v, bool):
        return False
    return isinstance(v, (_dt.datetime, _dt.date, _dt.time, _dt.timedelta))


def cell_str(v):
    """单元格 → 展示用字符串（空 → ""，浮点整数去掉 .0）。"""
    if is_blank(v):
        return ""
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return ""
        # 只有"确实是整数"的浮点才显示成整数：1e-12 曾被 round(v,10) 抹成 0，
        # 那是在报告里凭空造数据，绝不允许。
        if abs(v) < 1e15 and (v == 0 or abs(v) >= 1e-6) and abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return "%.12g" % v
    if isinstance(v, str):
        return v.strip()
    return str(v)


def norm_key(s):
    """归一化：去掉所有空白与全角空格，用于列名匹配与跨表 key 对齐。"""
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).replace("\u3000", "").strip()


_NUM_CLEAN_RE = re.compile(r"[,\s\u3000¥￥$€,、]")


def to_float(v):
    """
    单元格 → (float 值, 是否为数值)。
    非数值一律返回 (0.0, False) —— SKILL.md 规定：非数值单元格按 0.0 参与运算。
    支持：千分位、货币符号、百分号（除以 100）、会计式括号负数。
    """
    if is_blank(v):
        return 0.0, False
    if isinstance(v, bool):
        return (1.0 if v else 0.0), True
    if isinstance(v, (int, float)):
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return 0.0, False
        return f, True
    if is_datelike(v):
        # 日期/时间不是金额，绝不能悄悄按序列号参与运算
        return 0.0, False
    s = str(v).strip()
    if s in PLACEHOLDER_TOKENS:
        return 0.0, False
    # 全角减号 －、Unicode 减号 −、全角数字 ０-９ 在中文表里极常见，
    # 统一转半角后再解析（此前 "－500" 会被当成非数值按 0 处理，是实打实的算错）
    if _FULLWIDTH_RE.search(s):
        s = _FULLWIDTH_RE.sub(lambda m: FULLWIDTH_MAP[m.group(0)], s).strip()
        if s in PLACEHOLDER_TOKENS:
            return 0.0, False
    neg = False
    if len(s) >= 2 and ((s[0] == "(" and s[-1] == ")") or (s[0] == "（" and s[-1] == "）")):
        neg, s = True, s[1:-1].strip()
    pct = False
    if s.endswith("%") or s.endswith("％"):
        pct, s = True, s[:-1].strip()
    s = _NUM_CLEAN_RE.sub("", s)
    if s in ("", "-", "+", "."):
        return 0.0, False
    try:
        f = float(s)
    except (TypeError, ValueError):
        return 0.0, False
    if math.isnan(f) or math.isinf(f):
        return 0.0, False
    if pct:
        f /= 100.0
    if neg:
        f = -f
    return f, True


def md_escape(v):
    """Markdown 表格单元格转义：竖线与换行。"""
    s = cell_str(v)
    return s.replace("|", "\\|").replace("\r", "").replace("\n", "<br>")


def md_table(headers, rows):
    """生成 Markdown 表格（行为 list[list]）。"""
    out = ["| " + " | ".join(md_escape(h) for h in headers) + " |",
           "|" + "|".join(["---"] * len(headers)) + "|"]
    for r in rows:
        cells = list(r) + [""] * (len(headers) - len(r))
        out.append("| " + " | ".join(md_escape(c) for c in cells[:len(headers)]) + " |")
    return "\n".join(out)


def fmt_num(x, nd=2):
    """数值格式化（带千分位）。"""
    try:
        f = float(x)
    except (TypeError, ValueError):
        return cell_str(x)
    if math.isnan(f) or math.isinf(f):
        return ""
    return format(round(f, nd), ",.%df" % nd)


def ensure_dir(p):
    if not p or os.path.isdir(p):
        return p
    # -o 指到一个已存在的**文件**上时，os.makedirs 会抛 FileExistsError（未捕获的 traceback）
    if os.path.exists(p):
        raise UserError("输出路径 %s 已存在，但它是一个文件而不是目录。\n"
                        "    请把 -o 换成一个目录路径（脚本会自己创建）。" % p)
    try:
        os.makedirs(p, exist_ok=True)
    except OSError as e:
        raise UserError("无法创建输出目录 %s：%s" % (p, e))
    return p


def safe_name(s, maxlen=80):
    """把 Sheet / 文件名变成安全的文件名。"""
    s = re.sub(r'[\\/:*?"<>|\r\n\t]', "_", str(s)).strip().strip(".")
    return (s or "unnamed")[:maxlen]


def write_text(path, text):
    ensure_dir(os.path.dirname(path))
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)
    return path


# 用 Excel 打开 CSV 时会被当成公式起手的字符（CSV 注入 / DDE 注入）。
# `-` 也在列表里，但纯负数不能动，所以判定时会先看它是不是一个合法数值。
_INJECT_PREFIX = ("=", "+", "@", "-", "\t", "\r")


def is_formula_like(s):
    """这段文本被 Excel 打开时会不会被当成公式？（纯负数不算）"""
    if not isinstance(s, str) or s == "":
        return False
    if s[0] not in _INJECT_PREFIX:
        return False
    return not to_float(s)[1]      # "-500" 是数值，放行；"-2+3+cmd|..." 不是，拦下


def csv_guard(v):
    """
    写 CSV 前的防注入转义：给会被 Excel 当公式的文本加前导单引号。
    SKILL.md 明确要求 CSV 用 utf-8-sig「让用户用 Excel 打开」，那 CSV 注入就是真实攻击面：
    原表里一个叫 `=cmd|'/c calc'!A1` 的姓名，会被我们原样搬进交付物。
    """
    return "'" + v if is_formula_like(v) else v


def csv_unguard(s):
    """load_detail 读回时反转义，保证 output/analyze 拿到的还是原值。"""
    if isinstance(s, str) and len(s) >= 2 and s[0] == "'" and s[1] in _INJECT_PREFIX:
        return s[1:]
    return s


def write_csv(path, headers, rows):
    """一律 utf-8-sig，否则用户用 Excel 打开是乱码。"""
    ensure_dir(os.path.dirname(path))
    n_guard = 0
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([csv_guard(h) if isinstance(h, str) else h for h in headers])
        for r in rows:
            out = []
            for c in r:
                if isinstance(c, str) and is_formula_like(c):
                    out.append("'" + c)
                    n_guard += 1
                else:
                    out.append(c)
            w.writerow(out)
    if n_guard:
        sys.stderr.write("[提示] %s 中有 %d 个单元格的文本会被 Excel 当成公式"
                         "（`=`/`+`/`@` 开头），已加前导单引号防注入。\n"
                         % (os.path.basename(path), n_guard))
    return path


def clip(s, n=60):
    """截断长文本，避免把 3 万字的列名原样打进 stderr。"""
    s = str(s)
    return s if len(s) <= n else (s[:n] + "…(共%d字)" % len(s))


class UserError(Exception):
    """面向用户的中文错误（main 捕获后打印，不抛 traceback）。"""


# ---------------------------------------------------------------------------
# spec.json 的类型闸门：JSON 是用户/AI 手写的，写错类型是常态。
# 一律在这里转成 UserError，绝不让 AttributeError / ValueError 冒成 traceback。
# ---------------------------------------------------------------------------

_TYPE_CN = {dict: "对象 {}", list: "数组 []", str: "字符串", bool: "布尔", int: "整数", float: "小数"}


def _tname(v):
    for t in (bool, dict, list, str, int, float):
        if isinstance(v, t):
            return _TYPE_CN[t]
    return "null" if v is None else type(v).__name__


def as_dict(v, where, default=None):
    """要求是 JSON 对象；None 时给默认空 dict。"""
    if v is None:
        return {} if default is None else default
    if not isinstance(v, dict):
        raise UserError("spec 的 `%s` 必须是对象 {\"变量名\": ...}，实际写成了%s：%s"
                        % (where, _tname(v), clip(json.dumps(v, ensure_ascii=False), 120)))
    return v


def as_list(v, where):
    """要求是 JSON 数组；None → []。"""
    if v is None:
        return []
    if not isinstance(v, list):
        raise UserError("spec 的 `%s` 必须是数组 [...]，实际写成了%s：%s"
                        % (where, _tname(v), clip(json.dumps(v, ensure_ascii=False), 120)))
    return v


def as_item_dict(v, where):
    """数组里的每一项必须是对象。"""
    if not isinstance(v, dict):
        raise UserError("spec 的 `%s` 必须是对象 {...}，实际写成了%s：%s"
                        % (where, _tname(v), clip(json.dumps(v, ensure_ascii=False), 120)))
    return v


def as_int(v, where, default=None, minimum=None):
    if v is None:
        if default is None:
            raise UserError("spec 缺少必填字段 `%s`（整数）" % where)
        v = default
    if isinstance(v, bool):
        raise UserError("spec 的 `%s` 必须是整数，实际是布尔值" % where)
    if isinstance(v, float) and not float(v).is_integer():
        raise UserError("spec 的 `%s` 必须是整数，实际是小数 %s" % (where, v))
    try:
        n = int(v)
    except (TypeError, ValueError):
        raise UserError("spec 的 `%s` 必须是整数，实际写成了%s：%s"
                        % (where, _tname(v), clip(v, 40)))
    if minimum is not None and n < minimum:
        raise UserError("spec 的 `%s` 不能小于 %d，实际是 %d" % (where, minimum, n))
    return n


def as_float(v, where, default=None, minimum=None):
    if v is None:
        if default is None:
            raise UserError("spec 缺少必填字段 `%s`（数值）" % where)
        v = default
    if isinstance(v, bool):
        raise UserError("spec 的 `%s` 必须是数值，实际是布尔值" % where)
    try:
        f = float(v)
    except (TypeError, ValueError):
        raise UserError("spec 的 `%s` 必须是数值，实际写成了%s：%s"
                        % (where, _tname(v), clip(v, 40)))
    if math.isnan(f) or math.isinf(f):
        raise UserError("spec 的 `%s` 不能是 NaN / 无穷大" % where)
    if minimum is not None and f < minimum:
        raise UserError("spec 的 `%s` 不能小于 %s，实际是 %s" % (where, minimum, f))
    return f


def as_str(v, where, default=None, allow_empty=False):
    """
    要求是字符串。**凡是后面要拿去做 dict 查找的字段都必须过这道闸**：
    写成数组/对象时 `x in some_dict` 会抛 `TypeError: unhashable type`，
    那是一条未捕获的 traceback，不是中文报错。
    """
    if v is None:
        if default is None:
            raise UserError("spec 缺少必填字段 `%s`（字符串）" % where)
        return default
    if not isinstance(v, str):
        raise UserError("spec 的 `%s` 必须是字符串，实际写成了%s：%s"
                        % (where, _tname(v), clip(json.dumps(v, ensure_ascii=False), 80)))
    if not allow_empty and not v.strip():
        raise UserError("spec 的 `%s` 不能是空字符串" % where)
    return v


# ---------------------------------------------------------------------------
# 未知字段闸门：spec 是手写的，`tolerence` / `dimension` / `key` 这类拼错**最危险**——
# 它们不会报错，只会静默退回默认值（容差悄悄变回 0.01、维度列整段消失），
# 报告照样是漂亮的 100%。所以一律按"未知字段即错误"处理，并给出最接近的正确拼写。
# 想写自由备注请用 `_` 开头的键，或下面这批公认的文档字段。
# ---------------------------------------------------------------------------

DOC_KEYS = {"note", "notes", "comment", "comments", "description", "desc", "title",
            "version", "author", "date", "meta", "readme",
            "说明", "备注", "描述", "注释", "作者", "版本", "日期"}


def check_keys(obj, known, where):
    """obj 里出现 known / DOC_KEYS / `_前缀` 之外的键 → 报错并给拼写建议。"""
    if not isinstance(obj, dict):
        return obj
    for k in obj:
        if not isinstance(k, str):
            continue
        if k in known or k in DOC_KEYS or k.startswith("_"):
            continue
        near = difflib.get_close_matches(k, sorted(known), n=2, cutoff=0.6)
        # 单复数/大小写/下划线的常见手滑，get_close_matches 不一定捞得到
        if not near:
            flat = {re.sub(r"[_\s]", "", x.lower()): x for x in known}
            cand = flat.get(re.sub(r"[_\s]", "", k.lower().rstrip("s")))
            if cand:
                near = [cand]
        raise UserError(
            "spec 的 `%s` 里有未知字段 `%s`%s\n"
            "    合法字段：%s\n"
            "    拼错的字段不会报错、只会静默退回默认值（容差悄悄变回 0.01、维度整段消失），"
            "报告照样是漂亮的 100%%，所以这里一律拦下。\n"
            "    确实要写自由备注，请用 `_` 开头的键名。"
            % (where or "顶层", k,
               ("　你是不是想写 `%s`？" % "` 或 `".join(near)) if near else "",
               "、".join(sorted(known))))
    return obj


def as_str_list(v, where):
    """容忍单个字符串写法：`"合计"` 等价于 `["合计"]`。
    此前会被 list("合计") 拆成单字 ['合','计']，误伤面极大。"""
    if v is None:
        return []
    if isinstance(v, str):
        return [v] if v.strip() else []
    if not isinstance(v, list):
        raise UserError("spec 的 `%s` 必须是字符串数组 [\"...\"]（或单个字符串），实际写成了%s"
                        % (where, _tname(v)))
    out = []
    for i, x in enumerate(v):
        if not isinstance(x, str):
            raise UserError("spec 的 `%s[%d]` 必须是字符串，实际是%s" % (where, i, _tname(x)))
        if x.strip():
            out.append(x)
    return out


# ===========================================================================
# 二、安全表达式求值器（ast 白名单 + 递归求值，不使用 eval/exec/compile）
# ===========================================================================

class SafeEval(object):
    """
    只放行 SKILL.md 规定的语法：
        算术 : + - * / ** %
        比较 : == != < <= > >=
        布尔 : and / or / not
        三元 : x if c else y
        函数 : abs round min max int float sum floor ceil
        字面量: 数字 / 字符串 / True/False/None / 列表 / 元组（供 sum、max 使用）

    明确禁止（安全边界，务必写死）：
        属性访问 a.b、下标 a[i]、导入、lambda、推导式、赋值/海象、f-string、
        字典/集合字面量、以及上面白名单之外的任何调用。
    实现方式是**自己递归遍历 AST 求值**，全程不产生 code 对象，因此不存在
    通过 __class__ / __globals__ 逃逸的可能。
    """

    FUNCS = {
        "abs": abs,
        "round": round,
        "min": min,
        "max": max,
        "int": int,
        "float": float,
        "sum": sum,
        "floor": math.floor,
        "ceil": math.ceil,
    }

    BINOPS = {
        ast.Add: lambda a, b: a + b,
        ast.Sub: lambda a, b: a - b,
        ast.Mult: lambda a, b: a * b,
        ast.Div: lambda a, b: a / b,
        ast.Pow: lambda a, b: a ** b,
        ast.Mod: lambda a, b: a % b,
    }

    CMPOPS = {
        ast.Eq: lambda a, b: a == b,
        ast.NotEq: lambda a, b: a != b,
        ast.Lt: lambda a, b: a < b,
        ast.LtE: lambda a, b: a <= b,
        ast.Gt: lambda a, b: a > b,
        ast.GtE: lambda a, b: a >= b,
    }

    # -- 资源上限（防 DoS；业务公式再复杂也远够用） -------------------------
    MAX_EXPR_LEN = 4000      # 表达式字符数上限
    MAX_DEPTH = 200          # AST 嵌套深度上限（防递归求值把 Python 栈打爆）
    MAX_INT_BITS = 1 << 16   # 中间整数位数上限（约 2 万位十进制）
    MAX_SEQ_LEN = 1000000    # 字符串/列表长度上限（序列复制与 sum 共用）

    def __init__(self):
        self._cache = {}

    # -- 编译（只做解析与静态校验，不生成可执行代码） ----------------------
    def parse(self, expr, where=""):
        key = expr
        if key in self._cache:
            return self._cache[key]
        if not isinstance(expr, str) or expr.strip() == "":
            raise UserError("表达式为空%s" % (("（%s）" % where) if where else ""))
        # 长度上限：`"1+"*60000` 这种超长表达式会在 ast.parse 内部把解释器栈打爆
        # （RecursionError）或吃光内存（MemoryError），必须在解析前就拦掉。
        if len(expr) > self.MAX_EXPR_LEN:
            raise UserError("表达式过长%s：%d 字符，上限 %d。\n"
                            "    业务公式不该这么长，请拆成多个 derived 中间量。" %
                            (("（%s）" % where) if where else "", len(expr), self.MAX_EXPR_LEN))
        try:
            tree = ast.parse(expr, mode="eval")
        except SyntaxError as e:
            raise UserError("表达式语法错误%s：%s\n    %s" %
                            (("（%s）" % where) if where else "", expr, e))
        except (RecursionError, MemoryError):
            # 嵌套过深（如 6 万层一元负号）会让 CPython 的解析器自己先倒下
            raise UserError("表达式嵌套过深%s：%s…\n    请拆成多个 derived 中间量。" %
                            (("（%s）" % where) if where else "", expr[:60]))
        try:
            self._check(tree.body, expr, where)
        except RecursionError:
            raise UserError("表达式嵌套过深%s：%s…\n    请拆成多个 derived 中间量。" %
                            (("（%s）" % where) if where else "", expr[:60]))
        self._cache[key] = tree.body
        return tree.body

    def names(self, expr, where=""):
        """
        静态取出表达式引用的变量名（不含被调用的白名单函数名）。
        用来在**开始逐行求值之前**就查出未定义变量 / derived 循环依赖 ——
        1 万行的表跑到第 1 行才报错，用户看到的是一堆行号噪音而不是问题本身。
        """
        node = self.parse(expr, where)
        called = set()
        for nd in ast.walk(node):
            if isinstance(nd, ast.Call) and isinstance(nd.func, ast.Name):
                called.add(id(nd.func))
        return OrderedDict.fromkeys(
            nd.id for nd in ast.walk(node)
            if isinstance(nd, ast.Name) and id(nd) not in called)

    def _reject(self, node, expr, where, why):
        raise UserError(
            "表达式含不允许的语法%s：%s\n    原因：%s（节点 %s）\n"
            "    只允许：算术 + - * / ** %% 、比较、and/or/not、三元、"
            "以及函数 abs round min max int float sum floor ceil" %
            (("（%s）" % where) if where else "", expr, why, type(node).__name__))

    def _check(self, node, expr, where, depth=0):
        # 深度上限：递归求值器的栈深度 ≈ AST 深度。不设限就会被 `1+1+1+…` 这类
        # 超深表达式打成 RecursionError —— 那是未捕获的 traceback，不是中文报错。
        if depth > self.MAX_DEPTH:
            raise UserError("表达式嵌套过深%s（超过 %d 层）：%s…\n"
                            "    请拆成多个 derived 中间量。" %
                            (("（%s）" % where) if where else "", self.MAX_DEPTH, expr[:60]))
        d = depth + 1
        if isinstance(node, ast.Constant):
            if not isinstance(node.value, (int, float, str, bool, type(None))):
                self._reject(node, expr, where, "不支持的字面量类型")
            return
        if isinstance(node, ast.Name):
            if not isinstance(node.ctx, ast.Load):
                self._reject(node, expr, where, "禁止赋值")
            return
        if isinstance(node, ast.UnaryOp):
            if not isinstance(node.op, (ast.UAdd, ast.USub, ast.Not)):
                self._reject(node, expr, where, "不支持的一元运算")
            self._check(node.operand, expr, where, d)
            return
        if isinstance(node, ast.BinOp):
            if type(node.op) not in self.BINOPS:
                self._reject(node, expr, where, "不支持的运算符（只允许 + - * / ** %）")
            self._check(node.left, expr, where, d)
            self._check(node.right, expr, where, d)
            return
        if isinstance(node, ast.BoolOp):
            if not isinstance(node.op, (ast.And, ast.Or)):
                self._reject(node, expr, where, "不支持的布尔运算")
            for v in node.values:
                self._check(v, expr, where, d)
            return
        if isinstance(node, ast.Compare):
            for op in node.ops:
                if type(op) not in self.CMPOPS:
                    self._reject(node, expr, where, "不支持的比较运算（in/is 等一律禁止）")
            self._check(node.left, expr, where, d)
            for c in node.comparators:
                self._check(c, expr, where, d)
            return
        if isinstance(node, ast.IfExp):
            self._check(node.test, expr, where, d)
            self._check(node.body, expr, where, d)
            self._check(node.orelse, expr, where, d)
            return
        if isinstance(node, (ast.List, ast.Tuple)):
            if not isinstance(node.ctx, ast.Load):
                self._reject(node, expr, where, "禁止赋值")
            for e in node.elts:
                self._check(e, expr, where, d)
            return
        if isinstance(node, ast.Call):
            if not isinstance(node.func, ast.Name):
                self._reject(node, expr, where, "禁止属性访问式调用 a.b()")
            if node.func.id not in self.FUNCS:
                raise UserError(
                    "表达式调用了未授权的函数 %s()%s：%s\n    白名单：%s" %
                    (node.func.id, ("（%s）" % where) if where else "", expr,
                     " ".join(sorted(self.FUNCS))))
            if node.keywords:
                self._reject(node, expr, where, "函数调用不支持关键字参数")
            for a in node.args:
                if isinstance(a, ast.Starred):
                    self._reject(node, expr, where, "禁止 *args 展开")
                self._check(a, expr, where, d)
            return
        # 其余一律拒绝：Attribute / Subscript / Lambda / 推导式 / Dict / Set /
        # JoinedStr / NamedExpr / Await / Yield / Starred ...
        self._reject(node, expr, where, "该语法未在白名单内")

    # -- 求值期的资源守卫 ---------------------------------------------------
    def _guard_pow(self, a, b):
        """
        幂运算守卫。**必须在真正做 a ** b 之前拦截**，事后再检查已经晚了。

        只卡"指数大小"是不够的：指数每层都 ≤ 1000 合规，但底数可以是上一层
        算出来的巨型整数，于是 `((10**999)**999)**999` 三层叠加就能滚到 10 亿位、
        几百 MB，把进程 OOM 掉（实测被内核 SIGKILL）。所以这里同时卡**结果规模**：
        结果位数 ≈ 底数位数 × 指数，超过 MAX_INT_BITS 直接拒绝。
        """
        if isinstance(b, bool) or not isinstance(b, (int, float)):
            raise UserError("表达式的幂次必须是数字")
        try:
            if abs(float(b)) > 1000:
                raise UserError("表达式的幂次过大（|指数| > 1000），已拒绝求值")
        except OverflowError:  # b 本身就是个巨型整数
            raise UserError("表达式的幂次过大（|指数| > 1000），已拒绝求值")
        if isinstance(a, int) and not isinstance(a, bool) and isinstance(b, (int, float)) and b > 0:
            bits = a.bit_length() * float(b)
            if bits > self.MAX_INT_BITS:
                raise UserError(
                    "表达式的幂运算结果过大（约 %d 位二进制，上限 %d），已拒绝求值。\n"
                    "    注意：`(10**999)**999` 这类嵌套幂每层指数都不大，"
                    "但结果会指数级叠加，足以把内存吃干。"
                    % (int(bits), self.MAX_INT_BITS))

    def _guard_call(self, name, args):
        """
        白名单函数的参数守卫。

        `sum` 的第二个参数是累加初值；传一个 list 进去就变成列表拼接，
        `sum([[1]] * 300000, [])` 是 O(n²) 的内存炸弹（实测被内核 SIGKILL）。
        业务公式里 sum 只该用来加数字。
        """
        if name == "sum":
            if len(args) >= 2 and (isinstance(args[1], (list, tuple, str))
                                   or not isinstance(args[1], (int, float))):
                raise UserError("sum() 的第二个参数（累加初值）必须是数字，"
                                "不允许用它做列表/字符串拼接。")
            if args and isinstance(args[0], (list, tuple)) and len(args[0]) > self.MAX_SEQ_LEN:
                raise UserError("sum() 的序列长度 %d 超过上限 %d，已拒绝求值。"
                                % (len(args[0]), self.MAX_SEQ_LEN))

    # -- 求值 -------------------------------------------------------------
    def eval(self, expr, env, where=""):
        node = self.parse(expr, where)
        try:
            return self._ev(node, env)
        except UserError:
            raise
        except ZeroDivisionError:
            raise UserError("表达式除零%s：%s" % (("（%s）" % where) if where else "", expr))
        except Exception as e:
            raise UserError("表达式求值失败%s：%s\n    %s: %s" %
                            (("（%s）" % where) if where else "", expr, type(e).__name__, e))

    def _ev(self, node, env):
        if isinstance(node, ast.Constant):
            return node.value
        if isinstance(node, ast.Name):
            if node.id in env:
                return env[node.id]
            if node.id in self.FUNCS:
                # 绝不把白名单函数对象本身当值返回：调用走的是 ast.Call 分支
                # （直接查 FUNCS），这里返回只会把一个活的内建对象漏进数据流，
                # 而且 eval_num 会把它悄悄折成 0.0 —— 静默的错误答案。
                raise UserError("函数名 %s 必须写成 %s(...) 的调用形式，不能当作值使用。" %
                                (node.id, node.id))
            main = sorted(k for k in env if not k.endswith("__raw"))
            raise UserError("表达式引用了未定义的变量：%s\n    当前可用变量：%s\n"
                            "    （另外每个变量都有一个 `<变量名>__raw` 版本，取未转数值的原始文本）" %
                            (node.id, ", ".join(main) or "(无)"))
        if isinstance(node, ast.UnaryOp):
            v = self._ev(node.operand, env)
            if isinstance(node.op, ast.UAdd):
                return +v
            if isinstance(node.op, ast.USub):
                return -v
            return not v
        if isinstance(node, ast.BinOp):
            a = self._ev(node.left, env)
            b = self._ev(node.right, env)
            if isinstance(node.op, ast.Pow):
                self._guard_pow(a, b)
            if isinstance(node.op, ast.Mult):
                # 防 "x" * 10**9 这类序列复制把内存吃干（逐行求值会放大 N 倍）
                for x, y in ((a, b), (b, a)):
                    if isinstance(x, (str, list, tuple)) and isinstance(y, (int, float)) \
                            and not isinstance(y, bool) and abs(y) * max(1, len(x)) > self.MAX_SEQ_LEN:
                        raise UserError("表达式试图把字符串/序列复制 %d 次，已拒绝求值（上限 %d）"
                                        % (int(abs(y)), self.MAX_SEQ_LEN))
            return self.BINOPS[type(node.op)](a, b)
        if isinstance(node, ast.BoolOp):
            if isinstance(node.op, ast.And):
                r = True
                for v in node.values:
                    r = self._ev(v, env)
                    if not r:
                        return r
                return r
            r = False
            for v in node.values:
                r = self._ev(v, env)
                if r:
                    return r
            return r
        if isinstance(node, ast.Compare):
            left = self._ev(node.left, env)
            for op, comp in zip(node.ops, node.comparators):
                right = self._ev(comp, env)
                if not self.CMPOPS[type(op)](left, right):
                    return False
                left = right
            return True
        if isinstance(node, ast.IfExp):
            return self._ev(node.body, env) if self._ev(node.test, env) else self._ev(node.orelse, env)
        if isinstance(node, ast.List):
            return [self._ev(e, env) for e in node.elts]
        if isinstance(node, ast.Tuple):
            return tuple(self._ev(e, env) for e in node.elts)
        if isinstance(node, ast.Call):
            fn = self.FUNCS[node.func.id]
            argv = [self._ev(a, env) for a in node.args]
            self._guard_call(node.func.id, argv)
            return fn(*argv)
        raise UserError("内部错误：未处理的表达式节点 %s" % type(node).__name__)


SAFE = SafeEval()


def eval_num(expr, env, where=""):
    """求值并转成 float（布尔按 1/0）。"""
    v = SAFE.eval(expr, env, where)
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        # 巨型整数（如 10**999）转 float 会抛 OverflowError；那是 SAFE.eval
        # 之外的地方，不接住就是一条未捕获的 traceback。
        try:
            f = float(v)
        except OverflowError:
            raise UserError("表达式的计算结果超出浮点数范围%s：%s\n"
                            "    结果大到无法作为金额参与比对，请检查公式。"
                            % (("（%s）" % where) if where else "", expr))
        return 0.0 if (math.isnan(f) or math.isinf(f)) else f
    f, _ok = to_float(v)
    return f


def eval_bool(expr, env, where=""):
    return bool(SAFE.eval(expr, env, where))


# ===========================================================================
# 三、Excel 读取：pandas 读值 + openpyxl 读真实公式 + 合并单元格铺开
# ===========================================================================

class SheetData(object):
    """一个 Sheet 的完整快照。"""

    def __init__(self, path, sheet):
        self.path = path
        self.sheet = sheet
        self.values = []      # list[list]  pandas 读到的值（公式列是缓存结果）
        self.filled = []      # list[list]  合并单元格已铺开的值（仅用于拼表头）
        self.formulas = []    # list[list]  openpyxl data_only=False 抽到的真实公式（无则 ""）
        self.merged = []      # [(min_col, min_row, max_col, max_row)] 1-based
        self.nrows = 0
        self.ncols = 0

    def val(self, r, c):
        if 0 <= r < self.nrows and 0 <= c < len(self.values[r]):
            return self.values[r][c]
        return None

    def fml(self, r, c):
        if 0 <= r < self.nrows and 0 <= c < len(self.formulas[r]):
            return self.formulas[r][c]
        return ""


SUPPORTED_EXT = (".xlsx", ".xlsm", ".xltx", ".xltm")


def check_workbook_path(path):
    """
    打开任何工作簿之前都先过这道闸：不存在 / 旧版 .xls / 根本不是 Excel，
    一律给中文引导，而不是让 pandas 抛一句英文 ImportError。
    """
    if not os.path.isfile(path):
        raise UserError("找不到工作簿：%s" % path)
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        raise UserError(
            "这是旧版 .xls 格式，本脚本（openpyxl）只支持 .xlsx/.xlsm：%s\n"
            "    请先转成 xlsx 再跑，任选一种：\n"
            "      soffice --headless --convert-to xlsx --outdir <输出目录> '%s'\n"
            "      或用 Excel/WPS 打开后「另存为 → Excel 工作簿(*.xlsx)」" % (path, path))
    if ext == ".csv":
        raise UserError("这是 CSV 不是 Excel 工作簿：%s\n"
                        "    CSV 没有多行表头/合并单元格/公式，本脚本的价值也就无从谈起；"
                        "请先用 Excel 另存为 .xlsx。" % path)
    if ext not in SUPPORTED_EXT:
        raise UserError("不支持的文件类型 `%s`：%s\n    只支持：%s"
                        % (ext or "(无扩展名)", path, "、".join(SUPPORTED_EXT)))
    # 内容嗅探：xlsx 本质是 zip，必须以 PK 开头
    try:
        with open(path, "rb") as f:
            magic = f.read(4)
    except OSError as e:
        raise UserError("无法读取文件：%s\n    %s" % (path, e))
    if magic[:2] == b"\xd0\xcf":
        raise UserError("文件扩展名叫 %s，内容却是旧版 .xls（OLE2 复合文档）：%s\n"
                        "    请先转换：soffice --headless --convert-to xlsx --outdir <输出目录> '%s'"
                        % (ext, path, path))
    if magic[:2] != b"PK":
        raise UserError("这不是一个有效的 xlsx 文件（xlsx 本质是 zip，应以 PK 开头）：%s\n"
                        "    可能是被改了扩展名的文本/CSV/旧版 xls，请确认来源。" % path)


def load_sheet(path, sheet, fill_merged_rows=None):
    """
    读入一个 Sheet。
      * pandas header=None 读全表（不让 pandas 猜表头）
      * openpyxl data_only=False 抽真实公式
      * 用 range_boundaries 把合并单元格的值铺满其覆盖区（写进 .filled）

    铺开只写进 .filled（拼表头 / 识别大标题行用），**.values 永远是未铺开的原值**，
    因此数据区不会被复制翻倍。fill_merged_rows 给 int N 时只铺前 N 行。
    """
    require_deps()
    check_workbook_path(path)
    # Sheet 名必须是字符串。写成 null / 数组时 pandas 会**读回一个 dict**（多表模式），
    # 后面 df.itertuples 直接 AttributeError；写成数字 0 更糟——它会静默按下标读第一个
    # Sheet，报告里印着 `Sheet 0`，跨表比对全程对着一张错的表，还一路 rc=0。
    if not isinstance(sheet, str) or not sheet.strip():
        raise UserError(
            "Sheet 名必须是字符串（工作簿 %s），实际写成了%s：%s\n"
            "    写成数字会被按下标解释成另一张表、写成 null/数组会让 pandas 一次读回所有表，"
            "两种都会让后面的比对对着错的数据跑。" %
            (os.path.basename(path), _tname(sheet), clip(repr(sheet), 40)))

    sd = SheetData(path, sheet)
    try:
        df = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
    except Exception as e:
        raise UserError("pandas 读取失败：%s / %s\n    %s: %s" % (path, sheet, type(e).__name__, e))

    sd.values = [list(row) for row in df.itertuples(index=False, name=None)]
    sd.nrows = len(sd.values)
    sd.ncols = int(df.shape[1]) if sd.nrows else 0
    for r in range(sd.nrows):
        if len(sd.values[r]) < sd.ncols:
            sd.values[r] += [None] * (sd.ncols - len(sd.values[r]))

    # openpyxl：真实公式 + 合并区域
    sd.formulas = [[""] * sd.ncols for _ in range(sd.nrows)]
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
        ws = wb[sheet] if sheet in wb.sheetnames else None
        if ws is not None:
            for row in ws.iter_rows():
                for c in row:
                    r0, c0 = c.row - 1, c.column - 1
                    if r0 >= sd.nrows or c0 >= sd.ncols:
                        continue
                    v = c.value
                    if isinstance(v, str) and v.startswith("="):
                        sd.formulas[r0][c0] = v
                    elif v is not None and type(v).__name__ == "ArrayFormula":
                        sd.formulas[r0][c0] = str(getattr(v, "text", "") or "=<数组公式>")
            for rng in list(ws.merged_cells.ranges):
                sd.merged.append(range_boundaries(str(rng.coord if hasattr(rng, "coord") else rng)))
        wb.close()
    except UserError:
        raise
    except Exception as e:
        sys.stderr.write("[提示] openpyxl 读取公式/合并信息失败（不影响取值）：%s: %s\n"
                         % (type(e).__name__, e))

    # 合并单元格铺开（默认只铺表头扫描区）
    limit = sd.nrows if fill_merged_rows is None else int(fill_merged_rows)
    sd.filled = [list(r) for r in sd.values]
    for (min_col, min_row, max_col, max_row) in sd.merged:
        r0, c0 = min_row - 1, min_col - 1
        if r0 >= sd.nrows or c0 >= sd.ncols:
            continue
        v = sd.values[r0][c0]
        if is_blank(v):
            continue
        for rr in range(min_row - 1, min(max_row, sd.nrows)):
            if rr >= limit:
                continue
            for cc in range(min_col - 1, min(max_col, sd.ncols)):
                sd.filled[rr][cc] = v
    return sd


def is_big_title_row(row_vals, ncols):
    """
    '整行只有一个取值且横跨大半张表' 的大标题行（如"XX公司2024年5月工资表"）。
    合并单元格铺开后，这类行会变成同一个值重复 N 次。

    那个唯一取值**必须是非数值文本**：一行各项全为 0 的数据行（停薪留职、当月未发）
    同样满足"只有一个取值且横跨整表"，此前会被误判成大标题行而整行丢掉。
    """
    vals = [cell_str(v) for v in row_vals[:ncols]]
    nonblank = [v for v in vals if v != ""]
    if not nonblank:
        return False
    if len(set(nonblank)) != 1:
        return False
    only = nonblank[0]
    if to_float(only)[1] or len(only) < 2:   # 数字（含全 0 行）不可能是大标题
        return False
    return len(nonblank) >= max(3, int(math.ceil(ncols * 0.6)))


def row_is_summary(row_vals, words=None):
    """整行任意单元格含汇总词 → 判为汇总/小计行。"""
    ws = words if words is not None else DEFAULT_SUMMARY_WORDS
    for v in row_vals:
        s = cell_str(v)
        if not s:
            continue
        for w in ws:
            if w and w in s:
                return True
    return False


def detect_data_start(sd, scan=8):
    """
    启发式定位数据体起始行（返回 0-based 行号，同时也就是表头行数）。
    信号（取最早出现者）：
      A) 第一处出现真实公式的行 —— 对含公式的业务表最可靠
      B) 第一行"非空单元格中数值/日期占比 >= 25%"的行
    大标题行 / 全空行一律算表头区。

    注意 B 里必须把**日期/时间单元格也算作数据**：考勤表、入职日期这类列
    一整行只有一个金额是数字、其余全是日期，若只数数字就会把整个数据体误判成表头。
    """
    scan = max(1, int(scan))
    limit = min(scan + 4, sd.nrows)  # 允许略微超出扫描窗口找信号
    cand_formula = None
    cand_numeric = None
    for r in range(limit):
        row = sd.filled[r]
        vals = [row[c] for c in range(sd.ncols)]
        nonblank = [v for v in vals if not is_blank(v)]
        if not nonblank:
            continue
        if is_big_title_row(vals, sd.ncols):
            continue
        if cand_formula is None and any(sd.fml(r, c) for c in range(sd.ncols)):
            cand_formula = r
        if cand_numeric is None:
            num = sum(1 for v in nonblank if to_float(v)[1])
            nd = sum(1 for v in nonblank if is_datelike(v))
            other = len(nonblank) - num - nd
            # 日期只有**和别的内容混排**时才算数据信号：
            #   一整行全是日期 → 多半是"1日/2日/3日"那种日期表头行（考勤表常见）
            #   一行里既有工号又有日期又有金额 → 是数据行
            score = num + (nd if (num or other) else 0)
            if score / float(len(nonblank)) >= 0.25:
                cand_numeric = r
    cands = [c for c in (cand_formula, cand_numeric) if c is not None and c >= 1]
    if cands:
        return min(cands)
    # 兜底：最后一个非空的表头候选行 + 1
    last = 0
    for r in range(min(scan, sd.nrows)):
        if any(not is_blank(v) for v in sd.filled[r][:sd.ncols]):
            last = r
    return min(last + 1, sd.nrows)


def build_col_names(sd, header_rows):
    """
    多行表头 → 每列一个"拼接列名"。
    跳过大标题行与全空行；同列上下重复的片段只保留一次。
    """
    used = []
    for r in range(min(header_rows, sd.nrows)):
        vals = sd.filled[r][:sd.ncols]
        if all(is_blank(v) for v in vals):
            continue
        if is_big_title_row(vals, sd.ncols):
            continue
        used.append(r)
    names = []
    for c in range(sd.ncols):
        parts = []
        for r in used:
            s = cell_str(sd.filled[r][c]).replace("\n", " ").strip()
            if s and (not parts or parts[-1] != s):
                parts.append(s)
        names.append(" / ".join(parts) if parts else "(第%d列无表头)" % (c + 1))
    return names, used


def profile_column(sd, col, data_rows):
    """列画像：填充率 + 类型判定（pd.to_numeric 成功率）+ 样例值。"""
    raw = [sd.val(r, col) for r in data_rows]
    nonblank = [v for v in raw if not is_blank(v)]
    total = len(raw)
    fill_rate = (len(nonblank) / float(total)) if total else 0.0
    if not nonblank:
        return {"fill_rate": fill_rate, "type": "空列", "numeric_ratio": 0.0,
                "sample": "", "nonblank": 0, "total": total}
    # 日期/时间列单独判：pd.to_numeric 会把它们 coerce 成 NaN，于是整列被误报成"文本"
    ndate = sum(1 for v in nonblank if is_datelike(v))
    date_ratio = ndate / float(len(nonblank))

    ser = pd.Series(nonblank, dtype=object)
    try:
        conv = pd.to_numeric(ser, errors="coerce")
        ok = int(conv.notna().sum())
    except (TypeError, ValueError):     # 混了不可比较的对象时 pandas 也会直接抛
        ok = sum(1 for v in nonblank if to_float(v)[1])
    ratio = ok / float(len(nonblank))
    if date_ratio >= 0.95:
        typ = "日期/时间"
    elif date_ratio >= 0.5:
        typ = "日期/时间为主(%d%%)⚠️含非日期" % int(round(date_ratio * 100))
    elif ratio >= 0.95:
        typ = "数值"
    elif ratio >= 0.5:
        typ = "数值为主(%d%%)⚠️含非数值" % int(round(ratio * 100))
    elif ratio > 0:
        typ = "文本(混入%d%%数值)" % int(round(ratio * 100))
    else:
        typ = "文本"
    sample = " / ".join(cell_str(v) for v in nonblank[:3])
    return {"fill_rate": fill_rate, "type": typ, "numeric_ratio": ratio,
            "sample": sample[:120], "nonblank": len(nonblank), "total": total}


# ===========================================================================
# 四、tomd —— Step 1：Excel → 结构化 MD
# ===========================================================================

def sheet_to_md(sd, header_rows, preview_rows, summary_words):
    """把一个 Sheet 渲染成结构化 MD，返回 (md 文本, 概要 dict)。"""
    ncols = sd.ncols
    col_names, header_used = build_col_names(sd, header_rows)

    all_data_rows = list(range(header_rows, sd.nrows))
    body_rows, summary_rows, blank_rows = [], [], []
    for r in all_data_rows:
        vals = sd.values[r][:ncols]
        if all(is_blank(v) for v in vals):
            blank_rows.append(r)
        elif row_is_summary(vals, summary_words):
            summary_rows.append(r)
        else:
            body_rows.append(r)

    profiles = [profile_column(sd, c, body_rows) for c in range(ncols)]

    # 真实公式按列归集：数据体公式与汇总行公式分开——汇总行的 SUM 是另一类断言，
    # 混在一起会让"代表公式"指向 SUM，误导下一步的公式链还原。
    col_formula, col_formula_sum = [], []
    for c in range(ncols):
        col_formula.append([(r, sd.fml(r, c)) for r in body_rows if sd.fml(r, c)])
        col_formula_sum.append([(r, sd.fml(r, c)) for r in summary_rows if sd.fml(r, c)])

    merged_header = [m for m in sd.merged if m[1] - 1 < header_rows]
    merged_body = [m for m in sd.merged if m[1] - 1 >= header_rows]

    L = []
    L.append("# %s" % sd.sheet)
    L.append("")
    L.append("- 来源文件：`%s`" % sd.path)
    L.append("- 原始尺寸：%d 行 × %d 列" % (sd.nrows, ncols))
    L.append("- 推断表头行数：**%d**（数据体从 Excel 第 **%d** 行开始）" % (header_rows, header_rows + 1))
    L.append("- 数据体行数：**%d**（另有疑似汇总/小计行 %d 行、空行 %d 行，已从画像统计中排除）"
             % (len(body_rows), len(summary_rows), len(blank_rows)))
    L.append("- 合并单元格：%d 处（表头区 %d 处 / 数据区 %d 处）"
             % (len(sd.merged), len(merged_header), len(merged_body)))
    L.append("- 含真实公式的列：%d 列 / 公式单元格 %d 个"
             % (sum(1 for f in col_formula if f), sum(len(f) for f in col_formula)))
    L.append("")
    L.append("> 写 spec.json 时：`\"header_rows\": %d`。**表头行数数错一行全废，请先自己确认一眼。**"
             % header_rows)
    L.append("")

    # 一、表头区逐行解析
    L.append("## 一、表头区逐行解析")
    L.append("")
    rows = []
    for r in range(min(header_rows, sd.nrows)):
        vals = sd.filled[r][:ncols]
        if all(is_blank(v) for v in vals):
            kind = "空行（忽略）"
        elif is_big_title_row(vals, ncols):
            kind = "大标题行（跨表单值，已跳过）"
        elif r in header_used:
            kind = "表头行（参与拼接列名）"
        else:
            kind = "其它"
        parts = [cell_str(v) for v in vals if not is_blank(v)]
        if kind.startswith("大标题行"):  # 合并铺开后会重复 N 次，折叠展示
            content = "%s（跨 %d 列）" % (parts[0], len(parts))
        else:
            content = " | ".join(parts)
        rows.append([r + 1, kind, content[:400]])
    L.append(md_table(["Excel 行号", "类型", "内容（合并单元格已铺开）"], rows) if rows
             else "_（无表头行）_")
    L.append("")

    # 二、列字段定义表
    L.append("## 二、列字段定义表")
    L.append("")
    rows = []
    for c in range(ncols):
        p = profiles[c]
        f = col_formula[c]
        fml = ""
        if f:
            fml = "`%s`（%d 行）" % (f[0][1][:90], len(f))
        rows.append([c, get_column_letter(c + 1), col_names[c], p["sample"],
                     "%d%%" % int(round(p["fill_rate"] * 100)), p["type"], fml])
    L.append(md_table(["列号(0基)", "列标", "拼接列名", "样例值", "填充率", "类型", "真实公式"], rows))
    L.append("")
    L.append("> spec.json 的列引用可写「拼接列名（支持包含匹配）」或「列号(0基)」。"
             "列名命中多列时脚本会直接报错，那时改用列号。")
    L.append("")

    # 三、真实公式清单
    L.append("## 三、真实公式清单")
    L.append("")
    has_f = [c for c in range(ncols) if col_formula[c]]
    if not has_f:
        L.append("_本 Sheet 未抽到任何真实公式（可能是纯录入表，或公式已被转成静态值）。_")
    else:
        rows = []
        for c in has_f:
            f = col_formula[c]
            samples = "；".join("%s行:%s" % (r + 1, fm[:60]) for r, fm in f[1:3])
            rows.append([get_column_letter(c + 1), col_names[c], "`%s`" % f[0][1][:120],
                         len(f), samples])
        L.append(md_table(["列标", "拼接列名", "代表公式（首个）", "出现行数", "其它样本"], rows))
        L.append("")
        L.append("> 真实公式与你从列名推断的逻辑不一致时，**以真实公式为准**。")
    has_fs = [c for c in range(ncols) if col_formula_sum[c]]
    if has_fs:
        L.append("")
        L.append("**汇总行公式（原作者留下的断言，是白送的验证锚点）：**")
        L.append("")
        L.append(md_table(["列标", "拼接列名", "Excel行号", "公式"],
                          [[get_column_letter(c + 1), col_names[c], col_formula_sum[c][0][0] + 1,
                            "`%s`" % col_formula_sum[c][0][1][:120]] for c in has_fs]))
    L.append("")

    # 四、数据预览
    L.append("## 四、数据预览（数据体前 %d 行）" % preview_rows)
    L.append("")
    prev = body_rows[:preview_rows]
    if prev:
        headers = ["Excel行"] + ["%s(%d)" % (get_column_letter(c + 1), c) for c in range(ncols)]
        rows = [[r + 1] + [cell_str(sd.val(r, c)) for c in range(ncols)] for r in prev]
        L.append(md_table(headers, rows))
        L.append("")
        L.append("列号对应的拼接列名见上一节。")
    else:
        L.append("_（无数据体行）_")
    L.append("")

    # 五、汇总行
    if summary_rows:
        L.append("## 五、疑似汇总/小计行（验证时用 skip_when.label_in 排除，否则金额翻倍）")
        L.append("")
        rows = []
        for r in summary_rows[:20]:
            content = " | ".join(cell_str(v) for v in sd.values[r][:ncols] if not is_blank(v))
            rows.append([r + 1, content[:300]])
        L.append(md_table(["Excel 行号", "内容"], rows))
        if len(summary_rows) > 20:
            L.append("")
            L.append("_……共 %d 行，仅列前 20 行。_" % len(summary_rows))
        L.append("")

    # 六、告警
    warns = []
    for c in range(ncols):
        p, f = profiles[c], col_formula[c]
        if f and p["nonblank"] == 0:
            warns.append("- **列 %s「%s」有公式但取值全空**：这个 xlsx 没有缓存值（多半是程序生成的）。"
                         "用 Excel/LibreOffice 另存一次，或执行：\n"
                         "  `soffice --headless --convert-to xlsx --outdir <目录> '%s'` 重算后再跑一遍。"
                         % (get_column_letter(c + 1), col_names[c], sd.path))
        elif p["type"].startswith("日期/时间"):
            warns.append("- 列 %s「%s」是%s —— 日期不是金额，**按 0.0 参与运算**，"
                         "只能当维度/键用，别写进 checks 的算术表达式。"
                         % (get_column_letter(c + 1), col_names[c], p["type"]))
        elif "含非数值" in p["type"]:
            warns.append("- 列 %s「%s」%s —— 多半混了 `/`、`-`、备注文字，按 0 参与运算，"
                         "但业务上要与真正的 0 区分。"
                         % (get_column_letter(c + 1), col_names[c], p["type"]))
        elif p["type"] == "空列":
            warns.append("- 列 %s「%s」整列为空。" % (get_column_letter(c + 1), col_names[c]))
    if merged_body:
        warns.append("- 数据区存在 %d 处合并单元格（脚本**不会**把数据区的合并值铺开，以免金额被复制翻倍）。"
                     "若维度列因合并而出现空值，可在 spec.json 里用 `\"fill_merged\": [\"维度变量名\"]` 显式铺开。"
                     % len(merged_body))
    empty_all = [c for c in range(ncols) if profiles[c]["nonblank"] == 0 and not col_formula[c]]
    if len(empty_all) > ncols * 0.3 and ncols:
        warns.append("- 空列偏多（%d/%d），确认 header_rows 是否推断正确。" % (len(empty_all), ncols))
    L.append("## 六、告警与注意事项")
    L.append("")
    L.append("\n".join(warns) if warns else "_无。_")
    L.append("")

    summary = {
        "sheet": sd.sheet, "nrows": sd.nrows, "ncols": ncols,
        "header_rows": header_rows, "body_rows": len(body_rows),
        "summary_rows": len(summary_rows), "formula_cols": sum(1 for f in col_formula if f),
        "warns": len(warns),
    }
    return "\n".join(L), summary


def cmd_tomd(args):
    require_deps()
    outdir = ensure_dir(os.path.abspath(args.output))
    index = []
    total_sheets = 0

    if args.header_rows is not None and args.header_rows < 0:
        raise UserError("--header-rows 不能是负数：%d" % args.header_rows)
    if args.preview_rows < 0:
        raise UserError("--preview-rows 不能是负数：%d" % args.preview_rows)

    used_paths = {}   # 输出 md 全路径 → 已用次数（Sheet 名 safe_name 后可能撞车）
    for path in args.files:
        path = os.path.abspath(path)
        # 先过扩展名/魔数闸门，别让 pandas 抛一句英文 ImportError 就完事
        check_workbook_path(path)
        stem = safe_name(os.path.splitext(os.path.basename(path))[0])
        book_dir = ensure_dir(os.path.join(outdir, stem))
        try:
            xl = pd.ExcelFile(path)
            sheets = list(xl.sheet_names)
            xl.close()
        except Exception as e:
            raise UserError("无法打开工作簿 %s\n    %s: %s\n"
                            "    若它其实是旧版 .xls 或被改过扩展名，请先用 Excel/WPS 另存为 .xlsx。"
                            % (path, type(e).__name__, e))

        want = sheets
        if args.sheets:
            miss = [s for s in args.sheets if s not in sheets]
            if miss:
                raise UserError("工作簿 %s 中不存在 Sheet：%s\n    可选：%s"
                                % (os.path.basename(path), "、".join(miss), "、".join(sheets)))
            want = [s for s in sheets if s in args.sheets]

        for sh in want:
            sd = load_sheet(path, sh, fill_merged_rows=None)
            if sd.nrows == 0 or sd.ncols == 0:
                sys.stderr.write("[跳过] %s / %s 为空表\n" % (stem, sh))
                continue
            # 注意：合并单元格只铺进 sd.filled（供拼表头/识别大标题行用），
            # sd.values 始终是未铺开的原值，因此数据区不会被复制翻倍。
            # 注意用 is not None：--header-rows 0 是合法值（无表头的裸数据表），
            # 写成 `if args.header_rows` 会让 0 被当成"没指定"而回退自动推断。
            hr = (args.header_rows if args.header_rows is not None
                  else detect_data_start(sd, args.header_scan))
            hr = max(0, min(int(hr), sd.nrows))
            md, info = sheet_to_md(sd, hr, args.preview_rows, args.summary_words)
            # Sheet 名经 safe_name 后可能撞车（"报表." / "报表" / " 报表 " → 同一个文件名），
            # 撞了就加后缀，绝不能让后一个 Sheet 静默覆盖前一个。
            fp = os.path.join(book_dir, safe_name(sh) + ".md")
            if fp in used_paths:
                used_paths[fp] += 1
                fp = os.path.join(book_dir, "%s~%d.md" % (safe_name(sh), used_paths[fp]))
                sys.stderr.write("[提示] Sheet「%s」的文件名与前面的 Sheet 冲突，已改写为 %s\n"
                                 % (sh, os.path.basename(fp)))
            used_paths.setdefault(fp, 1)
            write_text(fp, md)
            info["file"] = os.path.relpath(fp, outdir)
            info["book"] = os.path.basename(path)
            index.append(info)
            total_sheets += 1
            print("[tomd] %s / %s → %s" % (os.path.basename(path), sh, fp))

    # 索引
    L = ["# 00-索引：Excel 结构化转换结果", "",
         "生成工具：`excel_ai.py tomd` v%s" % VERSION, "",
         "共 %d 个工作簿、%d 个 Sheet。" % (len(args.files), total_sheets), ""]
    rows = []
    for it in index:
        rows.append([it["book"], "[%s](%s)" % (it["sheet"], it["file"].replace(os.sep, "/")),
                     "%d×%d" % (it["nrows"], it["ncols"]), it["header_rows"], it["body_rows"],
                     it["summary_rows"], it["formula_cols"], it["warns"]])
    L.append(md_table(["工作簿", "Sheet", "尺寸", "表头行数", "数据行数", "汇总行", "含公式列", "告警数"], rows)
             if rows else "_（无内容）_")
    L += ["", "## 下一步", "",
          "1. 逐份 MD 确认四件事：数据体起始行 / 哪些是输入列与结果列 / 哪些是校验列 / 类型告警",
          "2. 手写 Step 2 字段本体与 Step 3 公式链+血缘（认知工作，脚本不能代劳）",
          "3. 把理解写成 `spec.json`，跑 `excel_ai.py verify` —— **跳过验证直接分析 = 没读懂代码就改代码**",
          ""]
    idx = write_text(os.path.join(outdir, "00-索引.md"), "\n".join(L))
    print("[tomd] 索引 → %s" % idx)
    print("[tomd] 完成：%d 个 Sheet，输出目录 %s" % (total_sheets, outdir))
    return 0


# ===========================================================================
# 五、spec 装载与列绑定
# ===========================================================================

def load_spec(path):
    if not os.path.isfile(path):
        raise UserError("找不到 spec 文件：%s" % path)
    with open(path, "r", encoding="utf-8-sig") as f:
        raw = f.read()
    # 容忍 jsonc 风格的整行 // 注释（不动字符串内部的 //）
    lines = []
    for ln in raw.splitlines():
        if ln.lstrip().startswith("//"):
            continue
        lines.append(ln)
    try:
        spec = json.loads("\n".join(lines))
    except Exception as e:
        raise UserError("spec.json 解析失败：%s\n    %s\n"
                        "    提示：JSON 不支持行尾注释，请只用整行 // 注释或干脆去掉注释。" % (path, e))
    if not isinstance(spec, dict):
        raise UserError("spec.json 的顶层必须是一个对象 { ... }，实际是%s：%s\n"
                        "    最小骨架：{\"workbook\":\"/abs/表.xlsx\",\"sheet\":\"明细\",\"header_rows\":1,"
                        "\"keys\":{},\"fields\":{},\"checks\":[]}" % (_tname(spec), clip(json.dumps(spec, ensure_ascii=False), 120)))
    return spec


def resolve_column(ref, col_names, where=""):
    """
    列引用解析：支持列号(int) 或 列名(str)。
    列名先精确匹配（归一化后全等）、再包含匹配；**命中多列一律直接报错并列出候选**。

    纯数字的字符串（"1"、"01"、"2024"）**先当列名找**，找不到才退回按列号解释——
    月份列 "01"、年份列 "2024" 是中文表里的常客，此前一律被当成列号，
    会静默绑到完全不相干的列上，通过率还是 100%，是最危险的一类错。
    """
    n = len(col_names)
    if isinstance(ref, bool):
        raise UserError("列引用不能是布尔值%s" % (("（%s）" % where) if where else ""))
    if isinstance(ref, float) and float(ref).is_integer():
        ref = int(ref)
    if isinstance(ref, int):
        if not (0 <= ref < n):
            raise UserError("列号越界%s：%d（本表共 %d 列，合法范围 0..%d）"
                            % (("（%s）" % where) if where else "", ref, n, n - 1))
        return ref
    if not isinstance(ref, str):
        raise UserError("列引用必须是列号(int)或列名(str)%s，收到：%s"
                        % (("（%s）" % where) if where else "", clip(repr(ref), 60)))
    s = ref.strip()
    if s == "":
        raise UserError("列引用不能是空字符串%s" % (("（%s）" % where) if where else ""))

    target = norm_key(s)
    normed = [norm_key(c) for c in col_names]
    wh = ("（%s）" % where) if where else ""

    def _cands(idxs):
        return "；".join("列号%d=%s" % (i, clip(col_names[i], 60)) for i in idxs[:12])

    exact = [i for i, c in enumerate(normed) if c == target]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        raise UserError("列名「%s」精确命中 %d 列%s，无法确定用哪一列——请改用列号。\n    候选：%s"
                        % (clip(ref), len(exact), wh, _cands(exact)))

    is_digits = re.fullmatch(r"\d+", s) is not None
    if not exact:
        part = [i for i, c in enumerate(normed) if target and target in c]
        # 纯数字串做包含匹配太容易误伤（"1" 会命中 "第1季度"），只在非数字时才用
        if not is_digits:
            if len(part) == 1:
                return part[0]
            if len(part) > 1:
                raise UserError("列名「%s」包含匹配命中 %d 列%s，无法确定用哪一列——请改用列号。\n    候选：%s"
                                % (clip(ref), len(part), wh, _cands(part)))

    if is_digits:   # 列名里确实没有这个名字，才退回"字符串形式的列号"
        idx = int(s)
        if not (0 <= idx < n):
            raise UserError("列引用「%s」%s既不是本表的列名，按列号解释又越界（本表共 %d 列，"
                            "合法范围 0..%d）。\n    如果它本来就是列名，请检查 header_rows 是否数对了。"
                            % (s, wh, n, n - 1))
        sys.stderr.write("[提示] 列引用「%s」%s在列名里找不到，已按列号 %d 解释 → 「%s」。"
                         "若本意是列名，请核对 header_rows。\n"
                         % (s, wh, idx, clip(col_names[idx], 40)))
        return idx

    near = [i for i, c in enumerate(normed) if c and (c in target)]
    hint = ("\n    近似列：%s" % _cands(near)) if near else ""
    raise UserError("列名「%s」在本表中找不到%s。%s\n    全部列名：%s"
                    % (clip(ref), wh, hint,
                       "；".join("%d=%s" % (i, clip(c, 40)) for i, c in enumerate(col_names))))


RAW_SUFFIX = "__raw"


def validate_var_name(var, where):
    """
    变量名闸门（keys / dimensions / fields / derived 共用一个命名空间）。

    这三条都是"不拦就会静默算错"的：
      1. 不是合法标识符 —— `"A-B"` 写进 expr 会被解析成 `A - B`，用户以为在取那一列，
         实际在做减法，结果完全是另一回事，而且照样 100% 通过；`"A B"` 则永远取不到。
      2. 以 `__raw` 结尾 —— 每个变量都会自动生成一个 `<变量名>__raw` 原始文本版本，
         自己再叫 `A__raw` 就会和 `A` 的那份撞车，谁后写谁赢，取到的值看运气。
      3. 与白名单函数同名 —— `sum`/`round` 既是函数又是列，表达式里到底指哪个全靠猜。
    """
    if not isinstance(var, str) or not var.strip():
        raise UserError("spec 的 `%s` 里出现了空变量名" % where)
    if var.endswith(RAW_SUFFIX):
        raise UserError(
            "变量名 %s 不能以 `%s` 结尾（%s）。\n"
            "    脚本会给每个变量自动生成一份 `<变量名>%s`（未转数值的原始文本），"
            "自己再叫这个名字就会和它撞车，表达式里取到哪一个全看书写顺序。"
            % (var, RAW_SUFFIX, where, RAW_SUFFIX))
    if var in SafeEval.FUNCS:
        raise UserError(
            "变量名 %s 与内置函数重名（%s）。\n"
            "    表达式里 `%s(...)` 是调用函数、`%s` 是取列值，同名会让人（和下一个 AI）"
            "根本分不清写的是哪一个。请改名，例如 `%s_x`。\n"
            "    内置函数：%s"
            % (var, where, var, var, var, " ".join(sorted(SafeEval.FUNCS))))
    if _kw.iskeyword(var) or var in ("True", "False", "None"):
        raise UserError("变量名 %s 是 Python 关键字（%s），表达式里没法引用它，请改名。"
                        % (var, where))
    if not var.isidentifier():
        raise UserError(
            "变量名 %s 不是合法标识符（%s），表达式里引用不到它。\n"
            "    只能用字母/数字/下划线/汉字，且不能以数字开头；空格、`-`、`.`、`(`、`·` 都不行。\n"
            "    **这一条必须拦**：`\"A-B\"` 这种名字写进 expr 会被当成 `A - B` 做减法，"
            "算出来的是另一回事，报告却照样 100%% 通过。\n"
            "    列名里的这些字符不用管——变量名只是你给列起的代号，"
            "真实列名写在 keys/dimensions/fields 的**值**里。" % (var, where))
    return var


# spec.json 的合法顶层字段（与 references/spec-schema.md 的字段表一一对应）
SPEC_TOP_KEYS = {"workbook", "sheet", "header_rows", "keys", "dimensions", "fields",
                 "derived", "checks", "skip_when", "fill_merged", "cross_checks",
                 "lineage", "ontology", "analysis"}


class Binding(object):
    """一次 verify 的表绑定：Sheet 数据 + 列名 + 变量→列号。"""

    def __init__(self, sd, header_rows, col_names):
        self.sd = sd
        self.header_rows = header_rows
        self.col_names = col_names
        self.var2col = OrderedDict()   # 变量名 → 列号
        self.var_kind = {}             # 变量名 → key/dim/field
        self.header_guess = header_rows  # 脚本用同一套启发式推断的表头行数


def bind_sheet(spec, base_dir):
    """按 spec 打开主表并绑定 keys / dimensions / fields。"""
    check_keys(spec, SPEC_TOP_KEYS, "")
    wb_path = spec.get("workbook")
    if not wb_path:
        raise UserError("spec.json 缺少必填字段 workbook（Excel 绝对路径）")
    if not isinstance(wb_path, str):
        raise UserError("spec.json 的 workbook 必须是字符串路径，实际是%s" % _tname(wb_path))
    if not os.path.isabs(wb_path):
        wb_path = os.path.normpath(os.path.join(base_dir, wb_path))
    sheet = spec.get("sheet")
    if not sheet:
        raise UserError("spec.json 缺少必填字段 sheet（Sheet 名）")
    if not isinstance(sheet, str):
        raise UserError("spec.json 的 sheet 必须是字符串，实际是%s" % _tname(sheet))
    if "header_rows" not in spec:
        raise UserError("spec.json 缺少必填字段 header_rows（表头占几行，数错一行全废）")
    # 负数会让 range(hr, nrows) 从 -1 起步，把最后一行和表头行当成数据行重复计入
    hr = as_int(spec["header_rows"], "header_rows", minimum=0)

    sd = load_sheet(wb_path, sheet, fill_merged_rows=None)
    if sd.nrows == 0:
        raise UserError("Sheet「%s」为空表" % sheet)
    if hr >= sd.nrows:
        raise UserError("header_rows=%d 超过总行数 %d，请核对 tomd 输出的表头行数" % (hr, sd.nrows))
    col_names, _ = build_col_names(sd, hr)
    # 表头拼接完成后，把数据区的合并铺开撤销（load_sheet 已整表铺开，这里只用 values）
    b = Binding(sd, hr, col_names)
    # 用 tomd 的同一套启发式独立推断一遍表头行数，和 spec 写的对一对。
    # "header_rows 数错一行全废"是 SKILL.md 的头号坑，而且数错时**多半不报错**：
    # 少数一行 → 字段名那行被当成数据（全是文本→按 0 参与运算→0==0 算通过）；
    # 多数一行 → 第一行真数据被当表头吃掉，谁也不会发现。
    try:
        b.header_guess = detect_data_start(sd, max(8, hr + 4))
    except Exception:      # 启发式只是辅助，任何意外都不该影响主流程
        b.header_guess = hr

    for kind, key in (("key", "keys"), ("dim", "dimensions"), ("field", "fields")):
        for var, ref in as_dict(spec.get(key), key).items():
            validate_var_name(var, key)
            if var in b.var2col:
                raise UserError("变量名重复定义：%s（出现在多个段中）" % var)
            if var in RESERVED_DETAIL_COLS:
                raise UserError("变量名 %s 与验证明细 CSV 的保留列名冲突，请换一个名字。\n"
                                "    保留列名：%s（CSV 里出现同名列会静默丢数据，"
                                "后面的 output/analyze 拿到的就是串了的表）"
                                % (var, "、".join(sorted(RESERVED_DETAIL_COLS))))
            b.var2col[var] = resolve_column(ref, col_names, where="%s.%s" % (key, var))
            b.var_kind[var] = kind

    # 可选：显式铺开某些变量列在数据区的合并单元格（默认不铺，避免金额翻倍）
    for var in as_str_list(spec.get("fill_merged"), "fill_merged"):
        if var not in b.var2col:
            raise UserError("fill_merged 里的变量 %s 未在 keys/dimensions/fields 中定义" % var)
        c = b.var2col[var]
        for (min_col, min_row, max_col, max_row) in sd.merged:
            if not (min_col - 1 <= c <= max_col - 1) or min_row - 1 < hr:
                continue
            v = sd.values[min_row - 1][min_col - 1]
            if is_blank(v):
                continue
            for rr in range(min_row - 1, min(max_row, sd.nrows)):
                sd.values[rr][c] = v
    return b, wb_path


def validate_expr_vars(expr, where, allowed, hint=""):
    """
    表达式引用的变量必须都已定义。**在逐行求值之前就查**——1 万行的表跑到第一行
    才报错，用户先看到的是行号噪音而不是问题本身。
    """
    for nm in SAFE.names(expr, where):
        if nm in allowed or nm in SafeEval.FUNCS:
            continue
        main = sorted(x for x in allowed if not x.endswith("__raw"))
        raise UserError("`%s` 的表达式引用了未定义的变量：%s\n    %s\n"
                        "    当前可用变量：%s\n"
                        "    （另外每个列变量都有一个 `<变量名>__raw` 版本，取未转数值的原始文本）"
                        % (where, nm, hint or "请检查是不是拼错了，或忘了把它写进 fields/derived。",
                           "、".join(main) or "(无)"))


def validate_derived(derived, b):
    """
    derived 的静态依赖检查：循环依赖 / 声明顺序颠倒 / 未定义变量。

    derived 是**按声明顺序**求值的，不是方程组求解器：
      * `{"S": "S+1"}`  自引用 —— 永远算不出来
      * `{"P": "Q+1", "Q": "P+1"}` 互相引用 —— 循环依赖
      * `{"TOT": "SUB+1", "SUB": "A+B"}` 顺序颠倒 —— 换个顺序就能跑
    这三种此前都只会在第一行求值时报一句"未定义的变量 Q"，用户很容易以为是列没绑上，
    对着 fields 找半天。这里一次性说清是哪一种。
    """
    deps = OrderedDict()
    for dvar, dexpr in derived.items():
        deps[dvar] = [nm for nm in SAFE.names(dexpr, "derived.%s" % dvar) if nm in derived]

    # 1) 先找环（含自环）。**必须用迭代式 DFS**：递归版在几千个链式 derived 上
    #    会自己把 Python 栈打爆（RecursionError），那又是一条未捕获的 traceback。
    WHITE, GRAY, BLACK = 0, 1, 2
    color = {k: WHITE for k in derived}
    for root in derived:
        if color[root] != WHITE:
            continue
        color[root] = GRAY
        path = [root]                 # 当前 DFS 路径（用来还原环）
        stack = [[root, 0]]           # [节点, 下一个待访问的依赖下标]
        while stack:
            u, i = stack[-1]
            if i < len(deps[u]):
                stack[-1][1] = i + 1
                v = deps[u][i]
                if color[v] == GRAY:                  # 回边 → 成环
                    cyc = path[path.index(v):] + [v]
                    if len(cyc) == 2 and cyc[0] == cyc[1]:
                        raise UserError(
                            "derived.%s 自己引用了自己：`%s`\n"
                            "    derived 是按声明顺序**逐个求值**的中间量，不是方程求解器，"
                            "自引用永远算不出来。\n"
                            "    如果本意是「用 Excel 里那一列的值」，请把它写进 `fields` 绑到真实列上。"
                            % (v, derived[v]))
                    raise UserError(
                        "derived 存在循环依赖：%s\n"
                        "    涉及的表达式：%s\n"
                        "    derived 按声明顺序逐个求值，互相引用的一组变量无论怎么排都算不出来，"
                        "必须打断环：把其中一个改成引用 `fields` 里的真实列。"
                        % (" → ".join(cyc),
                           "；".join("%s=`%s`" % (x, derived[x]) for x in cyc[:-1])))
                if color[v] == WHITE:
                    color[v] = GRAY
                    path.append(v)
                    stack.append([v, 0])
            else:
                color[u] = BLACK
                path.pop()
                stack.pop()

    # 2) 无环，再看声明顺序（按声明顺序求值，引用后面声明的中间量必然取不到）
    order = list(derived)
    pos = {k: i for i, k in enumerate(order)}
    for dvar in order:
        for v in deps[dvar]:
            if pos[v] > pos[dvar]:
                raise UserError(
                    "derived.%s 引用了在它**之后**声明的中间量 %s：`%s`\n"
                    "    derived 是按声明顺序求值的，算到 %s 时 %s 还不存在。\n"
                    "    把 %s 挪到 %s 前面即可（JSON 对象的书写顺序就是求值顺序）。"
                    % (dvar, v, derived[dvar], dvar, v, v, dvar))

    # 3) 剩下的未定义变量
    allowed = set(b.var2col) | {v + "__raw" for v in b.var2col}
    for dvar, dexpr in derived.items():
        validate_expr_vars(dexpr, "derived.%s" % dvar, allowed,
                           hint="derived 只能引用 keys/dimensions/fields 里的列变量，"
                                "以及**在它之前**声明的其它 derived。")
        allowed.add(dvar)
    return allowed


def row_env(b, r):
    """
    构造一行的求值环境：
      * field 变量 → float（非数值按 0.0）
      * key / dim 变量 → 原始字符串
      * 每个变量另有 `<变量名>__raw` → 原始字符串（供文本判断使用）
    """
    env = {}
    nonnum = []
    for var, c in b.var2col.items():
        raw = b.sd.val(r, c)
        s = cell_str(raw)
        env[var + "__raw"] = s
        if b.var_kind.get(var) == "field":
            f, ok = to_float(raw)
            env[var] = f
            if not ok and s != "":
                nonnum.append((var, s))
        else:
            env[var] = s
    return env, nonnum


# ===========================================================================
# 六、verify —— Step 4：全量数据验证（核心）
# ===========================================================================

def cmd_verify(args):
    require_deps()
    spec_path = os.path.abspath(args.spec)
    spec = load_spec(spec_path)
    base_dir = os.path.dirname(spec_path)
    outdir = ensure_dir(os.path.abspath(args.output))

    b, wb_path = bind_sheet(spec, base_dir)
    sd = b.sd
    derived = OrderedDict((k, v) for k, v in as_dict(spec.get("derived"), "derived").items())
    for d, dexpr in derived.items():
        validate_var_name(d, "derived")
        if d in b.var2col:
            raise UserError("derived 变量 %s 与 keys/dimensions/fields 中的变量重名" % d)
        if d in RESERVED_DETAIL_COLS:
            raise UserError("derived 变量名 %s 与验证明细 CSV 的保留列名冲突，请换一个名字。" % d)
        if not isinstance(dexpr, str):
            raise UserError("derived.%s 的值必须是表达式字符串，实际是%s" % (d, _tname(dexpr)))
    # 循环依赖 / 声明顺序 / 未定义变量，全部在开跑之前查掉
    allowed_vars = validate_derived(derived, b)
    checks = as_list(spec.get("checks"), "checks")
    if not checks:
        raise UserError("spec.json 的 checks 为空 —— 没有校验项就不叫验证。\n"
                        "    每项形如 {\"name\":\"应发工资G\",\"target\":\"G_x\","
                        "\"expr\":\"round(A+B-C+D+E-F,2)\",\"tolerance\":0.01}")
    loose_tol = []      # 容差被放大到远超建议值的校验项（会把真差异掩盖成"通过"）
    for i, ck in enumerate(checks):
        as_item_dict(ck, "checks[%d]" % i)
        check_keys(ck, {"name", "target", "expr", "tolerance"}, "checks[%d]" % i)
        for f in ("name", "target", "expr"):
            if not ck.get(f):
                raise UserError("checks[%d] 缺少字段 %s（target 必须是 Excel 里已有的结果列变量）" % (i, f))
            if not isinstance(ck[f], str):
                raise UserError("checks[%d].%s 必须是字符串，实际是%s" % (i, f, _tname(ck[f])))
        tgt = ck["target"]
        if tgt not in b.var2col:
            raise UserError("checks[%d]「%s」的 target=%s 未在 fields/keys/dimensions 中定义"
                            % (i, ck["name"], tgt))
        # target 是"Excel 里已有的数值结果列"，必须来自 fields：
        # 指向 keys/dimensions 时里面装的是字符串，直接 float() 会炸出 traceback。
        if b.var_kind.get(tgt) != "field":
            raise UserError(
                "checks[%d]「%s」的 target=%s 是%s变量，不能当校验目标。\n"
                "    target 必须是 `fields` 里的数值结果列（Excel 已经算好的那一列），"
                "keys/dimensions 装的是文本，没法和 AI 算出来的数字对撞。"
                % (i, ck["name"], tgt, {"key": "键(keys)", "dim": "维度(dimensions)"}.get(b.var_kind.get(tgt), "")))
        tol_i = as_float(ck.get("tolerance"), "checks[%d].tolerance" % i, default=0.01, minimum=0.0)
        if tol_i > LOOSE_TOL:
            loose_tol.append((ck["name"], tol_i))
        SAFE.parse(ck["expr"], "checks[%d].%s" % (i, ck["name"]))
        validate_expr_vars(ck["expr"], "checks[%d].%s" % (i, ck["name"]), allowed_vars,
                           hint="checks 的表达式只能引用 keys/dimensions/fields 里的列变量与 "
                                "derived 中间量；引用 derived 之前请先在 `derived` 段里定义它。")
    names = [ck["name"] for ck in checks]
    if len(set(names)) != len(names):
        raise UserError("checks 里存在同名校验项：%s" % names)
    for nm in names:
        if nm in b.var2col or nm in derived or nm in RESERVED_DETAIL_COLS:
            raise UserError("校验项名「%s」与变量名/保留列名重复，会让明细 CSV 出现同名列，请改名。" % nm)
        if nm.endswith(RAW_SUFFIX):
            raise UserError("校验项名「%s」不能以 `%s` 结尾：analyze 里每个变量都有一份"
                            "自动生成的 `<变量名>%s`（原始文本），同名会让分析报告把文本"
                            "当成金额去求和。" % (nm, RAW_SUFFIX, RAW_SUFFIX))

    skip_when = check_keys(as_dict(spec.get("skip_when"), "skip_when"),
                           {"empty", "label_in"}, "skip_when")
    skip_empty = as_str_list(skip_when.get("empty"), "skip_when.empty")
    skip_labels = as_str_list(skip_when.get("label_in"), "skip_when.label_in")
    for v in skip_empty:
        if v not in b.var2col:
            raise UserError("skip_when.empty 里的 %s 未在 keys/dimensions/fields 中定义" % v)

    # ---- 逐行回算 ----------------------------------------------------
    key_vars = [v for v in b.var2col if b.var_kind[v] == "key"]
    dim_vars = [v for v in b.var2col if b.var_kind[v] == "dim"]
    field_vars = [v for v in b.var2col if b.var_kind[v] == "field"]

    detail_headers = (["Excel行号"] + key_vars + dim_vars + field_vars + list(derived) +
                      sum([["%s%sAI" % (n, SEP), "%s%sExcel" % (n, SEP), "%s%s差异" % (n, SEP)]
                           for n in names], []))
    # 明细 CSV 出现同名列 = 交付物直接损坏（DictReader 会静默丢列，output/analyze
    # 拿到的是串了的表）。变量名和校验项名各自不重复还不够：一个叫 `X·AI` 的字段变量
    # 和一个叫 `X` 的校验项拼出来就是同一个列名。这里做最后一道闸。
    dup_h = sorted({h for h in detail_headers if detail_headers.count(h) > 1})
    if dup_h:
        raise UserError(
            "明细 CSV 会出现同名列：%s\n"
            "    列名由「变量名」和「校验项名%sAI / %sExcel / %s差异」共同拼成，"
            "两边撞在一起会让 CSV 静默丢列，后面 output/analyze 拿到的全是串了的数据。\n"
            "    请给其中一个改名（例如给字段变量加后缀 `_x`）。"
            % ("、".join(dup_h), SEP, SEP, SEP))
    detail_rows = []
    mismatch_rows = []
    stats = OrderedDict((n, {"cmp": 0, "ok": 0, "bad": 0, "maxdiff": 0.0,
                             "sum_ai": 0.0, "sum_ex": 0.0, "tiny": 0, "loose": 0})
                        for n in names)
    skipped_empty = skipped_label = 0
    nonnum_hits = {}
    n_rows = 0
    row_envs = []
    suspect_summary = []   # 没被 skip_when 排掉、但长得像汇总/小计的行（金额会翻倍）
    n_suspect = 0
    headerish = []         # 每个数值字段都是"文本"的行 —— header_rows 少数了一行的铁证
    n_headerish = 0

    for r in range(b.header_rows, sd.nrows):
        vals = sd.values[r][:sd.ncols]
        if all(is_blank(v) for v in vals):
            continue
        if skip_labels and row_is_summary(vals, skip_labels):
            skipped_label += 1
            continue
        env, nonnum = row_env(b, r)
        if skip_empty and any(env[v + "__raw"] == "" for v in skip_empty):
            skipped_empty += 1
            continue
        # 没配 skip_when、或配漏了词的情况下，汇总行会混进数据体让金额直接翻倍
        # （SKILL.md 踩坑清单第 6 条）。这里不擅自跳过，但必须点名。
        if row_is_summary(vals, DEFAULT_SUMMARY_WORDS):
            n_suspect += 1
            if len(suspect_summary) < 20:
                suspect_summary.append([r + 1, clip(" | ".join(
                    cell_str(v) for v in vals if not is_blank(v)), 120)])
        for var, s in nonnum:
            nonnum_hits.setdefault(var, []).append(s)

        # header_rows 少数一行时，字段名那一行会被当成数据行：每个数值字段里装的都是
        # 列名文本 → 按 0.0 参与运算 → AI 算出 0、Excel 现值也是 0 → **0==0 判为通过**。
        # 于是报告里凭空多一行、通过率照样 100%，是最危险的一类静默错误。
        # 判据：字段变量 ≥2 且全部是"非空、非数值、也不是 / - 这类占位符"的实义文本。
        if len(field_vars) >= 2 and len(nonnum) == len(field_vars) \
                and all(s not in PLACEHOLDER_TOKENS for _v, s in nonnum):
            n_headerish += 1
            if len(headerish) < 20:
                headerish.append([r + 1, clip(" | ".join(
                    "%s=%s" % (v, s) for v, s in nonnum), 120)])

        # 求值出错时必须报出是哪一行哪个人，否则 1 万行表里根本没法定位
        rowtag = "Excel第%d行%s" % (r + 1, ("／" + "／".join(
            "%s=%s" % (v, clip(env[v + "__raw"], 20)) for v in key_vars[:2])) if key_vars else "")

        # 先按声明顺序算 derived，再算 checks
        for dv, dexpr in derived.items():
            env[dv] = eval_num(dexpr, env, "derived.%s，%s" % (dv, rowtag))

        drow = [r + 1] + [env[v + "__raw"] for v in key_vars + dim_vars]
        drow += [round(env[v], 6) for v in field_vars]
        drow += [round(env[v], 6) for v in derived]

        for ck in checks:
            n = ck["name"]
            tol = as_float(ck.get("tolerance"), "checks.%s.tolerance" % n, default=0.01, minimum=0.0)
            ai = eval_num(ck["expr"], env, "checks.%s，%s" % (n, rowtag))
            ex = float(env[ck["target"]])   # target 已强制为 field，这里必是 float
            diff = ai - ex
            st = stats[n]
            st["cmp"] += 1
            st["sum_ai"] += ai
            st["sum_ex"] += ex
            st["maxdiff"] = max(st["maxdiff"], abs(diff))
            if abs(diff) > tol:
                st["bad"] += 1
                mismatch_rows.append(
                    [r + 1] + [env[v + "__raw"] for v in key_vars + dim_vars] +
                    [n, round(ai, 4), round(ex, 4), round(diff, 4), tol, ck["expr"]])
            else:
                st["ok"] += 1
                if diff != 0:
                    st["tiny"] += 1
                # 差异其实超过了 0.01（金额的合理容差），只是被放大的 tolerance 兜住了。
                # 这种"通过"必须点名，否则就是拿容差把真差异洗成 100%。
                if abs(diff) > STRICT_TOL:
                    st["loose"] += 1
            drow += [round(ai, 6), round(ex, 6), round(diff, 6)]

        detail_rows.append(drow)
        row_envs.append(env)
        n_rows += 1

    if n_rows == 0:
        raise UserError("没有任何数据行参与比对。请检查 header_rows(%d) 与 skip_when 是否过严。"
                        % b.header_rows)

    # ---- 跨表一致性 ---------------------------------------------------
    cross_results = []
    for i, cc in enumerate(as_list(spec.get("cross_checks"), "cross_checks")):
        cross_results.append(run_cross_check(as_item_dict(cc, "cross_checks[%d]" % i),
                                             i, b, base_dir, key_vars, dim_vars))

    # ---- 公式列全空告警 ------------------------------------------------
    warns = []
    empty_formula_cols = []      # 有公式但取值全空 → 该 xlsx 没有缓存值，验证结果不可信
    body = list(range(b.header_rows, sd.nrows))
    for var, c in b.var2col.items():
        nonblank = sum(1 for r in body if not is_blank(sd.val(r, c)))
        has_f = any(sd.fml(r, c) for r in body)
        if nonblank == 0 and has_f:
            empty_formula_cols.append("%s（列 %s「%s」）" % (var, get_column_letter(c + 1), b.col_names[c]))
        elif nonblank == 0:
            warns.append("变量 %s（列 %s「%s」）在数据体内整列为空，确认列引用是否指错。"
                         % (var, get_column_letter(c + 1), clip(b.col_names[c], 60)))
    # 键/维度列若在数据区被合并，只有合并区首行有值，其余全空 → 分组统计会整片掉进"(空)"
    filled_vars = set(as_str_list(spec.get("fill_merged"), "fill_merged"))
    for var, c in b.var2col.items():
        if b.var_kind.get(var) not in ("key", "dim") or var in filled_vars:
            continue
        n_m = sum(1 for (mc0, mr0, mc1, mr1) in sd.merged
                  if mc0 - 1 <= c <= mc1 - 1 and mr0 - 1 >= b.header_rows and mr1 > mr0)
        if n_m:
            warns.append("变量 %s（列 %s「%s」）在数据区有 %d 处纵向合并单元格，"
                         "合并区里只有第一行有值、其余是空——分组/连接会整片掉进「(空)」。"
                         "如需铺开，在 spec 里加 `\"fill_merged\": [\"%s\"]`。"
                         % (var, get_column_letter(c + 1), clip(b.col_names[c], 40), n_m, var))
    for var, ss in sorted(nonnum_hits.items()):
        uniq = sorted(set(ss))[:6]
        warns.append("字段 %s 有 %d 个非数值单元格（已按 0.0 参与运算）：%s —— "
                     "`/`、`-` 在中文表里多半表示\"不参加/无此项\"，业务上要与真正的 0 区分。"
                     % (var, len(ss), "、".join("`%s`" % clip(u, 30) for u in uniq)))
    if n_suspect:
        warns.append("**有 %d 行长得像汇总/小计行却仍在参与比对**（含 %s 等词），金额会被重复计入、"
                     "汇总额直接翻倍。请在 spec 里补 `\"skip_when\": {\"label_in\": [\"合计\",\"小计\",\"总计\"]}` "
                     "后重跑。明细见报告「疑似未排除的汇总行」一节。"
                     % (n_suspect, "、".join(DEFAULT_SUMMARY_WORDS[:4])))
    if n_headerish:
        warns.append("**疑似表头行混入数据体：%d 行**——这些行的每个数值字段装的都是文本，"
                     "会按 0.0 参与运算，于是 AI 算 0、Excel 也读成 0，**0==0 被判为通过**，"
                     "凭空推高通过率。多半是 `header_rows` 少数了一行。明细见报告「疑似表头行」一节。"
                     % n_headerish)
    if b.header_guess != b.header_rows:
        warns.append("**header_rows 与脚本推断不一致**：spec 写的是 %d，脚本按 tomd 的同一套启发式"
                     "推断数据体从第 %d 行（0基）开始。SKILL.md 的头号坑就是「表头行数数错一行全废」，"
                     "而数错时通常**不会报错**：少数一行→字段名那行被当成数据（0==0 算通过）；"
                     "多数一行→第一行真数据被当表头吃掉，谁也不会发现。请对照 tomd 的 MD 再确认一眼。"
                     % (b.header_rows, b.header_guess))
    for nm_l, tol_l in loose_tol:
        warns.append("校验项「%s」的容差被设成 %s，远大于建议值（金额 0.01 / 比率 0.0001）。"
                     "SKILL.md：**不要为了让报告好看去放大容差**。" % (nm_l, tol_l))
    n_loose = sum(s["loose"] for s in stats.values())
    if n_loose:
        warns.append("**有 %d 次比对的绝对差异超过 %s，仅因为容差被放大才判为「通过」**（逐项：%s）。"
                     "这些不是真的对上了，请把容差调回 0.01 重跑，再看剩下的差异到底是谁的错。"
                     % (n_loose, STRICT_TOL,
                        "；".join("%s %d 次" % (n, stats[n]["loose"]) for n in names
                                  if stats[n]["loose"])))

    # ---- 落盘 ---------------------------------------------------------
    detail_csv = write_csv(os.path.join(outdir, "verification_detail.csv"), detail_headers, detail_rows)
    mm_headers = (["Excel行号"] + key_vars + dim_vars +
                  ["校验项", "AI值", "Excel值", "差异", "容差", "AI 表达式"])
    mm_csv = write_csv(os.path.join(outdir, "mismatches.csv"), mm_headers, mismatch_rows)

    total_cmp = sum(s["cmp"] for s in stats.values())
    total_bad = sum(s["bad"] for s in stats.values())
    pass_rate = (total_cmp - total_bad) / float(total_cmp) * 100 if total_cmp else 0.0

    report = render_verify_report(b, wb_path, names, checks, derived, stats, mismatch_rows,
                                  mm_headers, cross_results, warns, n_rows, skipped_empty,
                                  skipped_label, total_cmp, total_bad, pass_rate,
                                  detail_csv, mm_csv, empty_formula_cols,
                                  suspect_summary, n_suspect,
                                  headerish=headerish, n_headerish=n_headerish,
                                  n_loose=n_loose)
    rp = write_text(os.path.join(outdir, "验证报告.md"), report)

    if n_headerish:
        sys.stderr.write(
            "\n[警告] 有 %d 行的每个数值字段装的都是文本，极可能是**表头行被当成了数据行**"
            "（header_rows 少数了一行）：\n    %s\n"
            "    这类行 AI 算 0、Excel 也读成 0，会被判为「通过」，凭空推高通过率。\n"
            "    修复：把 spec 的 header_rows 从 %d 改成 %d 后重跑。\n\n"
            % (n_headerish, "；".join("第%s行 %s" % (r[0], r[1]) for r in headerish[:5]),
               b.header_rows, b.header_rows + 1))
    if b.header_guess != b.header_rows:
        sys.stderr.write(
            "[警告] spec 的 header_rows=%d，脚本自己推断的是 %d —— 表头行数数错一行全废，"
            "请对照 tomd 的 MD 确认一眼。\n" % (b.header_rows, b.header_guess))
    if n_loose:
        sys.stderr.write(
            "[警告] 有 %d 次比对的差异其实超过 %s，只是被放大的容差兜住才算「通过」。"
            "把容差调回 0.01 重跑再下结论。\n" % (n_loose, STRICT_TOL))
    if n_suspect:
        sys.stderr.write(
            "\n[警告] 有 %d 行长得像汇总/小计行却仍被当成数据行参与了比对，金额会翻倍：\n"
            "    %s\n"
            "    修复：在 spec.json 里加 \"skip_when\": {\"label_in\": [\"合计\", \"小计\", \"总计\"]}\n\n"
            % (n_suspect, "；".join("第%s行 %s" % (r[0], r[1]) for r in suspect_summary[:5])))
    if empty_formula_cols:
        sys.stderr.write(
            "\n[严重] 以下列有公式但取值全空，说明这个 xlsx 没有缓存值（多半是程序生成的），"
            "**本次验证结果不可信**：\n    %s\n"
            "    修复：用 Excel/LibreOffice 把原表另存一次，或执行\n"
            "    soffice --headless --convert-to xlsx --outdir <目录> '%s'\n"
            "    重算后再跑一遍 verify。\n\n" % ("；".join(empty_formula_cols), wb_path))
    print("[verify] 数据行 %d，校验项 %d，比对 %d 次，不匹配 %d 次，通过率 %.2f%%"
          % (n_rows, len(checks), total_cmp, total_bad, pass_rate))
    for n in names:
        s = stats[n]
        print("         - %-24s 比对 %4d  不匹配 %4d  通过率 %6.2f%%  最大差 %.4f"
              % (n, s["cmp"], s["bad"], (s["ok"] / float(s["cmp"]) * 100 if s["cmp"] else 0), s["maxdiff"]))
    print("[verify] 报告 → %s" % rp)
    print("[verify] 明细 → %s" % detail_csv)
    print("[verify] 不匹配 → %s" % mm_csv)
    if total_bad:
        print("[verify] 注意：有不匹配时先别假设是自己错——也可能是这张 Excel 本身算错了。"
              "差固定值→漏项；差比例→系数口径；只差零星几行→多半是人为覆盖，这恰恰是最有价值的发现。")
    return 0


def run_cross_check(cc, idx, b, base_dir, key_vars, dim_vars):
    """跨表一致性：按 key 对齐右表，比对若干字段。"""
    check_keys(cc, {"name", "workbook", "sheet", "header_rows", "key", "left_key", "compare"},
               "cross_checks[%d]" % idx)
    name = as_str(cc.get("name"), "cross_checks[%d].name" % idx,
                  default="跨表校验%d" % (idx + 1))
    for f in ("workbook", "sheet", "key", "left_key", "compare"):
        if f not in cc:
            raise UserError("cross_checks[%d]「%s」缺少字段 %s" % (idx, name, f))
    path = cc["workbook"]
    if not isinstance(path, str):
        raise UserError("cross_checks[%d].workbook 必须是字符串路径，实际是%s" % (idx, _tname(path)))
    if not os.path.isabs(path):
        path = os.path.normpath(os.path.join(base_dir, path))
    hr = as_int(cc.get("header_rows"), "cross_checks[%d].header_rows" % idx, default=1, minimum=0)
    # sheet / left_key 后面都要拿去做 dict 查找，写成数组/对象会抛 unhashable，
    # 写成数字会被 pandas 当成 Sheet 下标静默读到另一张表 —— 先把类型闸死
    sheet_name = as_str(cc.get("sheet"), "cross_checks[%d].sheet" % idx)
    left_key = as_str(cc.get("left_key"), "cross_checks[%d].left_key" % idx)
    rsd = load_sheet(path, sheet_name, fill_merged_rows=None)
    if hr >= rsd.nrows:
        raise UserError("cross_checks[%d]「%s」header_rows=%d 超过右表总行数 %d" % (idx, name, hr, rsd.nrows))
    rcols, _ = build_col_names(rsd, hr)
    kcol = resolve_column(cc["key"], rcols, where="cross_checks[%d].key" % idx)
    if left_key not in b.var2col:
        raise UserError("cross_checks[%d] 的 left_key=%s 未在主表 keys/dimensions/fields 中定义"
                        % (idx, left_key))

    pairs = []
    for j, cp in enumerate(as_list(cc["compare"], "cross_checks[%d].compare" % idx)):
        as_item_dict(cp, "cross_checks[%d].compare[%d]" % (idx, j))
        check_keys(cp, {"left", "right", "tolerance"},
                   "cross_checks[%d].compare[%d]" % (idx, j))
        for f in ("left", "right"):
            if f not in cp:
                raise UserError("cross_checks[%d].compare[%d] 缺少字段 %s" % (idx, j, f))
        lv = as_str(cp["left"], "cross_checks[%d].compare[%d].left" % (idx, j))
        if lv not in b.var2col:
            raise UserError("cross_checks[%d].compare[%d] 的 left=%s 未在主表中定义（derived 变量"
                            "不支持跨表比对，请先把它写进 fields 或换成输入字段）" % (idx, j, lv))
        rc = resolve_column(cp["right"], rcols, where="cross_checks[%d].compare[%d].right" % (idx, j))
        tol = as_float(cp.get("tolerance"), "cross_checks[%d].compare[%d].tolerance" % (idx, j),
                       default=0.001, minimum=0.0)
        pairs.append((lv, cp["right"], rc, tol))

    rmap = {}
    dup = set()
    for r in range(hr, rsd.nrows):
        k = norm_key(cell_str(rsd.val(r, kcol)))
        if not k:
            continue
        if k in rmap:
            dup.add(k)
        else:
            rmap[k] = r

    res = {"name": name, "workbook": path, "sheet": sheet_name, "key": rcols[kcol],
           "left_key": left_key, "dup_keys": sorted(dup)[:10], "n_dup": len(dup),
           "items": [], "not_found": [], "matched_rows": 0, "total_rows": 0}
    per = {p[0]: {"cmp": 0, "bad": 0, "maxdiff": 0.0, "right": p[1], "tol": p[3], "bad_rows": []}
           for p in pairs}

    for r in range(b.header_rows, b.sd.nrows):
        vals = b.sd.values[r][:b.sd.ncols]
        if all(is_blank(v) for v in vals):
            continue
        lk = norm_key(cell_str(b.sd.val(r, b.var2col[left_key])))
        if not lk:
            continue
        res["total_rows"] += 1
        if lk not in rmap:
            if len(res["not_found"]) < 30:
                res["not_found"].append([r + 1, lk])
            continue
        res["matched_rows"] += 1
        rr = rmap[lk]
        for (lv, rname, rc, tol) in pairs:
            a = to_float(b.sd.val(r, b.var2col[lv]))[0]
            bb = to_float(rsd.val(rr, rc))[0]
            d = a - bb
            st = per[lv]
            st["cmp"] += 1
            st["maxdiff"] = max(st["maxdiff"], abs(d))
            if abs(d) > tol:
                st["bad"] += 1
                if len(st["bad_rows"]) < 20:
                    st["bad_rows"].append([r + 1, lk, round(a, 4), round(bb, 4), round(d, 4)])
    res["items"] = [(lv, per[lv]) for (lv, _n, _c, _t) in pairs]
    return res


def render_verify_report(b, wb_path, names, checks, derived, stats, mismatch_rows, mm_headers,
                         cross_results, warns, n_rows, skipped_empty, skipped_label,
                         total_cmp, total_bad, pass_rate, detail_csv, mm_csv,
                         empty_formula_cols=(), suspect_summary=(), n_suspect=0,
                         headerish=(), n_headerish=0, n_loose=0):
    L = ["# 验证报告（Step 4：用全量真实数据回归测试）", "",
         "> 跳过验证直接分析 = 在没读懂代码的情况下改代码。通过率 100% 才算 AI 学会了这套业务逻辑。", ""]
    if empty_formula_cols:
        L += ["> ## ⛔ 本次结果不可信，请先修表再看下面的数字",
              ">",
              "> 以下列**有公式但取值全空**，说明这个 xlsx 没有缓存值（多半是程序生成的）：",
              ">",
              "> " + "；".join(empty_formula_cols),
              ">",
              "> 脚本把空单元格按 0.0 处理，于是所有以它们为 target 的校验项必然全错——"
              "这不是 AI 理解错了，是数据没读到。",
              ">",
              "> **修复：** 用 Excel/LibreOffice 把原表另存一次，或执行",
              "> `soffice --headless --convert-to xlsx --outdir <目录> '%s'`，重算后再跑一遍 verify。" % wb_path,
              ""]
    L.append("## 一、概况")
    L.append("")
    L.append(md_table(["项", "值"], [
        ["工作簿", "`%s`" % wb_path],
        ["Sheet", b.sd.sheet],
        ["表头行数", b.header_rows],
        ["原始行数", b.sd.nrows],
        ["参与比对的数据行", n_rows],
        ["按 skip_when.empty 跳过", skipped_empty],
        ["按 skip_when.label_in 跳过（汇总/小计行）", skipped_label],
        ["校验项数", len(checks)],
        ["比对次数（行 × 校验项）", total_cmp],
        ["不匹配次数", total_bad],
        ["**总体通过率**", "**%.2f%%**" % pass_rate],
    ]))
    L.append("")
    caveat = ""
    if n_headerish:
        caveat += ("\n\n> ⚠️ **但这个通过率不能信**：有 %d 行的每个数值字段装的都是文本"
                   "（疑似表头行被当成了数据行），它们 AI 算 0、Excel 也读成 0，"
                   "是被「0==0」白送的通过。先把 `header_rows` 数对再看这份报告。" % n_headerish)
    if n_loose:
        caveat += ("\n\n> ⚠️ **有 %d 次「通过」是靠放大容差换来的**：它们的绝对差异其实超过 %s，"
                   "只是没超过 spec 里写的 tolerance。把容差调回 0.01（比率 0.0001）重跑，"
                   "再决定这些差异是谁的错。" % (n_loose, STRICT_TOL))
    if b.header_guess != b.header_rows:
        caveat += ("\n\n> ⚠️ **header_rows 存疑**：spec 写 %d，脚本推断 %d。"
                   "表头行数数错一行全废，请对照 tomd 的 MD 确认。" % (b.header_rows, b.header_guess))
    # 有水分的时候绝不能让第一行写着"✅ 全部通过"——读报告的人只会看这一句
    clean_pass = ("✅ 全部通过 —— AI 对这套业务逻辑的理解与 Excel 实际值完全一致，可以进入 Step 5。"
                  if not caveat else
                  "⚠️ 字面上全部通过，但**这个通过率有水分**（原因见下），"
                  "先照着提示修好 spec 再重跑，不要就这么进 Step 5。")
    verdict = (clean_pass if total_bad == 0 else
               "⚠️ 存在 %d 处不匹配 —— **先别假设是自己错**：可能是理解错了，"
               "也可能是这张 Excel 本身算错了。看差异模式："
               "差一个固定值→几乎总是漏了某一项；差一个比例→系数/口径不同；"
               "只有零星几行差→多半是人为覆盖或错误，这恰恰是最有价值的发现。" % total_bad)
    L += ["**结论：** %s%s" % (verdict, caveat), ""]

    L.append("## 二、列绑定")
    L.append("")
    kindname = {"key": "键", "dim": "维度", "field": "字段"}
    rows = [[v, kindname.get(b.var_kind[v], ""), c, get_column_letter(c + 1), b.col_names[c]]
            for v, c in b.var2col.items()]
    L.append(md_table(["变量", "角色", "列号(0基)", "列标", "拼接列名"], rows))
    L.append("")
    if derived:
        L.append("**中间量（derived，按声明顺序求值）：**")
        L.append("")
        L.append(md_table(["变量", "表达式"], [[k, "`%s`" % v] for k, v in derived.items()]))
        L.append("")

    L.append("## 三、校验项定义与结果")
    L.append("")
    rows = []
    for ck in checks:
        n = ck["name"]
        s = stats[n]
        rate = (s["ok"] / float(s["cmp"]) * 100) if s["cmp"] else 0.0
        rows.append([n, ck["target"], "`%s`" % ck["expr"], ck.get("tolerance", 0.01),
                     s["cmp"], s["cmp"] - s["bad"], s["bad"],
                     "%.2f%%" % rate, fmt_num(s["maxdiff"], 4),
                     "✅" if s["bad"] == 0 else "❌"])
    L.append(md_table(["校验项", "target(Excel列)", "AI 表达式", "容差", "比对行数",
                       "匹配", "不匹配", "通过率", "最大绝对差", ""], rows))
    L.append("")
    tiny = sum(s["tiny"] for s in stats.values())
    if tiny:
        L.append("_另有 %d 次差异非零但在容差内（浮点尾差），不计为不匹配。_" % tiny)
        L.append("")
    if n_loose:
        L.append("**其中 %d 次的绝对差异超过 %s，是「仅因容差被放大才算通过」的——逐项：%s。**"
                 % (n_loose, STRICT_TOL,
                    "；".join("%s %d 次（容差 %s）" % (n, stats[n]["loose"],
                                                  next((ck.get("tolerance", 0.01)
                                                        for ck in checks if ck["name"] == n), 0.01))
                              for n in names if stats[n]["loose"])))
        L.append("")

    L.append("## 四、不匹配明细")
    L.append("")
    if not mismatch_rows:
        L.append("_无不匹配。_")
    else:
        show = mismatch_rows[:60]
        L.append(md_table(mm_headers, show))
        if len(mismatch_rows) > 60:
            L.append("")
            L.append("_……共 %d 条，仅列前 60 条，全量见 `%s`。_"
                     % (len(mismatch_rows), os.path.basename(mm_csv)))
        L.append("")
        # 差异模式提示
        L.append("**差异模式速查：**")
        L.append("")
        prows = []
        # mm_headers 末六列固定为：校验项 / AI值 / Excel值 / 差异 / 容差 / AI 表达式
        i_name, i_ai, i_ex, i_diff = (len(mm_headers) - 6, len(mm_headers) - 5,
                                      len(mm_headers) - 4, len(mm_headers) - 3)
        for n in names:
            hits = [m for m in mismatch_rows if m[i_name] == n]
            if not hits:
                continue
            ds = [float(m[i_diff]) for m in hits]
            uniq = sorted(set(round(d, 2) for d in ds))
            ratios = [float(m[i_ai]) / float(m[i_ex]) for m in hits if float(m[i_ex]) != 0]
            total = stats[n]["cmp"]
            if len(hits) <= max(3, int(total * 0.05)):
                mode = "只有零星 %d 行差（%.1f%%）→ 多半是 Excel 里的人为覆盖或错误，" \
                       "**这恰恰是最有价值的发现**，请逐行人工核查" % (len(hits), len(hits) / float(total) * 100)
            elif len(uniq) == 1:
                mode = "差值恒为 %s → 几乎总是漏了某一项" % fmt_num(uniq[0])
            elif len(ratios) == len(hits) and max(ratios) - min(ratios) < 1e-6:
                mode = "AI/Excel 恒为 %.6f → 系数或口径不同" % ratios[0]
            else:
                mode = "差值有 %d 种取值（%s…）→ 逐行看，可能混了多种情形" \
                       % (len(uniq), "、".join(fmt_num(u) for u in uniq[:5]))
            prows.append([n, len(ds), mode])
        L.append(md_table(["校验项", "不匹配条数", "差异模式"], prows) if prows else "_无。_")
    L.append("")

    L.append("## 五、汇总额对比（AI 合计 vs Excel 合计）")
    L.append("")
    rows = []
    for n in names:
        s = stats[n]
        d = s["sum_ai"] - s["sum_ex"]
        rel = (abs(d) / abs(s["sum_ex"]) * 100) if s["sum_ex"] else 0.0
        rows.append([n, fmt_num(s["sum_ai"]), fmt_num(s["sum_ex"]), fmt_num(d),
                     "%.4f%%" % rel, "✅" if abs(d) < 0.01 else "⚠️"])
    L.append(md_table(["校验项", "AI 合计", "Excel 合计", "差额", "相对差异", ""], rows))
    L.append("")

    L.append("## 六、跨表一致性")
    L.append("")
    if not cross_results:
        L.append("_spec.json 未配置 cross_checks。_")
    for cr in cross_results:
        L.append("### %s" % cr["name"])
        L.append("")
        L.append("- 右表：`%s` / Sheet `%s`，连接键：右 `%s` ← 左 `%s`"
                 % (cr["workbook"], cr["sheet"], cr["key"], cr["left_key"]))
        L.append("- 左表参与 %d 行，其中 %d 行在右表命中，%d 行未找到"
                 % (cr["total_rows"], cr["matched_rows"], cr["total_rows"] - cr["matched_rows"]))
        if cr["n_dup"]:
            L.append("- ⚠️ 右表连接键有 %d 个重复值（只取首次出现行）：%s"
                     % (cr["n_dup"], "、".join(cr["dup_keys"])))
        L.append("")
        rows = []
        for lv, st in cr["items"]:
            rate = ((st["cmp"] - st["bad"]) / float(st["cmp"]) * 100) if st["cmp"] else 0.0
            rows.append([lv, st["right"], st["tol"], st["cmp"], st["bad"],
                         "%.2f%%" % rate, fmt_num(st["maxdiff"], 4),
                         "✅" if st["bad"] == 0 else "❌"])
        L.append(md_table(["左字段", "右字段", "容差", "比对行数", "不匹配", "通过率", "最大绝对差", ""], rows))
        L.append("")
        for lv, st in cr["items"]:
            if st["bad_rows"]:
                L.append("**%s 不匹配明细（前 %d 条）：**" % (lv, len(st["bad_rows"])))
                L.append("")
                L.append(md_table(["Excel行号", "连接键", "左值", "右值", "差异"], st["bad_rows"]))
                L.append("")
        if cr["not_found"]:
            L.append("**右表未找到的连接键（前 %d 条，常见于离职/中途入职人员）：**" % len(cr["not_found"]))
            L.append("")
            L.append(md_table(["Excel行号", "连接键"], cr["not_found"]))
            L.append("")

    if n_suspect:
        L.append("## 六之二、疑似未排除的汇总行（金额会翻倍）")
        L.append("")
        L.append("以下 **%d** 行含「%s」等汇总词，却仍被当成数据行参与了比对——"
                 "它们的金额会和明细行重复计入，上面的「汇总额对比」直接翻倍。"
                 % (n_suspect, "、".join(DEFAULT_SUMMARY_WORDS[:4])))
        L.append("")
        L.append("修复：在 spec.json 里加 `\"skip_when\": {\"label_in\": [\"合计\", \"小计\", \"总计\"]}` 后重跑。")
        L.append("")
        L.append(md_table(["Excel 行号", "内容"], list(suspect_summary)))
        if n_suspect > len(suspect_summary):
            L.append("")
            L.append("_……共 %d 行，仅列前 %d 行。_" % (n_suspect, len(suspect_summary)))
        L.append("")

    if n_headerish:
        L.append("## 六之三、疑似表头行混入数据体（0==0 白送的通过）")
        L.append("")
        L.append("以下 **%d** 行的**每一个数值字段装的都是文本**——这正是 `header_rows` 少数了一行时"
                 "的样子：字段名那一行被当成了数据行。" % n_headerish)
        L.append("")
        L.append("这类行按 0.0 参与运算，于是 AI 算出 0、Excel 现值也读成 0，"
                 "**0==0 被判为通过**，凭空推高通过率、还多出一行数据。")
        L.append("")
        L.append("修复：spec 的 `header_rows` 从 %d 改成 %d（或对照 tomd 的 MD 再数一遍）后重跑。"
                 % (b.header_rows, b.header_rows + 1))
        L.append("")
        L.append(md_table(["Excel 行号", "该行各数值字段的实际内容"], list(headerish)))
        if n_headerish > len(headerish):
            L.append("")
            L.append("_……共 %d 行，仅列前 %d 行。_" % (n_headerish, len(headerish)))
        L.append("")

    L.append("## 七、告警")
    L.append("")
    allw = ["**有公式但取值全空（xlsx 无缓存值，见开头的红色横幅）：** %s" % w
            for w in empty_formula_cols] + list(warns)
    L.append("\n".join("- " + w for w in allw) if allw else "_无。_")
    L.append("")
    L.append("## 八、产物")
    L.append("")
    L.append("- 逐行明细：`%s`（每个校验项三列：`名%sAI` / `名%sExcel` / `名%s差异`，utf-8-sig）"
             % (os.path.basename(detail_csv), SEP, SEP, SEP))
    L.append("- 不匹配清单：`%s`" % os.path.basename(mm_csv))
    L.append("")
    L.append("下一步：`excel_ai.py output <spec> -d <明细csv> -o <目录>` 与 "
             "`excel_ai.py analyze <spec> -d <明细csv> -o <目录>`。")
    L.append("")
    return "\n".join(L)


# ===========================================================================
# 七、detail CSV 读取（output / analyze 共用）
# ===========================================================================

class Detail(object):
    def __init__(self, headers, rows, spec):
        self.headers = headers
        self.rows = rows                      # list[dict]  列名 → 原始字符串
        self.key_vars = list(as_dict(spec.get("keys"), "keys").keys())
        self.dim_vars = list(as_dict(spec.get("dimensions"), "dimensions").keys())
        self.field_vars = list(as_dict(spec.get("fields"), "fields").keys())
        self.derived_vars = list(as_dict(spec.get("derived"), "derived").keys())
        check_keys(spec, SPEC_TOP_KEYS, "")
        self.checks = as_list(spec.get("checks"), "checks")
        for i, c in enumerate(self.checks):
            as_item_dict(c, "checks[%d]" % i)
            check_keys(c, {"name", "target", "expr", "tolerance"}, "checks[%d]" % i)
            if not c.get("name"):
                raise UserError("checks[%d] 缺少字段 name" % i)
            as_str(c["name"], "checks[%d].name" % i)
            if c["name"].endswith(RAW_SUFFIX):
                raise UserError("checks[%d] 的校验项名「%s」不能以 `%s` 结尾——"
                                "它会和自动生成的原始文本变量撞名，分析时会把文本当金额求和。"
                                % (i, c["name"], RAW_SUFFIX))
        for v in self.key_vars + self.dim_vars + self.field_vars + self.derived_vars:
            validate_var_name(v, "keys/dimensions/fields/derived")
        self.check_names = [c["name"] for c in self.checks]
        # 数值型可分析列（analyze 的 field 只能取这些，取文本列会让 pandas 直接崩）
        self.numeric_vars = list(self.field_vars) + list(self.derived_vars) + \
            [n for n in self.check_names if n not in self.field_vars and n not in self.derived_vars]

    def require_numeric(self, name, where):
        """analyze 的 field 必须是数值列，否则给中文提示而不是 pandas 的 ValueError。"""
        if name in self.numeric_vars:
            return
        if name in self.key_vars or name in self.dim_vars:
            raise UserError("%s 的 field=%s 是键/维度列（文本），不能做数值分析。\n"
                            "    可用的数值列：%s\n"
                            "    （想看文本列的分布请改用 analysis.distributions）"
                            % (where, name, "、".join(self.numeric_vars) or "(无)"))
        raise UserError("%s 的 field=%s 不在明细列中。可用的数值列：%s"
                        % (where, name, "、".join(self.numeric_vars) or "(无)"))

    def col_kind(self, h):
        """列 → 类别：input / ai / diff / excel / other"""
        if h == "Excel行号":
            return "other"
        if h in self.key_vars or h in self.dim_vars or h in self.field_vars:
            return "input"
        if h in self.derived_vars:
            return "ai"
        for n in self.check_names:
            if h == "%s%sAI" % (n, SEP):
                return "ai"
            if h == "%s%sExcel" % (n, SEP):
                return "excel"
            if h == "%s%s差异" % (n, SEP):
                return "diff"
        return "other"

    def num(self, row, h, default=0.0):
        if h not in row:
            return default
        f, ok = to_float(row[h])
        return f if ok else default

    def env(self, row):
        """行 → 求值环境（字段/中间量为 float，键/维度为字符串）。"""
        e = {}
        for v in self.key_vars + self.dim_vars:
            s = str(row.get(v, "") or "")
            e[v] = s
            e[v + "__raw"] = s
        for v in self.field_vars + self.derived_vars:
            e[v] = self.num(row, v)
            e[v + "__raw"] = str(row.get(v, "") or "")
        # 校验项结果（AI 值）以校验项名暴露。
        # 这里**不能**只放合法标识符：`analysis` 的 group_by.metrics / top_n.field /
        # outliers.field 是按名字查字典的，校验项名叫「应发工资G=A+B-C」完全合法，
        # 但它进不了 env 就会在 `e[m]` 上抛 KeyError（未捕获的 traceback）。
        # 非标识符的名字本来就没法被表达式解析出来，放进 env 不会污染命名空间。
        for n in self.check_names:
            if n not in e:
                e[n] = self.num(row, "%s%sAI" % (n, SEP))
        return e


def load_detail(path, spec):
    if not os.path.isfile(path):
        raise UserError("找不到验证明细 CSV：%s\n    先跑 `excel_ai.py verify` 生成 verification_detail.csv" % path)
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rd = csv.DictReader(f)
        headers = [csv_unguard(h) for h in (rd.fieldnames or [])]
        rows = [{csv_unguard(k): csv_unguard(v) for k, v in r.items()} for r in rd]
    if not headers:
        raise UserError("验证明细 CSV 为空：%s" % path)
    dups = sorted({h for h in headers if headers.count(h) > 1})
    if dups:
        raise UserError("验证明细 CSV 存在同名列：%s\n"
                        "    同名列会被静默合并成一列，后面所有数字都不可信。请检查 spec 的变量名"
                        "与校验项名是否撞车，改名后重跑 verify。" % "、".join(dups))
    d = Detail(headers, rows, spec)
    missing = [v for v in (d.key_vars + d.dim_vars + d.field_vars + d.derived_vars) if v not in headers]
    # 校验项的三列同样要在场。少了它们时 d.num() 会一路返回默认 0.0：
    # 「校验汇总」印出比对 N 次、不匹配 0 次、通过率 100%，分析报告里 TopN 全是 0 ——
    # 一份格式完好、数字全假的交付物。改了 check 名却忘了重跑 verify 就是这样。
    for n in d.check_names:
        for suf in ("AI", "Excel", "差异"):
            col = "%s%s%s" % (n, SEP, suf)
            if col not in headers:
                missing.append(col)
    if missing:
        raise UserError("明细 CSV 与 spec 不匹配，缺少列：%s\n"
                        "    该 CSV 是用另一份 spec 生成的（多半是改了变量名/校验项名却没重跑 verify）。\n"
                        "    请先重跑：`excel_ai.py verify <spec> -o <目录>`，再拿新的"
                        " verification_detail.csv 跑 output/analyze。\n"
                        "    CSV 现有列：%s"
                        % ("、".join(clip(m, 40) for m in missing),
                           "、".join(clip(h, 30) for h in headers)))
    if not rows:
        raise UserError("验证明细 CSV 没有数据行：%s" % path)
    return d


def render_block(v, indent=0):
    """
    把 spec 里的 lineage / ontology 段渲染成 (行列表)，供写入 Excel。
    支持：str / list[str] / list[dict] / dict（值可再嵌套）。
    返回 [(缩进级别, [单元格...], 是否表头行)]
    """
    out = []
    if v is None:
        return out
    if isinstance(v, str):
        out.append((indent, [v], False))
    elif isinstance(v, (int, float, bool)):
        out.append((indent, [str(v)], False))
    elif isinstance(v, list):
        if v and all(isinstance(x, dict) for x in v):
            cols = []
            for x in v:
                for k in x:
                    if k not in cols:
                        cols.append(k)
            out.append((indent, cols, True))
            for x in v:
                out.append((indent, [_flat(x.get(k)) for k in cols], False))
        else:
            for x in v:
                out.extend(render_block(x, indent))
    elif isinstance(v, dict):
        for k, sub in v.items():
            out.append((indent, ["【%s】" % k], True))
            out.extend(render_block(sub, indent + 1))
    return out


def _flat(v):
    if v is None:
        return ""
    if isinstance(v, (list, tuple)):
        return " / ".join(_flat(x) for x in v)
    if isinstance(v, dict):
        return "; ".join("%s=%s" % (k, _flat(x)) for k, x in v.items())
    return str(v)


# ===========================================================================
# 八、output —— Step 5：带配色标注的结果 xlsx
# ===========================================================================

def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def put(ws, row, col, value):
    """
    写单元格 + 防公式注入。
    openpyxl 会把以 "=" 开头的字符串自动写成**活公式**（data_type='f'），
    于是原表里一个叫 `=cmd|'/c calc'!A1` 的姓名，会在我们交付的 xlsx 里变成
    一个真正的 DDE 触发点。强制成文本即可，值本身不改（无损）。
    """
    c = ws.cell(row=row, column=col, value=value)
    if isinstance(value, str) and value[:1] == "=":
        c.data_type = "s"
    return c


def _autosize(ws, maxw=42):
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            ln = max(len(seg) for seg in str(c.value).split("\n"))
            ln = ln + sum(1 for ch in str(c.value) if ord(ch) > 127) * 0.7
            widths[c.column] = min(max(widths.get(c.column, 8), ln + 2), maxw)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def cmd_output(args):
    require_deps()
    spec_path = os.path.abspath(args.spec)
    spec = load_spec(spec_path)
    outdir = ensure_dir(os.path.abspath(args.output))
    d = load_detail(os.path.abspath(args.detail), spec)

    tol_of = {c["name"]: as_float(c.get("tolerance"), "checks.%s.tolerance" % c["name"],
                                  default=0.01, minimum=0.0) for c in d.checks}
    expr_of = {c["name"]: c.get("expr", "") for c in d.checks}
    target_of = {c["name"]: c.get("target", "") for c in d.checks}

    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(bold=True)
    title_font = Font(bold=True, size=13)

    # ---------- Sheet 1：AI处理结果 ----------
    ws = wb.active
    ws.title = "AI处理结果"
    put(ws, 1, 1, value="AI 处理结果（%s / %s）"
            % (os.path.basename(str(spec.get("workbook", ""))), spec.get("sheet", ""))).font = title_font
    legend = [("原始输入（取自 Excel 原表）", FILL_INPUT),
              ("AI 计算（按 spec 回算）", FILL_AI),
              ("差异非零（超出容差，需人工核）", FILL_DIFF)]
    put(ws, 2, 1, value="图例：").font = hdr_font
    for i, (txt, color) in enumerate(legend):
        c = put(ws, 2, 2 + i, value=txt)
        c.fill = _fill(color)
        c.border = border
    put(ws, 3, 1, value="共 %d 行 × %d 列；校验项 %d 个（公式说明见「校验汇总」Sheet）"
            % (len(d.rows), len(d.headers), len(d.check_names)))

    HROW = 5
    for j, h in enumerate(d.headers, start=1):
        c = put(ws, HROW, j, value=h)
        c.font = hdr_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        kind = d.col_kind(h)
        if kind in ("input", "excel"):
            c.fill = _fill(FILL_INPUT)
        elif kind == "ai":
            c.fill = _fill(FILL_AI)
        elif kind == "diff":
            c.fill = _fill(FILL_DIFF)

    diff_of_check = {"%s%s差异" % (n, SEP): n for n in d.check_names}
    # 键/维度列一律按文本写出，避免 "007" 这类工号被转成数字丢掉前导零
    text_cols = set(d.key_vars) | set(d.dim_vars)
    n_flag = 0
    for i, row in enumerate(d.rows):
        for j, h in enumerate(d.headers, start=1):
            raw = row.get(h, "")
            if h in text_cols:
                f, ok = 0.0, False
                val = raw if raw != "" else None
            else:
                f, ok = to_float(raw)
                if h == "Excel行号":
                    val = int(f) if ok else (raw or None)
                elif ok:
                    val = round(f, 6)
                else:
                    val = raw if raw != "" else None
            c = put(ws, HROW + 1 + i, j, value=val)
            c.border = border
            kind = d.col_kind(h)
            if kind in ("input", "excel"):
                c.fill = _fill(FILL_INPUT)
            elif kind == "ai":
                c.fill = _fill(FILL_AI)
            elif kind == "diff":
                n = diff_of_check.get(h, "")
                if ok and abs(f) > tol_of.get(n, 0.01):
                    c.fill = _fill(FILL_DIFF)
                    c.font = Font(bold=True, color="C00000")
                    n_flag += 1
            if ok and h != "Excel行号" and abs(f) >= 0.005:
                c.number_format = "#,##0.00"
    ws.freeze_panes = ws.cell(row=HROW + 1, column=1)   # 只取坐标，不能写值
    _autosize(ws)

    # ---------- Sheet 2：校验汇总 ----------
    ws2 = wb.create_sheet("校验汇总")
    r = 1
    put(ws2, r, 1, value="校验汇总").font = title_font
    r += 2
    heads = ["校验项", "target(Excel列)", "AI 表达式", "容差", "比对行数", "匹配", "不匹配",
             "通过率", "最大绝对差", "AI 合计", "Excel 合计", "差额"]
    for j, h in enumerate(heads, start=1):
        c = put(ws2, r, j, value=h)
        c.font = hdr_font
        c.fill = _fill(FILL_INPUT)
        c.border = border
    r += 1
    tot_cmp = tot_bad = 0
    for n in d.check_names:
        tol = tol_of.get(n, 0.01)
        cmpn = ok = bad = 0
        mx = sai = sex = 0.0
        for row in d.rows:
            ai = d.num(row, "%s%sAI" % (n, SEP))
            ex = d.num(row, "%s%sExcel" % (n, SEP))
            df = d.num(row, "%s%s差异" % (n, SEP))
            cmpn += 1
            sai += ai
            sex += ex
            mx = max(mx, abs(df))
            if abs(df) > tol:
                bad += 1
            else:
                ok += 1
        tot_cmp += cmpn
        tot_bad += bad
        vals = [n, target_of.get(n, ""), expr_of.get(n, ""), tol, cmpn, ok, bad,
                (ok / float(cmpn) if cmpn else 0), round(mx, 4), round(sai, 2),
                round(sex, 2), round(sai - sex, 2)]
        for j, v in enumerate(vals, start=1):
            c = put(ws2, r, j, value=v)
            c.border = border
            if j == 8:
                c.number_format = "0.00%"
            if j == 7 and bad:
                c.fill = _fill(FILL_DIFF)
        r += 1
    r += 1
    put(ws2, r, 1, value="总体通过率").font = hdr_font
    put(ws2, r, 2, value=((tot_cmp - tot_bad) / float(tot_cmp) if tot_cmp else 0)
             ).number_format = "0.00%"
    put(ws2, r, 3, value="比对 %d 次，不匹配 %d 次" % (tot_cmp, tot_bad))
    r += 2

    put(ws2, r, 1, value="公式说明区").font = title_font
    r += 1
    for j, h in enumerate(["校验项", "AI 表达式（脚本按此回算）", "Excel 中对应的结果列", "容差"], start=1):
        c = put(ws2, r, j, value=h)
        c.font = hdr_font
        c.fill = _fill(FILL_AI)
        c.border = border
    r += 1
    for n in d.check_names:
        for j, v in enumerate([n, expr_of.get(n, ""), target_of.get(n, ""), tol_of.get(n, 0.01)], start=1):
            put(ws2, r, j, value=v).border = border
        r += 1
    if spec.get("derived"):
        r += 1
        put(ws2, r, 1, value="中间量（derived，按声明顺序求值）").font = hdr_font
        r += 1
        for j, h in enumerate(["变量", "表达式"], start=1):
            c = put(ws2, r, j, value=h)
            c.font = hdr_font
            c.fill = _fill(FILL_AI)
            c.border = border
        r += 1
        for k, v in spec["derived"].items():
            put(ws2, r, 1, value=k).border = border
            put(ws2, r, 2, value=v).border = border
            r += 1
    r += 1
    put(ws2, r, 1, value="图例").font = title_font
    r += 1
    for txt, color in legend:
        c = put(ws2, r, 1, value="")
        c.fill = _fill(color)
        c.border = border
        put(ws2, r, 2, value=txt)
        r += 1
    _autosize(ws2, maxw=70)

    # ---------- Sheet 3/4：数据血缘 / 字段本体 ----------
    for title, key, hint in (("数据血缘", "lineage", "Step 3 的跨表数据流向（源 Sheet.字段 → 目标 Sheet.字段 / 连接键 / 说明）"),
                             ("字段本体", "ontology", "Step 2 的字段定义（字段名 / 业务含义 / 类型 / 计算关系 / 所属表）")):
        wsx = wb.create_sheet(title)
        put(wsx, 1, 1, value=title).font = title_font
        put(wsx, 2, 1, value=hint)
        blocks = render_block(spec.get(key))
        rr = 4
        if not blocks:
            put(wsx, rr, 1,
                     value="（spec.json 未提供 `%s` 段。这一段是 AI 手写的认知产出，"
                           "写进 spec 后重跑 output 即可带出来。）" % key)
        for indent, cells, is_head in blocks:
            for j, v in enumerate(cells, start=1):
                c = put(wsx, rr, indent + j, value=_flat(v))
                c.border = border
                c.alignment = Alignment(vertical="top", wrap_text=True)
                if is_head:
                    c.font = hdr_font
                    c.fill = _fill(FILL_INPUT if title == "字段本体" else FILL_AI)
            rr += 1
        _autosize(wsx, maxw=60)

    fp = os.path.join(outdir, "AI处理结果.xlsx")
    ensure_dir(outdir)
    wb.save(fp)
    print("[output] 结果表 → %s" % fp)
    print("[output] %d 行 × %d 列；标橙（差异超容差）单元格 %d 个；"
          "Sheet：AI处理结果 / 校验汇总 / 数据血缘 / 字段本体"
          % (len(d.rows), len(d.headers), n_flag))
    if not spec.get("lineage"):
        print("[output] 提示：spec.json 未提供 lineage 段，「数据血缘」Sheet 为空。")
    if not spec.get("ontology"):
        print("[output] 提示：spec.json 未提供 ontology 段，「字段本体」Sheet 为空。")
    return 0


# ===========================================================================
# 九、analyze —== Step 5：业务分析报告
# ===========================================================================

def _pct(x, total):
    return "%.2f%%" % (x / float(total) * 100) if total else "-"


def cmd_analyze(args):
    require_deps()
    spec_path = os.path.abspath(args.spec)
    spec = load_spec(spec_path)
    outdir = ensure_dir(os.path.abspath(args.output))
    d = load_detail(os.path.abspath(args.detail), spec)
    an = check_keys(as_dict(spec.get("analysis"), "analysis"),
                    {"group_by", "distributions", "top_n", "outliers", "rules", "what_if"},
                    "analysis")

    envs = [d.env(r) for r in d.rows]
    n = len(envs)

    L = ["# 分析报告（Step 5）", "",
         "- 数据源：`%s` / Sheet `%s`" % (spec.get("workbook", ""), spec.get("sheet", "")),
         "- 明细：`%s`（%d 行）" % (os.path.basename(args.detail), n),
         "- 生成：`excel_ai.py analyze` v%s" % VERSION, ""]
    if not an:
        L += ["> ⚠️ spec.json 未提供 `analysis` 段，本报告只有数据概览。",
              "> 补上 `analysis`（group_by / distributions / top_n / outliers / rules / what_if）后重跑。", ""]

    # ---- 〇、数值字段概览 ----
    L += ["## 〇、数值字段概览", ""]
    rows = []
    for v in d.field_vars + d.derived_vars:
        xs = [e[v] for e in envs]
        nz = [x for x in xs if x != 0]
        rows.append([v, len(xs), len(nz), fmt_num(sum(xs)), fmt_num(sum(xs) / n if n else 0),
                     fmt_num(min(xs) if xs else 0), fmt_num(max(xs) if xs else 0)])
    L.append(md_table(["字段", "行数", "非零行数", "合计", "均值", "最小", "最大"], rows) if rows else "_无。_")
    L.append("")

    sec = 0

    # ---- 一、分组汇总 ----
    if an.get("group_by"):
        sec += 1
        L += ["## %d、分组汇总" % sec, ""]
        for gi, g in enumerate(as_list(an["group_by"], "analysis.group_by")):
            as_item_dict(g, "analysis.group_by[%d]" % gi)
            check_keys(g, {"dim", "metrics", "count_as"}, "analysis.group_by[%d]" % gi)
            dim = g.get("dim")
            if not dim:
                raise UserError("analysis.group_by[%d] 每项必须有 dim" % gi)
            dim = as_str(dim, "analysis.group_by[%d].dim" % gi)
            if dim not in envs[0]:
                raise UserError("analysis.group_by 的 dim=%s 不在明细列中。可用：%s"
                                % (dim, "、".join(d.key_vars + d.dim_vars)))
            metrics = as_str_list(g.get("metrics"), "analysis.group_by[%d].metrics" % gi)
            for m in metrics:
                d.require_numeric(m, "analysis.group_by[%d].metrics" % gi)
            cnt_as = as_str(g.get("count_as"), "analysis.group_by[%d].count_as" % gi,
                            default="数量")
            groups = OrderedDict()
            for e in envs:
                k = str(e.get(dim, "")) or "(空)"
                groups.setdefault(k, []).append(e)
            totals = {m: sum(e[m] for e in envs) for m in metrics}
            heads = [dim, cnt_as, "占比"]
            for m in metrics:
                heads += ["%s 总额" % m, "%s 人均" % m, "%s 占比" % m]
            rows = []
            if metrics:  # 按第一个指标的总额降序
                order = sorted(groups.items(), key=lambda kv: -sum(e[metrics[0]] for e in kv[1]))
            else:        # 无指标时按数量降序
                order = sorted(groups.items(), key=lambda kv: -len(kv[1]))
            for k, es in order:
                row = [k, len(es), _pct(len(es), n)]
                for m in metrics:
                    s = sum(e[m] for e in es)
                    row += [fmt_num(s), fmt_num(s / len(es) if es else 0), _pct(s, totals[m])]
                rows.append(row)
            trow = ["**合计**", n, "100.00%"]
            for m in metrics:
                trow += ["**%s**" % fmt_num(totals[m]), fmt_num(totals[m] / n if n else 0), "100.00%"]
            rows.append(trow)
            L += ["### 按「%s」分组（%d 组）" % (dim, len(groups)), "", md_table(heads, rows), ""]

    # ---- 二、取值分布 ----
    if an.get("distributions"):
        sec += 1
        L += ["## %d、取值分布" % sec, ""]
        for di, it in enumerate(as_list(an["distributions"], "analysis.distributions")):
            as_item_dict(it, "analysis.distributions[%d]" % di)
            check_keys(it, {"dim", "field", "bins"}, "analysis.distributions[%d]" % di)
            dim = it.get("dim") or it.get("field")
            if not dim:
                raise UserError("analysis.distributions[%d] 每项必须有 dim 或 field" % di)
            dim = as_str(dim, "analysis.distributions[%d].dim" % di)
            if dim not in envs[0]:
                raise UserError("analysis.distributions 的 %s 不在明细列中。可用：%s"
                                % (dim, "、".join(d.key_vars + d.dim_vars + d.numeric_vars)))
            bins = as_list(it.get("bins"), "analysis.distributions[%d].bins" % di)
            if bins:
                d.require_numeric(dim, "analysis.distributions[%d]（配了 bins 就必须是数值列）" % di)
                edges = [as_float(x, "analysis.distributions[%d].bins[%d]" % (di, k))
                         for k, x in enumerate(bins)]
                if sorted(edges) != edges:
                    raise UserError("analysis.distributions[%d].bins 必须按从小到大排列，实际是 %s"
                                    % (di, edges))
                labels = ["< %s" % fmt_num(edges[0])]
                for i in range(len(edges) - 1):
                    labels.append("%s ~ %s" % (fmt_num(edges[i]), fmt_num(edges[i + 1])))
                labels.append(">= %s" % fmt_num(edges[-1]))
                cnt = OrderedDict((lb, 0) for lb in labels)
                for e in envs:
                    x = e[dim] if isinstance(e[dim], (int, float)) else to_float(e[dim])[0]
                    idx = len(edges)
                    for i, ed in enumerate(edges):
                        if x < ed:
                            idx = i
                            break
                    cnt[labels[idx]] += 1
                rows = [[lb, c, _pct(c, n)] for lb, c in cnt.items()]
                L += ["### 「%s」分箱分布" % dim, "", md_table(["区间", "数量", "占比"], rows), ""]
            else:
                cnt = {}
                for e in envs:
                    k = str(e.get(dim, "")) or "(空)"
                    cnt[k] = cnt.get(k, 0) + 1
                rows = [[k, c, _pct(c, n)] for k, c in sorted(cnt.items(), key=lambda kv: -kv[1])]
                L += ["### 「%s」取值分布（%d 种取值）" % (dim, len(cnt)), "",
                      md_table(["取值", "数量", "占比"], rows[:40]), ""]
                if len(rows) > 40:
                    L += ["_……共 %d 种取值，仅列前 40。_" % len(rows), ""]

    # ---- 三、Top / Bottom ----
    if an.get("top_n"):
        sec += 1
        L += ["## %d、Top / Bottom" % sec, ""]
        for ti, it in enumerate(as_list(an["top_n"], "analysis.top_n")):
            as_item_dict(it, "analysis.top_n[%d]" % ti)
            check_keys(it, {"field", "n", "label"}, "analysis.top_n[%d]" % ti)
            f = it.get("field")
            if not f:
                raise UserError("analysis.top_n[%d] 每项必须有 field" % ti)
            f = as_str(f, "analysis.top_n[%d].field" % ti)
            d.require_numeric(f, "analysis.top_n[%d]" % ti)
            k = as_int(it.get("n"), "analysis.top_n[%d].n" % ti, default=5, minimum=1)
            lab = as_str(it.get("label"), "analysis.top_n[%d].label" % ti,
                         default="", allow_empty=True) \
                or (d.key_vars[0] if d.key_vars else None)
            labs = [x for x in ([lab] if lab else []) + d.key_vars + d.dim_vars
                    if x in envs[0] and x != f]
            labs = list(OrderedDict.fromkeys(labs))[:3]
            ordered = sorted(envs, key=lambda e: e[f], reverse=True)
            for tag, part in (("Top %d" % k, ordered[:k]), ("Bottom %d" % k, ordered[::-1][:k])):
                rows = []
                for i, e in enumerate(part, start=1):
                    rows.append([i] + [str(e.get(x, "")) for x in labs] + [fmt_num(e[f])])
                L += ["### %s —— %s" % (tag, f), "",
                      md_table(["#"] + labs + [f], rows), ""]

    # ---- 四、离群检测 ----
    if an.get("outliers"):
        sec += 1
        L += ["## %d、离群检测" % sec, ""]
        for oi, it in enumerate(as_list(an["outliers"], "analysis.outliers")):
            as_item_dict(it, "analysis.outliers[%d]" % oi)
            check_keys(it, {"field", "method", "k", "z", "label"}, "analysis.outliers[%d]" % oi)
            f = it.get("field")
            if not f:
                raise UserError("analysis.outliers[%d] 每项必须有 field" % oi)
            f = as_str(f, "analysis.outliers[%d].field" % oi)
            d.require_numeric(f, "analysis.outliers[%d]" % oi)
            method = as_str(it.get("method"), "analysis.outliers[%d].method" % oi,
                            default="iqr").lower()
            if method not in ("iqr", "z", "zscore", "z-score"):
                raise UserError("analysis.outliers[%d].method 只支持 iqr 或 zscore，收到：%s"
                                % (oi, clip(method, 30)))
            lab = as_str(it.get("label"), "analysis.outliers[%d].label" % oi,
                         default="", allow_empty=True)
            labs = [x for x in ([lab] if lab else []) + d.key_vars + d.dim_vars
                    if x in envs[0] and x != f]
            labs = list(OrderedDict.fromkeys(labs))[:3]
            xs = [float(e[f]) if isinstance(e[f], (int, float)) else to_float(e[f])[0] for e in envs]
            ser = pd.Series(xs, dtype="float64")
            if method == "iqr":
                k = as_float(it.get("k"), "analysis.outliers[%d].k" % oi, default=1.5, minimum=0.0)
                q1, q3 = float(ser.quantile(0.25)), float(ser.quantile(0.75))
                iqr = q3 - q1
                lo, hi = q1 - k * iqr, q3 + k * iqr
                desc = "IQR 法：Q1=%s，Q3=%s，IQR=%s，k=%s → 正常区间 [%s, %s]" % (
                    fmt_num(q1), fmt_num(q3), fmt_num(iqr), k, fmt_num(lo), fmt_num(hi))
                flagged = [(e, e[f]) for e in envs if e[f] < lo or e[f] > hi]
                extra_head, extra = "偏离", lambda v: fmt_num(v - hi if v > hi else v - lo)
            else:
                z = as_float(it.get("z"), "analysis.outliers[%d].z" % oi, default=3.0, minimum=0.0)
                mu = float(ser.mean())
                sd_ = float(ser.std(ddof=0))
                desc = "Z-score 法：均值=%s，标准差=%s，阈值 |z|>%s" % (fmt_num(mu), fmt_num(sd_), z)
                flagged = [(e, e[f]) for e in envs if sd_ > 0 and abs((e[f] - mu) / sd_) > z]
                extra_head, extra = "z 值", lambda v: "%.2f" % ((v - mu) / sd_ if sd_ else 0)
            flagged.sort(key=lambda t: -abs(t[1]))
            rows = [[i] + [str(e.get(x, "")) for x in labs] + [fmt_num(v), extra(v)]
                    for i, (e, v) in enumerate(flagged[:40], start=1)]
            L += ["### %s 的离群值（%s）" % (f, method.upper()), "", desc, "",
                  "命中 **%d** 条 / 共 %d 条（%s）。" % (len(flagged), n, _pct(len(flagged), n)), ""]
            L += [md_table(["#"] + labs + [f, extra_head], rows) if rows else "_无离群值。_", ""]
            if len(flagged) > 40:
                L += ["_……共 %d 条，仅列前 40。_" % len(flagged), ""]

    # ---- 五、规则异常 ----
    if an.get("rules"):
        sec += 1
        L += ["## %d、规则异常" % sec, ""]
        summ = []
        blocks = []
        for ri, it in enumerate(as_list(an["rules"], "analysis.rules")):
            as_item_dict(it, "analysis.rules[%d]" % ri)
            check_keys(it, {"name", "when", "label", "show"}, "analysis.rules[%d]" % ri)
            nm = it.get("name") or it.get("when", "规则%d" % ri)
            when = it.get("when")
            if not when:
                raise UserError("analysis.rules[%d]「%s」缺少 when 表达式" % (ri, clip(nm, 40)))
            if not isinstance(when, str):
                raise UserError("analysis.rules[%d].when 必须是表达式字符串，实际是%s" % (ri, _tname(when)))
            nm = as_str(nm, "analysis.rules[%d].name" % ri)
            lab = as_str(it.get("label"), "analysis.rules[%d].label" % ri,
                         default="", allow_empty=True)
            labs = [x for x in ([lab] if lab else []) + d.key_vars + d.dim_vars if x in envs[0]]
            labs = list(OrderedDict.fromkeys(labs))[:3]
            show = [x for x in as_str_list(it.get("show"), "analysis.rules[%d].show" % ri)
                    if x in envs[0]]
            hit = [e for e in envs if eval_bool(when, e, "rules.%s" % nm)]
            summ.append([nm, "`%s`" % when, len(hit), _pct(len(hit), n),
                         "✅" if not hit else "⚠️"])
            rows = [[i] + [str(e.get(x, "")) for x in labs] +
                    [fmt_num(e[x]) if isinstance(e.get(x), (int, float)) else str(e.get(x, ""))
                     for x in show]
                    for i, e in enumerate(hit[:40], start=1)]
            blk = ["### %s" % nm, "", "规则：`%s`　命中 **%d** 条。" % (when, len(hit)), ""]
            blk += [md_table(["#"] + labs + show, rows) if rows else "_无命中。_", ""]
            if len(hit) > 40:
                blk += ["_……共 %d 条，仅列前 40。_" % len(hit), ""]
            blocks += blk
        L += [md_table(["规则", "表达式", "命中数", "占比", ""], summ), ""] + blocks

    # ---- 六、What-If ----
    if an.get("what_if"):
        sec += 1
        L += ["## %d、What-If 推演" % sec, ""]
        derived_spec = OrderedDict((k, v) for k, v in as_dict(spec.get("derived"), "derived").items())
        for wi, it in enumerate(as_list(an["what_if"], "analysis.what_if")):
            as_item_dict(it, "analysis.what_if[%d]" % wi)
            check_keys(it, {"name", "set", "recompute", "targets", "label"},
                       "analysis.what_if[%d]" % wi)
            nm = as_str(it.get("name"), "analysis.what_if[%d].name" % wi,
                        default="情景%d" % wi)
            sets = as_dict(it.get("set"), "analysis.what_if[%d].set" % wi)
            if not sets:
                raise UserError("analysis.what_if[%s] 缺少 set（要扰动的输入点）" % nm)
            recompute = as_str_list(it.get("recompute"), "analysis.what_if[%d].recompute" % wi)
            for rc in recompute:
                if rc not in derived_spec:
                    raise UserError("analysis.what_if[%s].recompute 中的 %s 不是 spec.derived 里的变量。"
                                    "\n    可用中间量：%s"
                                    % (nm, rc, "、".join(derived_spec) or "(spec 未定义 derived)"))
            # 同一个中间量既被 set 扰动、又被 recompute 重算 —— 重算会按原公式把它算回去，
            # 你刚扰动的值被静默抹掉，报告里每行「变化」都是 0.00：一份格式完好、
            # 结论完全是假的 What-If。
            both = [v for v in sets if v in recompute]
            if both:
                raise UserError(
                    "analysis.what_if[%s] 里 %s 既出现在 `set`（要扰动）又出现在 `recompute`（要重算）。\n"
                    "    执行顺序是「先按 set 改值 → 再按 derived 公式重算 recompute 里的变量」，"
                    "所以重算会把你刚改的值按原公式算回去，推演出来必然是「毫无变化」。\n"
                    "    二选一：想直接指定它的新值就从 recompute 里去掉；"
                    "想让它跟着上游变就改成扰动上游的输入字段。"
                    % (nm, "、".join(both)))
            targets = as_list(it.get("targets"), "analysis.what_if[%d].targets" % wi)
            if not targets:
                raise UserError("analysis.what_if[%s] 缺少 targets" % nm)
            for tj, t in enumerate(targets):
                as_item_dict(t, "analysis.what_if[%d].targets[%d]" % (wi, tj))
                check_keys(t, {"name", "expr"}, "analysis.what_if[%d].targets[%d]" % (wi, tj))
                if not t.get("name") or not t.get("expr"):
                    raise UserError("analysis.what_if[%s].targets 每项需要 name 与 expr" % nm)
                as_str(t["name"], "analysis.what_if[%d].targets[%d].name" % (wi, tj))
                as_str(t["expr"], "analysis.what_if[%d].targets[%d].expr" % (wi, tj))
            if len({t["name"] for t in targets}) != len(targets):
                raise UserError("analysis.what_if[%s].targets 里存在同名目标，请改名。" % nm)
            # 扰动会顺着 derived 链往下传，但脚本只重算 recompute 里点名的那些。
            # 漏掉链条中间任何一环，下游 target 取到的就是**扰动前的陈旧值**，
            # 报告会一本正经地告诉你"这个情景毫无影响"。这里把整条链算出来对一遍。
            dirty = set(sets) | set(recompute)
            stale = OrderedDict()
            for dv, dexpr in derived_spec.items():
                if dv in recompute:
                    continue
                if set(SAFE.names(dexpr, "derived.%s" % dv)) & dirty:
                    stale[dv] = dexpr
                    dirty.add(dv)      # 它的真实值本该变，下游同样受影响
            if stale:
                hit = OrderedDict()
                for t in targets:
                    for x in SAFE.names(t["expr"], "what_if.targets.%s" % t["name"]):
                        if x in stale:
                            hit.setdefault(x, []).append(t["name"])
                if hit:
                    raise UserError(
                        "analysis.what_if[%s]：中间量 %s 依赖了被扰动的 %s，却没被列进 `recompute`，"
                        "而 targets（%s）又用到了它。\n"
                        "    结果是 targets 拿着**扰动前的陈旧值**去算，报告会告诉你"
                        "「这个情景毫无影响」——数字全是 0，看着正常其实什么都没推演。\n"
                        "    修复：`\"recompute\": [%s]`（按 derived 的声明顺序把整条链都列上）。"
                        % (nm, "、".join(hit), "、".join(sorted(sets)),
                           "、".join(sorted({n2 for v2 in hit.values() for n2 in v2})),
                           ", ".join('"%s"' % x for x in derived_spec
                                     if x in recompute or x in stale)))
            # 扰动点必须是数值变量。把文本列（键/维度）写进 set 会静默跑通：
            # "张三" * 2 = "张三张三"，转数值得 0.0，于是报告里每行"变化"都是
            # 0.00 —— 一份格式完好、结论却完全是假的 What-If。宁可报错也不能这样。
            text_vars = set(d.key_vars) | set(d.dim_vars)
            for var in sets:
                if var in text_vars:
                    raise UserError(
                        "analysis.what_if[%s].set 的扰动点 %s 是文本列（键/维度），不能做数值推演。\n"
                        "    文本参与运算会被折成 0.0，跑出来的「变化」全是 0，"
                        "看着正常其实什么都没推演。\n"
                        "    可用的数值扰动点：%s"
                        % (nm, var, "、".join(d.field_vars + d.derived_vars) or "(无)"))
            lab = as_str(it.get("label"), "analysis.what_if[%d].label" % wi,
                         default="", allow_empty=True)
            labs = [x for x in ([lab] if lab else []) + d.key_vars + d.dim_vars if x in envs[0]]
            labs = list(OrderedDict.fromkeys(labs))[:2]

            per_row = []
            tot = OrderedDict((t["name"], [0.0, 0.0]) for t in targets)
            for e in envs:
                base = dict(e)
                new = {}
                for var, expr in sets.items():
                    if var not in base:
                        raise UserError("analysis.what_if[%s].set 的 %s 不在明细列中" % (nm, var))
                    new[var] = eval_num(expr, base, "what_if.%s.set.%s" % (nm, var))
                sc = dict(base)
                sc.update(new)
                # 按 spec.derived 的声明顺序重算被指定的中间量（保证依赖顺序）
                for dv, dexpr in derived_spec.items():
                    if dv in recompute:
                        sc[dv] = eval_num(dexpr, sc, "what_if.%s.recompute.%s" % (nm, dv))
                rec = {"labs": [str(e.get(x, "")) for x in labs], "t": OrderedDict()}
                for t in targets:
                    b0 = eval_num(t["expr"], base, "what_if.%s.targets.%s(现状)" % (nm, t["name"]))
                    s0 = eval_num(t["expr"], sc, "what_if.%s.targets.%s(情景)" % (nm, t["name"]))
                    rec["t"][t["name"]] = (round(b0, 2), round(s0, 2), round(s0 - b0, 2))
                    tot[t["name"]][0] += b0
                    tot[t["name"]][1] += s0
                per_row.append(rec)

            L += ["### %s" % nm, "",
                  "扰动：%s" % "；".join("`%s → %s`" % (k, v) for k, v in sets.items()), ""]
            if recompute:
                L += ["重算中间量：%s" % "、".join("`%s`" % r for r in recompute), ""]
            trows = []
            for t in targets:
                b0, s0 = tot[t["name"]]
                trows.append([t["name"], "`%s`" % t["expr"], fmt_num(b0), fmt_num(s0),
                              fmt_num(s0 - b0), _pct(s0 - b0, abs(b0)) if b0 else "-"])
            L += ["**合计影响：**", "",
                  md_table(["目标", "表达式", "现状合计", "情景合计", "变化", "变化率"], trows), ""]

            first = targets[0]["name"]
            per_row.sort(key=lambda x: -abs(x["t"][first][2]))
            heads = ["#"] + labs
            for t in targets:
                heads += ["%s·现状" % t["name"], "%s·情景" % t["name"], "%s·变化" % t["name"]]
            rows = []
            for i, rec in enumerate(per_row[:25], start=1):
                row = [i] + rec["labs"]
                for t in targets:
                    a, bb, c = rec["t"][t["name"]]
                    row += [fmt_num(a), fmt_num(bb), fmt_num(c)]
                rows.append(row)
            trow = ["**合计**"] + [""] * len(labs)
            for t in targets:
                b0, s0 = tot[t["name"]]
                trow += ["**%s**" % fmt_num(b0), "**%s**" % fmt_num(s0), "**%s**" % fmt_num(s0 - b0)]
            rows.append(trow)
            L += ["**逐行影响（按「%s·变化」绝对值降序，前 25 行 + 合计）：**" % first, "",
                  md_table(heads, rows), ""]
            if len(per_row) > 25:
                L += ["_……共 %d 行，仅列前 25。合计行为全量口径。_" % len(per_row), ""]

    L += ["---", "",
          "> 以上是脚本给的骨架。**真正的洞察要在此基础上再写一段结论**：",
          "> 钱花在哪、哪些异常需要人去核、哪个情景对总成本的杠杆最大。", ""]

    fp = write_text(os.path.join(outdir, "分析报告.md"), "\n".join(L))
    print("[analyze] 分析报告 → %s（%d 行数据，%d 个分析小节）" % (fp, n, sec))
    return 0


# ===========================================================================
# 十、CLI
# ===========================================================================

EPILOG = """\
五步法（详见 excel-ai-analyst 技能的 SKILL.md）：
  1 结构化转 MD  tomd      ← 脚本
  2 字段本体                ← AI 手写
  3 公式链与血缘            ← AI 手写
  4 全量数据验证 verify     ← 脚本（核心，跳过验证直接分析 = 没读懂代码就改代码）
  5 交付        output/analyze ← 脚本

典型流程：
  python3 excel_ai.py tomd    表A.xlsx 表B.xlsx -o ./01_raw_md
  # 读 MD，手写本体/公式链，写出 spec.json
  python3 excel_ai.py verify  spec.json -o ./03_verify
  python3 excel_ai.py output  spec.json -d ./03_verify/verification_detail.csv -o ./04_output
  python3 excel_ai.py analyze spec.json -d ./03_verify/verification_detail.csv -o ./04_output
"""


def build_parser():
    p = argparse.ArgumentParser(
        prog="excel_ai.py",
        description="Excel 智能分析工具（excel-ai-analyst 技能配套）：把含公式的业务 Excel "
                    "当作没有文档的遗留代码来逆向工程。",
        epilog=EPILOG, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("-V", "--version", action="version", version="excel_ai.py %s" % VERSION)
    sub = p.add_subparsers(dest="cmd", metavar="子命令")

    # ---- tomd ----
    p1 = sub.add_parser(
        "tomd", help="Step 1：Excel → 结构化 MD（多行表头/合并单元格/真实公式/列画像）",
        description="把 Excel 每个 Sheet 转成一份结构化 MD：表头区逐行解析、列字段定义表"
                    "（拼接列名/样例值/填充率/类型/真实公式）、真实公式清单、数据预览、告警；"
                    "并生成 00-索引.md。原表只读，绝不修改。",
        epilog="示例：\n"
               "  python3 excel_ai.py tomd 工资表.xlsx 绩效表.xlsx -o ./01_raw_md\n"
               "  python3 excel_ai.py tomd 工资表.xlsx --sheets 明细表 汇总表 --header-rows 4 -o ./01_raw_md\n",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p1.add_argument("files", nargs="+", help="一个或多个 .xlsx/.xlsm 路径")
    p1.add_argument("-o", "--output", default="./01_raw_md", help="输出目录（默认 ./01_raw_md）")
    p1.add_argument("--sheets", nargs="*", default=None, help="只转指定 Sheet（默认全部）")
    p1.add_argument("--header-scan", type=int, default=8, help="表头启发式扫描的行数上限（默认 8）")
    p1.add_argument("--header-rows", type=int, default=None,
                    help="强制指定表头行数（不指定则自动推断；推断值请自己确认一眼）")
    p1.add_argument("--preview-rows", type=int, default=10, help="数据预览行数（默认 10）")
    p1.add_argument("--summary-words", nargs="*", default=None,
                    help="汇总/小计行识别词（默认：%s）" % " ".join(DEFAULT_SUMMARY_WORDS))
    p1.set_defaults(func=cmd_tomd)

    # ---- verify ----
    p2 = sub.add_parser(
        "verify", help="Step 4：按 spec.json 用全量真实数据逐行回算并与 Excel 现值比对（核心）",
        description="读 spec.json（workbook/sheet/header_rows/keys/dimensions/fields/derived/"
                    "checks/skip_when/cross_checks），逐行回算 AI 理解的公式并与 Excel 现有结果列比对，"
                    "产出 验证报告.md + verification_detail.csv + mismatches.csv（均 utf-8-sig）。\n"
                    "表达式白名单：算术 + - * / ** %、比较、and/or/not、三元、"
                    "函数 abs round min max int float sum floor ceil；非数值单元格按 0.0 参与运算。",
        epilog="示例：\n  python3 excel_ai.py verify spec.json -o ./03_verify\n",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p2.add_argument("spec", help="spec.json 路径（AI 对业务逻辑的形式化理解）")
    p2.add_argument("-o", "--output", default="./03_verify", help="输出目录（默认 ./03_verify）")
    p2.set_defaults(func=cmd_verify)

    # ---- output ----
    p3 = sub.add_parser(
        "output", help="Step 5：产出带配色标注的 AI处理结果.xlsx",
        description="读验证明细 CSV，写 AI处理结果.xlsx：蓝 DAEEF3=原始输入、绿 E2EFDA=AI 计算、"
                    "橙 FCE4D6=差异非零（超出容差）；另附「校验汇总」「数据血缘」「字段本体」三个 Sheet "
                    "以及公式说明区 + 图例。血缘/本体内容取自 spec 的 lineage / ontology 段。",
        epilog="示例：\n"
               "  python3 excel_ai.py output spec.json -d ./03_verify/verification_detail.csv -o ./04_output\n",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p3.add_argument("spec", help="spec.json 路径")
    p3.add_argument("-d", "--detail", required=True, help="verify 产出的 verification_detail.csv")
    p3.add_argument("-o", "--output", default="./04_output", help="输出目录（默认 ./04_output）")
    p3.set_defaults(func=cmd_output)

    # ---- analyze ----
    p4 = sub.add_parser(
        "analyze", help="Step 5：产出业务分析报告（分组/分布/TopN/离群/规则/What-If）",
        description="读验证明细 CSV 与 spec.analysis，产出 分析报告.md：分组汇总（数量/总额/人均/占比）、"
                    "取值分布、Top/Bottom、离群（IQR k / Z-score z）、规则异常（表达式）、"
                    "What-If（set 改动 → recompute 重算中间量 → 对 targets 出「现状/情景/变化」三组列 + 合计行）。",
        epilog="示例：\n"
               "  python3 excel_ai.py analyze spec.json -d ./03_verify/verification_detail.csv -o ./04_output\n",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p4.add_argument("spec", help="spec.json 路径（需含 analysis 段）")
    p4.add_argument("-d", "--detail", required=True, help="verify 产出的 verification_detail.csv")
    p4.add_argument("-o", "--output", default="./04_output", help="输出目录（默认 ./04_output）")
    p4.set_defaults(func=cmd_analyze)

    return p


def main(argv=None):
    parser = build_parser()
    args = parser.parse_args(argv)
    if not getattr(args, "cmd", None):
        parser.print_help()
        return 1
    if args.cmd == "tomd" and args.summary_words is None:
        args.summary_words = DEFAULT_SUMMARY_WORDS
    try:
        return args.func(args)
    except UserError as e:
        sys.stderr.write("\n[错误] %s\n\n" % e)
        return 1
    except KeyboardInterrupt:
        sys.stderr.write("\n[中断] 用户取消。\n")
        return 130
    except FileNotFoundError as e:
        sys.stderr.write("\n[错误] 找不到文件：%s\n\n" % e)
        return 1
    except PermissionError as e:
        sys.stderr.write("\n[错误] 没有权限读写：%s\n\n" % e)
        return 1
    except OSError as e:
        # 磁盘满、路径过长、文件名非法、-o 指到文件上……一律给中文，别吐 traceback
        sys.stderr.write("\n[错误] 文件系统操作失败：%s\n\n" % e)
        return 1


if __name__ == "__main__":
    sys.exit(main())
