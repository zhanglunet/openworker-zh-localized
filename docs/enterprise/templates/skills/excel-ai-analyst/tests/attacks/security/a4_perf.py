#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""A4：性能——1 万行全量验证必须在 60 秒内跑完（含 cross_checks 与 output/analyze）。

顺带压两件事：
  * 汇总行必须被 skip_when 排掉，不能被算进金额（"汇总行被算进金额"属于静默错误）
  * 明细 CSV 行数必须等于真实数据行数
"""
import json
import os
import shutil
import subprocess
import sys
import time

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
SCRIPT = os.path.join(ROOT, "scripts", "excel_ai.py")
WORK = os.path.join(HERE, "_work_a4")
N = 10000
BUDGET = 60.0


def build(path, n, cross=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工资明细表"
    ws["A1"] = "大表 %d 行压测" % n
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    heads = ["工号", "姓名", "一级部门", "基本工资A", "绩效工资B1", "社保C1", "补贴D", "应发工资G"]
    for j, h in enumerate(heads, start=1):
        ws.cell(row=2, column=j, value=h)
    depts = ["研发中心", "财务部", "市场部", "供应链", "人力资源"]
    tot = 0.0
    for i in range(n):
        r = 3 + i
        a, b, c, d = 8000 + i % 900, 1500 + i % 400, 900 + i % 60, 300 + i % 50
        g = round(a + b - c + d, 2)
        tot += g
        ws.cell(row=r, column=1, value="E%06d" % i)
        ws.cell(row=r, column=2, value="员工%06d" % i)
        ws.cell(row=r, column=3, value=depts[i % len(depts)])
        ws.cell(row=r, column=4, value=a)
        ws.cell(row=r, column=5, value=b)
        ws.cell(row=r, column=6, value=c)
        ws.cell(row=r, column=7, value=d)
        ws.cell(row=r, column=8, value=g)
    # 汇总行：不排除的话金额会翻倍
    r = 3 + n
    ws.cell(row=r, column=2, value="合计")
    ws.cell(row=r, column=8, value=round(tot, 2))
    wb.save(path)

    if cross:
        wb2 = openpyxl.Workbook()
        w2 = wb2.active
        w2.title = "月考月发"
        for j, h in enumerate(["工号", "绩效工资"], start=1):
            w2.cell(row=1, column=j, value=h)
        for i in range(n):
            w2.cell(row=2 + i, column=1, value="E%06d" % i)
            w2.cell(row=2 + i, column=2, value=1500 + i % 400)
        wb2.save(path.replace(".xlsx", "_绩效.xlsx"))
    return round(tot, 2)


def main():
    if os.path.isdir(WORK):
        shutil.rmtree(WORK)
    os.makedirs(WORK)
    xlsx = os.path.join(WORK, "大表.xlsx")
    print("构造 %d 行测试表…" % N)
    t0 = time.time()
    expect_total = build(xlsx, N, cross=True)
    print("  构造耗时 %.1fs，Excel 侧应发合计 = %.2f" % (time.time() - t0, expect_total))

    spec = {
        "workbook": xlsx, "sheet": "工资明细表", "header_rows": 2,
        "keys": {"ID": "工号", "NAME": "姓名"},
        "dimensions": {"DEPT": "一级部门"},
        "fields": {"A": "基本工资A", "B1": "绩效工资B1", "C1": "社保C1",
                   "D": "补贴D", "G_x": "应发工资G"},
        "derived": {"A_ALL": "round(A, 2)", "B_ALL": "round(B1, 2)"},
        "checks": [
            {"name": "应发G", "target": "G_x", "expr": "round(A + B1 - C1 + D, 2)", "tolerance": 0.01},
            {"name": "应发G分层", "target": "G_x", "expr": "round(A_ALL + B_ALL - C1 + D, 2)",
             "tolerance": 0.01},
        ],
        "skip_when": {"empty": ["ID"], "label_in": ["汇总", "合计", "总计"]},
        "cross_checks": [{"name": "绩效传递", "workbook": xlsx.replace(".xlsx", "_绩效.xlsx"),
                          "sheet": "月考月发", "header_rows": 1, "key": "工号", "left_key": "ID",
                          "compare": [{"left": "B1", "right": "绩效工资", "tolerance": 0.001}]}],
        "analysis": {
            "group_by": [{"dim": "DEPT", "metrics": ["G_x", "A"], "count_as": "人数"}],
            "distributions": [{"dim": "DEPT"}, {"dim": "A", "bins": [8000, 8500, 9000]}],
            "top_n": [{"field": "G_x", "n": 5, "label": "NAME"}],
            "outliers": [{"field": "G_x", "method": "iqr", "k": 1.5, "label": "NAME"}],
            "rules": [{"name": "社保为零", "when": "C1 == 0", "label": "NAME"}],
            "what_if": [{"name": "基本工资上调10%", "set": {"A": "A*1.1"},
                         "recompute": ["A_ALL"],
                         "targets": [{"name": "应发G", "expr": "round(A + B1 - C1 + D, 2)"}],
                         "label": "NAME"}],
        },
    }
    sp = os.path.join(WORK, "spec.json")
    with open(sp, "w", encoding="utf-8") as f:
        json.dump(spec, f, ensure_ascii=False, indent=1)

    v_out = os.path.join(WORK, "03_verify")
    o_out = os.path.join(WORK, "04_output")
    broken = []
    timings = []

    for tag, args, budget in (
        ("verify（含 cross_checks）", ["verify", sp, "-o", v_out], BUDGET),
        ("output", ["output", sp, "-d", os.path.join(v_out, "verification_detail.csv"),
                    "-o", o_out], BUDGET),
        ("analyze（含 What-If）", ["analyze", sp, "-d", os.path.join(v_out, "verification_detail.csv"),
                                 "-o", o_out], BUDGET),
        ("tomd", ["tomd", xlsx, "-o", os.path.join(WORK, "01_md")], BUDGET * 2),
    ):
        t0 = time.time()
        p = subprocess.run([sys.executable, SCRIPT] + args, capture_output=True, timeout=1800)
        dt = time.time() - t0
        err = p.stderr.decode("utf-8", "replace")
        out = p.stdout.decode("utf-8", "replace")
        status = "PASS" if (p.returncode == 0 and dt <= budget and "Traceback" not in err) else "FAIL"
        timings.append((tag, dt, budget, status))
        print("  [%s] %-24s %6.1fs / 预算 %.0fs（退出码 %d）" % (status, tag, dt, budget, p.returncode))
        if p.returncode != 0 or "Traceback" in err:
            broken.append((tag, "退出码 %d %s" % (p.returncode, err.strip().splitlines()[-1][:90] if err.strip() else "")))
        elif dt > budget:
            broken.append((tag, "耗时 %.1fs 超预算 %.0fs" % (dt, budget)))
        if tag.startswith("verify"):
            print("     " + "\n     ".join(out.strip().splitlines()[:4]))

    # ---- 正确性：汇总行不得被算进金额 ----
    print()
    print("正确性核对：")
    rep = os.path.join(v_out, "验证报告.md")
    if os.path.isfile(rep):
        txt = open(rep, encoding="utf-8").read()
        import re
        m = re.search(r"参与比对的数据行 \| (\d+)", txt)
        got = int(m.group(1)) if m else -1
        if got != N:
            broken.append(("参与比对行数", "期望 %d，实际 %d（汇总行被算进来了？）" % (N, got)))
            print("  [FAIL] 参与比对的数据行 = %s，期望 %d" % (got, N))
        else:
            print("  [PASS] 参与比对的数据行 = %d（汇总行已排除）" % got)
        m = re.search(r"\| 应发G \| ([\d,\.]+) \| ([\d,\.]+) \|", txt)
        if m:
            ai = float(m.group(1).replace(",", ""))
            ex = float(m.group(2).replace(",", ""))
            if abs(ex - expect_total) > 0.05:
                broken.append(("Excel 合计", "报告 %.2f vs 真实 %.2f（汇总行混入 → 金额翻倍）"
                               % (ex, expect_total)))
                print("  [FAIL] Excel 合计 %.2f ≠ 真实 %.2f" % (ex, expect_total))
            else:
                print("  [PASS] Excel 合计 %.2f = 真实 %.2f" % (ex, expect_total))
            if abs(ai - ex) > 0.05:
                broken.append(("AI 合计", "%.2f vs %.2f" % (ai, ex)))
        if "100.00%" not in txt:
            broken.append(("通过率", "构造的表本身完全自洽，通过率却不是 100%"))
            print("  [FAIL] 通过率不是 100%（构造的数据是自洽的）")
        else:
            print("  [PASS] 通过率 100%")

    d = os.path.join(v_out, "verification_detail.csv")
    if os.path.isfile(d):
        nl = sum(1 for _ in open(d, encoding="utf-8-sig")) - 1
        print("  明细 CSV 数据行 = %d（期望 %d）" % (nl, N))
        if nl != N:
            broken.append(("明细行数", "%d ≠ %d" % (nl, N)))

    print()
    print("=" * 80)
    print("A4 小结：攻破 %d 项" % len(broken))
    for t, dd in broken:
        print("  * %s：%s" % (t, dd))
    return 1 if broken else 0


if __name__ == "__main__":
    sys.exit(main())
