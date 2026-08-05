#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
spec.json 鲁棒性攻击夹具生成器。

生成四个工作簿（都写到本目录的 fx/ 下，绝不碰别人的 fixtures/）：

  B01_单行表头.xlsx     最普通的一张工资表，1 行表头，8 行数据，G = A + B - C
  B02_两行表头.xlsx     大标题 + 字段名 两行表头 —— 用来攻击 "header_rows 数错一行"
  B03_右表.xlsx         跨表校验用的右表（绩效表）
  B04_万行.xlsx         10000 行 × 12 列 —— 性能攻击（verify 必须 60 秒内跑完）
  B05_含汇总行.xlsx     数据体里混了合计行 —— 攻击 skip_when 拼错时的静默翻倍
"""

import os
import random

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
FX = os.path.join(HERE, "fx")


def _save(wb, name):
    os.makedirs(FX, exist_ok=True)
    p = os.path.join(FX, name)
    wb.save(p)
    print("[fx] %s" % p)
    return p


def b01():
    """1 行表头的标准表：工号 姓名 部门 A B C 应发G。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工资明细表"
    ws.append(["工号", "姓名", "一级部门", "基本工资A", "绩效B", "扣款C", "应发工资G"])
    rows = [
        ("A001", "张三", "研发部", 12000, 3000, 1500, 13500),
        ("A002", "李四", "研发部", 9000, 2500, 1200, 10300),
        ("A003", "王五", "市场部", 8000, 4000, 1100, 10900),
        ("A004", "赵六", "市场部", 7500, 1500, 900, 8100),
        ("A005", "钱七", "财务部", 11000, 2000, 1400, 11600),
        ("A006", "孙八", "财务部", 6800, 1200, 800, 7200),
        ("A007", "周九", "研发部", 15000, 5000, 2000, 18000),
        ("A008", "吴十", "市场部", 5200, 800, 600, 5400),
    ]
    for r in rows:
        ws.append(list(r))
    return _save(wb, "B01_单行表头.xlsx")


def b02():
    """两行表头（大标题 + 字段名）。正确 header_rows=2，spec 写 1 就会把字段名行当数据。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工资明细表"
    ws.append(["某某公司2026年6月工资表"] + [None] * 6)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.append(["工号", "姓名", "一级部门", "基本工资A", "绩效B", "扣款C", "应发工资G"])
    rows = [
        ("A001", "张三", "研发部", 12000, 3000, 1500, 13500),
        ("A002", "李四", "研发部", 9000, 2500, 1200, 10300),
        ("A003", "王五", "市场部", 8000, 4000, 1100, 10900),
        ("A004", "赵六", "市场部", 7500, 1500, 900, 8100),
        ("A005", "钱七", "财务部", 11000, 2000, 1400, 11600),
    ]
    for r in rows:
        ws.append(list(r))
    return _save(wb, "B02_两行表头.xlsx")


def b03():
    """跨表校验的右表。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "绩效表"
    ws.append(["工号", "姓名", "绩效系数"])
    for i, (k, n, c) in enumerate([("A001", "张三", 1.2), ("A002", "李四", 1.0),
                                   ("A003", "王五", 1.5), ("A004", "赵六", 0.8)]):
        ws.append([k, n, c])
    ws2 = wb.create_sheet("另一个表")
    ws2.append(["占位"])
    return _save(wb, "B03_右表.xlsx")


def b04(n=10000):
    """性能攻击：10000 行 × 12 列。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "大表"
    heads = ["工号", "姓名", "一级部门", "A1", "A2", "A3", "B1", "B2", "C1", "C2", "D", "应发G"]
    ws.append(heads)
    rnd = random.Random(20260805)
    depts = ["研发部", "市场部", "财务部", "生产部", "行政部"]
    for i in range(n):
        a1 = rnd.randint(5000, 20000)
        a2 = rnd.randint(0, 3000)
        a3 = rnd.randint(0, 1500)
        b1 = rnd.randint(0, 8000)
        b2 = rnd.randint(0, 2000)
        c1 = rnd.randint(500, 2500)
        c2 = rnd.randint(300, 1800)
        dd = rnd.choice([0, 200, 500])
        g = a1 + a2 + a3 + b1 + b2 - c1 - c2 + dd
        ws.append(["E%05d" % i, "员工%05d" % i, depts[i % len(depts)],
                   a1, a2, a3, b1, b2, c1, c2, dd, g])
    return _save(wb, "B04_万行.xlsx")


def b05():
    """数据体里混了部门小计与总计行。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工资明细表"
    ws.append(["工号", "姓名", "一级部门", "基本工资A", "绩效B", "扣款C", "应发工资G"])
    body = [
        ("A001", "张三", "研发部", 12000, 3000, 1500, 13500),
        ("A002", "李四", "研发部", 9000, 2500, 1200, 10300),
        ("", "研发部小计", "", 21000, 5500, 2700, 23800),
        ("A003", "王五", "市场部", 8000, 4000, 1100, 10900),
        ("A004", "赵六", "市场部", 7500, 1500, 900, 8100),
        ("", "市场部小计", "", 15500, 5500, 2000, 19000),
        ("", "合计", "", 36500, 11000, 4700, 42800),
    ]
    for r in body:
        ws.append(list(r))
    return _save(wb, "B05_含汇总行.xlsx")


if __name__ == "__main__":
    b01()
    b02()
    b03()
    b04()
    b05()
