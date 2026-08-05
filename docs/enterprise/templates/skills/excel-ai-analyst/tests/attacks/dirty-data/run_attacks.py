#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
脏数据 / 边界对抗测试。目标：把 scripts/excel_ai.py 弄崩。

判定标准（任一命中即算"攻破"）：
  1. 未捕获的 traceback（应当是清晰的中文错误 + 非零退出码）
  2. 错误的静默通过（把不匹配算成匹配、汇总行被算进金额、绑错列还报 100%）
  3. 明显不合理的结果（数值错误、报告为空、产物缺失）
  4. 性能不可接受（1 万行验证超过 60 秒）

用法：
    python3 make_attack_fixtures.py     # 先生成夹具
    python3 run_attacks.py              # 再跑攻击
"""

import json
import os
import shutil
import subprocess
import sys
import time

HERE = os.path.dirname(os.path.abspath(__file__))
FX = os.path.join(HERE, "fx")
RUN = os.path.join(HERE, "run")
SCRIPT = os.path.abspath(os.path.join(HERE, "..", "..", "scripts", "excel_ai.py"))

RESULTS = []          # (编号, 名称, 是否通过, 说明)
PERF = []


# ---------------------------------------------------------------- 基础设施
def sh(args, timeout=300):
    """跑一次子命令，返回 (rc, stdout, stderr, 秒)。"""
    t0 = time.time()
    p = subprocess.run([sys.executable, SCRIPT] + args, capture_output=True,
                       text=True, timeout=timeout)
    return p.returncode, p.stdout, p.stderr, time.time() - t0


def spec(name, obj):
    d = os.path.join(RUN, "specs")
    os.makedirs(d, exist_ok=True)
    p = os.path.join(d, name + ".json")
    with open(p, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=1)
    return p


def outdir(name):
    p = os.path.join(RUN, name)
    if os.path.isdir(p):
        shutil.rmtree(p)
    return p


def record(no, title, ok, note=""):
    RESULTS.append((no, title, ok, note))
    print("  [%s] %-46s %s" % ("守住" if ok else "攻破", title, note))


def expect_clean_error(no, title, args, must_contain=(), timeout=300):
    """期望：非零退出 + 中文错误 + 没有 traceback。"""
    rc, out, err, _ = sh(args, timeout)
    if "Traceback" in err:
        return record(no, title, False, "出现未捕获 traceback：%s" % err.strip().splitlines()[-1][:90])
    if rc == 0:
        return record(no, title, False, "本该报错却 rc=0")
    blob = out + err
    miss = [k for k in must_contain if k not in blob]
    if miss:
        return record(no, title, False, "错误信息缺关键词 %s：%s" % (miss, blob.strip()[-120:]))
    record(no, title, True, "rc=%d 中文报错" % rc)


def expect_ok(no, title, args, timeout=300):
    rc, out, err, sec = sh(args, timeout)
    if "Traceback" in err:
        return record(no, title, False, "traceback：%s" % err.strip().splitlines()[-1][:90])
    if rc != 0:
        return record(no, title, False, "rc=%d：%s" % (rc, (out + err).strip()[-120:]))
    record(no, title, True, "%.1fs" % sec)
    return out + err


def fx(n):
    return os.path.join(FX, n)


def base_spec(book, sheet, **kw):
    s = {"workbook": fx(book), "sheet": sheet, "header_rows": 1}
    s.update(kw)
    return s


def read(p):
    with open(p, "r", encoding="utf-8-sig") as f:
        return f.read()


# =========================================================================
# 一、tomd —— 脏表结构
# =========================================================================
def group_tomd():
    print("\n【一】tomd：脏表结构")

    # A1 空表 / 只有空格的表：应跳过而不是崩
    o = outdir("t_空表")
    res = expect_ok("A1", "空表 / 只有空格的 Sheet 不崩", ["tomd", fx("A01_空表.xlsx"), "-o", o])
    if res is not None and "为空表" not in res:
        record("A1b", "空表应给出跳过提示", False, "没有提示")
    else:
        record("A1b", "空表应给出跳过提示", True)

    # A2 单行表 / 单列表
    o = outdir("t_单行")
    expect_ok("A2", "只有表头 / 只有一行 / 单列表", ["tomd", fx("A02_单行表.xlsx"), "-o", o])
    md = read(os.path.join(o, "A02_单行表", "只有一行数据.md"))
    ok = "推断表头行数：**1**" in md and "E001" in md
    record("A2b", "单行数据表：表头 1 行 + 数据体保留", ok,
           "" if ok else "表头行数或数据体不对")

    # A3 全空列 + 中间夹全空行
    o = outdir("t_全空列")
    expect_ok("A3", "全空列 + 中间全空行", ["tomd", fx("A03_全空列.xlsx"), "-o", o])
    md = read(os.path.join(o, "A03_全空列", "含全空列.md"))
    ok = "空列" in md and "数据体行数：**8**" in md
    record("A3b", "全空列被点名 + 空行不计入数据体", ok, "" if ok else "统计不对")

    # A4 超长列名（2400 字 + 32000 字）
    o = outdir("t_超长列名")
    expect_ok("A4", "超长列名（3.2 万字）不崩", ["tomd", fx("A04_超长列名.xlsx"), "-o", o])

    # A5 重复列名
    o = outdir("t_重复列名")
    expect_ok("A5", "重复列名（两个「金额」）不崩", ["tomd", fx("A05_重复列名.xlsx"), "-o", o])

    # A6 中文数字混排
    o = outdir("t_中文数字")
    expect_ok("A6", "中文数字/全角/百分号/括号负数", ["tomd", fx("A06_中文数字混排.xlsx"), "-o", o])

    # A7 日期列 —— 曾把整个数据体误判成表头
    o = outdir("t_日期列")
    expect_ok("A7", "日期列不崩", ["tomd", fx("A07_日期列.xlsx"), "-o", o])
    md = read(os.path.join(o, "A07_日期列", "日期列.md"))
    ok = "推断表头行数：**1**" in md
    record("A7b", "日期列不得被当成表头（数据体起点=1）", ok,
           "" if ok else "表头行数推断错了，整个数据体被吞")
    ok2 = "日期/时间" in md
    record("A7c", "日期列的类型标成「日期/时间」而非「空列」", ok2,
           "" if ok2 else "类型判定错")

    # A8 科学计数法
    o = outdir("t_科学计数")
    expect_ok("A8", "科学计数法（1e18 / 1.5e-12）", ["tomd", fx("A08_科学计数法.xlsx"), "-o", o])
    md = read(os.path.join(o, "A08_科学计数法", "科学计数法.md"))
    ok = "1.5e-12" in md
    record("A8b", "极小数不得被显示成 0", ok, "" if ok else "1.5e-12 被抹成了 0，等于凭空造数据")

    # A9 负数
    o = outdir("t_负数")
    expect_ok("A9", "负数 / 负扣款", ["tomd", fx("A09_负数.xlsx"), "-o", o])

    # A14 Sheet 名 safe_name 撞车
    o = outdir("t_名字冲突")
    expect_ok("A14", "Sheet 名「报表.」「报表」「 报表 」撞车", ["tomd", fx("A14_Sheet名冲突.xlsx"), "-o", o])
    files = os.listdir(os.path.join(o, "A14_Sheet名冲突"))
    ok = len(files) == 3
    record("A14b", "3 个 Sheet 必须产出 3 份 MD（不得静默覆盖）", ok,
           "" if ok else "只产出了 %d 份：%s" % (len(files), files))

    # A15 特殊文本
    o = outdir("t_特殊文本")
    expect_ok("A15", "竖线/换行/制表符/前导零/伪公式文本", ["tomd", fx("A15_特殊文本.xlsx"), "-o", o])

    # A16 .xls 与非表格文件的优雅报错
    expect_clean_error("A16", ".xls 旧格式 → 中文引导（tomd）",
                       ["tomd", fx("A12_真旧格式.xls"), "-o", outdir("t_xls")],
                       must_contain=["旧版 .xls", "soffice"])
    expect_clean_error("A17", "xlsx 内容却叫 .xls → 中文引导",
                       ["tomd", fx("A12_其实是xlsx.xls"), "-o", outdir("t_xls2")],
                       must_contain=["旧版 .xls"])
    expect_clean_error("A18", "纯文本改名成 .xlsx → 中文引导",
                       ["tomd", fx("A12_不是表格.xlsx"), "-o", outdir("t_notxlsx")],
                       must_contain=["不是一个有效的 xlsx"])
    expect_clean_error("A19", "--header-rows 负数被拒",
                       ["tomd", fx("A09_负数.xlsx"), "-o", outdir("t_hrneg"), "--header-rows", "-1"],
                       must_contain=["不能是负数"])

    # A20 --header-rows 0 必须被当成"无表头"，而不是回退自动推断
    o = outdir("t_hr0")
    expect_ok("A20", "--header-rows 0 生效", ["tomd", fx("A09_负数.xlsx"), "-o", o, "--header-rows", "0"])
    md = read(os.path.join(o, "A09_负数", "负数.md"))
    ok = "推断表头行数：**0**" in md
    record("A20b", "--header-rows 0 不得被当成「没指定」", ok, "" if ok else "0 被忽略了")


# =========================================================================
# 二、verify —— spec 类型污染（这些以前全是 traceback）
# =========================================================================
def group_spec_poison():
    print("\n【二】verify：spec 类型污染")

    def B(**kw):
        d = {"keys": {"ID": "工号"}, "fields": {"G": "应发", "C": "扣款", "S": "实发"},
             "checks": [{"name": "实发", "target": "S", "expr": "G-C"}]}
        d.update(kw)
        return base_spec("A09_负数.xlsx", "负数", **d)

    cases = [
        ("B1", "checks[].target 指向键列（文本）",
         B(checks=[{"name": "x", "target": "ID", "expr": "G-C"}]), ["target", "fields"]),
        ("B2", "header_rows 写成中文「四」",
         base_spec("A09_负数.xlsx", "负数", keys={"ID": "工号"}, fields={"G": "应发"},
                   checks=[{"name": "x", "target": "G", "expr": "G"}]) | {"header_rows": "四"},
         ["header_rows", "整数"]),
        ("B3", "header_rows 为负数",
         B() | {"header_rows": -1}, ["header_rows", "不能小于"]),
        ("B4", "header_rows 为小数 2.5",
         B() | {"header_rows": 2.5}, ["header_rows", "整数"]),
        ("B5", "fields 写成数组",
         base_spec("A09_负数.xlsx", "负数", keys={"ID": "工号"}, fields=["应发"],
                   checks=[{"name": "x", "target": "G", "expr": "G"}]), ["fields", "对象"]),
        ("B6", "checks 写成对象",
         base_spec("A09_负数.xlsx", "负数", keys={"ID": "工号"}, fields={"G": "应发"},
                   checks={"a": {"name": "x", "target": "G", "expr": "G"}}), ["checks", "数组"]),
        ("B7", "checks 元素是字符串",
         base_spec("A09_负数.xlsx", "负数", keys={"ID": "工号"}, fields={"G": "应发"},
                   checks=["应发=G"]), ["checks[0]", "对象"]),
        ("B8", "tolerance 写成中文",
         B(checks=[{"name": "x", "target": "S", "expr": "G-C", "tolerance": "很小"}]),
         ["tolerance", "数值"]),
        ("B9", "tolerance 为负",
         B(checks=[{"name": "x", "target": "S", "expr": "G-C", "tolerance": -1}]),
         ["tolerance", "不能小于"]),
        ("B10", "derived 写成数组",
         B(derived=["G-C"]), ["derived", "对象"]),
        ("B11", "skip_when 写成字符串",
         B(skip_when="合计"), ["skip_when", "对象"]),
        ("B12", "skip_when.label_in 元素不是字符串",
         B(skip_when={"label_in": [123]}), ["label_in", "字符串"]),
        ("B13", "workbook 写成数字",
         B() | {"workbook": 123}, ["workbook", "字符串"]),
        ("B14", "cross_checks.compare 写成对象",
         B(cross_checks=[{"name": "c", "workbook": fx("A09_负数.xlsx"), "sheet": "负数",
                          "header_rows": 1, "key": "工号", "left_key": "ID",
                          "compare": {"left": "G", "right": "应发"}}]), ["compare", "数组"]),
        ("B15", "变量名占用保留列名 Excel行号",
         base_spec("A09_负数.xlsx", "负数", keys={"Excel行号": "工号"}, fields={"G": "应发"},
                   checks=[{"name": "x", "target": "G", "expr": "G"}]), ["保留列名"]),
        ("B16", "校验项名与字段名撞车",
         B(checks=[{"name": "G", "target": "G", "expr": "G"}]), ["重复", "改名"]),
        ("B17", "列引用是空字符串",
         base_spec("A09_负数.xlsx", "负数", keys={"ID": ""}, fields={"G": "应发"},
                   checks=[{"name": "x", "target": "G", "expr": "G"}]), ["空字符串"]),
    ]
    for no, title, sp, kws in cases:
        expect_clean_error(no, title, ["verify", spec(no, sp), "-o", outdir("v_" + no)], kws)

    # spec 顶层不是对象
    p = os.path.join(RUN, "specs")
    os.makedirs(p, exist_ok=True)
    bad = os.path.join(p, "B18.json")
    with open(bad, "w", encoding="utf-8") as f:
        f.write("[1, 2, 3]")
    expect_clean_error("B18", "spec 顶层是数组", ["verify", bad, "-o", outdir("v_B18")],
                       ["顶层必须是一个对象"])


# =========================================================================
# 三、verify —— 静默算错
# =========================================================================
def group_silent():
    print("\n【三】verify：静默算错（最危险的一类）")

    # C1 数字型列名 "1"/"01" 不得被当成列号绑到别的列
    sp = base_spec("A05_重复列名.xlsx", "重复列名", keys={"ID": "工号"},
                   fields={"X": "1", "Y": "01"},
                   checks=[{"name": "x", "target": "X", "expr": "X"}])
    o = outdir("v_C1")
    rc, out, err, _ = sh(["verify", spec("C1", sp), "-o", o])
    if "Traceback" in err:
        record("C1", "数字型列名 \"1\"/\"01\"", False, "traceback")
    elif rc == 0:
        rep = read(os.path.join(o, "验证报告.md"))
        # 列 4 的表头就是 "1"，列 5 是 "01"；绑对了才行
        ok = "| X | 字段 | 4 |" in rep and "| Y | 字段 | 5 |" in rep
        record("C1", "数字型列名 \"1\"/\"01\" 绑到同名列而非列号", ok,
               "" if ok else "绑错列却报 100% 通过")
    else:
        record("C1", "数字型列名 \"1\"/\"01\"", False, "rc=%d %s" % (rc, (out + err)[-120:]))

    # C2 列名里没有的纯数字串仍可当列号，但要有提示
    sp = base_spec("A09_负数.xlsx", "负数", keys={"ID": 0}, fields={"G": "1"},
                   checks=[{"name": "x", "target": "G", "expr": "G"}])
    rc, out, err, _ = sh(["verify", spec("C2", sp), "-o", outdir("v_C2")])
    ok = rc == 0 and "已按列号" in (out + err) and "Traceback" not in err
    record("C2", "列名里没有的 \"1\" 退回按列号 + 明确提示", ok, "" if ok else "无提示或崩了")

    # C3 汇总行未配 skip_when → 必须点名（金额翻倍）
    sp = base_spec("A13_汇总行.xlsx", "带汇总", keys={"ID": "工号"},
                   fields={"A": "基本工资", "B": "补贴", "G": "应发"},
                   checks=[{"name": "应发", "target": "G", "expr": "A+B"}])
    o = outdir("v_C3")
    rc, out, err, _ = sh(["verify", spec("C3", sp), "-o", o])
    rep = read(os.path.join(o, "验证报告.md")) if rc == 0 else ""
    ok = rc == 0 and "疑似未排除的汇总行" in rep and "汇总" in err
    record("C3", "未排除的汇总行必须报警（否则金额翻倍）", ok,
           "" if ok else "通过率 100% 却对翻倍的汇总额只字不提")

    # C4 配了 skip_when 之后不该再报警，且金额正确
    sp["skip_when"] = {"label_in": ["合计", "小计"]}
    o = outdir("v_C4")
    rc, out, err, _ = sh(["verify", spec("C4", sp), "-o", o])
    rep = read(os.path.join(o, "验证报告.md")) if rc == 0 else ""
    ok = rc == 0 and "疑似未排除的汇总行" not in rep and "16,500.00" in rep
    record("C4", "配了 skip_when 后汇总额回到 16500", ok, "" if ok else "金额或告警不对")

    # C5 skip_when.label_in 写成单个字符串，不得被拆成单字
    sp2 = dict(sp)
    sp2["skip_when"] = {"label_in": "合计"}
    o = outdir("v_C5")
    rc, out, err, _ = sh(["verify", spec("C5", sp2), "-o", o])
    ok = rc == 0 and "数据行 6" in (out + err)     # 只排掉「合计」，「小计」还在
    record("C5", "label_in 写成字符串按整词处理（不拆成单字）", ok,
           "" if ok else (out + err).strip()[-120:])

    # C6 中文数字混排的解析正确性（逐格核对）
    sp = base_spec("A06_中文数字混排.xlsx", "中文数字混排", keys={"ID": "工号"},
                   fields={"K": "绩效系数", "T": "补贴", "G": "应发"},
                   checks=[{"name": "应发", "target": "G", "expr": "round(K*T,2)"}])
    o = outdir("v_C6")
    rc, out, err, _ = sh(["verify", spec("C6", sp), "-o", o])
    detail = read(os.path.join(o, "verification_detail.csv")) if rc == 0 else ""
    want = {"E002": ("0.0", "-500.0"),     # "/" → 0；全角减号 －500 → -500
            "E003": ("0.8", "1234.0"),     # 全角数字 ０.８ / １２３４
            "E005": ("1.5", "-300.0"),     # 会计括号 (300) → -300
            "E007": ("-1.0", "3000.0")}    # U+2212 减号；"3 000" 空格千分位
    bad = []
    for line in detail.splitlines()[1:]:
        c = line.split(",")
        if len(c) > 4 and c[1] in want:
            if (c[2], c[3]) != want[c[1]]:
                bad.append("%s 得到 (%s,%s) 期望 %s" % (c[1], c[2], c[3], want[c[1]]))
    record("C6", "全角减号 －/− 与全角数字必须正确解析", rc == 0 and not bad,
           "；".join(bad) if bad else "")

    # C7 日期列绝不能悄悄按序列号参与金额运算
    sp = base_spec("A07_日期列.xlsx", "日期列", keys={"ID": "工号"},
                   fields={"D": "入职日期", "M": "金额"},
                   checks=[{"name": "日期按0", "target": "M", "expr": "M + D"}])
    o = outdir("v_C7")
    rc, out, err, _ = sh(["verify", spec("C7", sp), "-o", o])
    ok = rc == 0 and "不匹配    0" in (out + err)
    record("C7", "日期字段按 0.0 参与运算（不当序列号）", ok, "" if ok else (out + err)[-120:])

    # C8 全空列做 target：必须告警而不是静默 0
    sp = base_spec("A03_全空列.xlsx", "含全空列", keys={"ID": "工号"},
                   fields={"E": "备用列1", "A": "基本工资"},
                   checks=[{"name": "空列", "target": "E", "expr": "0"}])
    o = outdir("v_C8")
    rc, out, err, _ = sh(["verify", spec("C8", sp), "-o", o])
    rep = read(os.path.join(o, "验证报告.md")) if rc == 0 else ""
    ok = rc == 0 and "整列为空" in rep
    record("C8", "字段绑到全空列必须告警", ok, "" if ok else "无告警")

    # C9 无缓存值 xlsx：必须给出"结果不可信"的红牌
    sp = base_spec("A11_无缓存值.xlsx", "无缓存", keys={"ID": "工号"},
                   fields={"A": "基本工资", "B": "补贴", "G": "应发"},
                   checks=[{"name": "应发", "target": "G", "expr": "A+B"}])
    o = outdir("v_C9")
    rc, out, err, _ = sh(["verify", spec("C9", sp), "-o", o])
    rep = read(os.path.join(o, "验证报告.md")) if rc == 0 else ""
    ok = rc == 0 and "结果不可信" in rep and "没有缓存值" in err
    record("C9", "xlsx 无缓存值 → 报告红牌 + stderr 严重告警", ok, "" if ok else "红牌没出")

    # C10 空表 / 只有表头：清晰报错而不是产出空报告
    expect_clean_error("C10", "空 Sheet → 清晰报错",
                       ["verify", spec("C10", base_spec("A01_空表.xlsx", "全空", keys={"ID": 0},
                                                        fields={"G": 1},
                                                        checks=[{"name": "x", "target": "G", "expr": "G"}])),
                        "-o", outdir("v_C10")], ["为空表"])
    expect_clean_error("C11", "只有表头没有数据 → 清晰报错",
                       ["verify", spec("C11", base_spec("A02_单行表.xlsx", "只有表头", keys={"ID": "工号"},
                                                        fields={"G": "应发"},
                                                        checks=[{"name": "x", "target": "G", "expr": "G"}])),
                        "-o", outdir("v_C11")], ["header_rows"])

    # C12 重复列名必须报错并列出候选
    expect_clean_error("C12", "重复列名「金额」→ 报错列候选",
                       ["verify", spec("C12", base_spec("A05_重复列名.xlsx", "重复列名",
                                                        keys={"ID": "工号"}, fields={"M": "金额"},
                                                        checks=[{"name": "x", "target": "M", "expr": "M"}])),
                        "-o", outdir("v_C12")], ["精确命中 2 列", "候选"])

    # C13 超长列名找不到时报错信息不得刷屏
    rc, out, err, _ = sh(["verify", spec("C13", base_spec("A04_超长列名.xlsx", "超长列名",
                                                          keys={"ID": "工号"}, fields={"G": "查无此列"},
                                                          checks=[{"name": "x", "target": "G", "expr": "G"}])
                                          | {"header_rows": 2}),
                          "-o", outdir("v_C13")])
    ok = rc != 0 and "Traceback" not in err and len(err) < 4000
    record("C13", "超长列名的报错信息被截断（<4000 字）", ok, "stderr %d 字" % len(err))

    # C14 表达式字符串复制炸弹
    sp = base_spec("A09_负数.xlsx", "负数", keys={"ID": "工号"}, fields={"G": "应发"},
                   checks=[{"name": "x", "target": "G", "expr": "G if ID__raw*900000000 else G"}])
    expect_clean_error("C14", "表达式字符串复制炸弹被拒",
                       ["verify", spec("C14", sp), "-o", outdir("v_C14")],
                       ["拒绝求值"], timeout=120)

    # C15 危险表达式仍然被拒（安全边界回归）
    for i, ex in enumerate(["__import__('os').system('id')", "G.__class__", "open('/etc/passwd')",
                            "[].__class__.__base__", "G[0]"]):
        sp = base_spec("A09_负数.xlsx", "负数", keys={"ID": "工号"}, fields={"G": "应发"},
                       checks=[{"name": "x", "target": "G", "expr": ex}])
        expect_clean_error("C15.%d" % i, "危险表达式被拒：%s" % ex[:28],
                           ["verify", spec("C15_%d" % i, sp), "-o", outdir("v_C15_%d" % i)], [])


# =========================================================================
# 四、analyze / output —— 分析段污染
# =========================================================================
def group_analyze():
    print("\n【四】analyze / output：分析段污染")

    sp = base_spec("A10_超大表.xlsx", "万行明细",
                   keys={"ID": "工号", "NAME": "姓名"}, dimensions={"DEPT": "部门"},
                   fields={"A": "基本工资A", "B1": "绩效B1", "B2": "补贴B2",
                           "C1": "社保C1", "C2": "公积金C2", "G_x": "应发G"},
                   derived={"B": "B1+B2", "C": "C1+C2"},
                   checks=[{"name": "应发G", "target": "G_x", "expr": "round(A+B-C,2)"}],
                   skip_when={"empty": ["ID"], "label_in": ["合计"]})
    sp["header_rows"] = 2
    o = outdir("v_perf")
    rc, out, err, sec = sh(["verify", spec("D0", sp), "-o", o])
    PERF.append(("verify 1 万行", sec))
    ok = rc == 0 and sec < 60 and "数据行 10000" in out
    record("D0", "1 万行 verify < 60 秒", ok, "%.1fs" % sec)
    detail = os.path.join(o, "verification_detail.csv")

    def an(name, analysis):
        s = dict(sp)
        s["analysis"] = analysis
        return spec(name, s)

    cases = [
        ("D1", "outliers.field 指向文本列", {"outliers": [{"field": "NAME"}]}, ["键/维度列"]),
        ("D2", "top_n.field 指向维度列", {"top_n": [{"field": "DEPT"}]}, ["键/维度列"]),
        ("D3", "top_n.n 写成中文", {"top_n": [{"field": "A", "n": "五"}]}, ["n", "整数"]),
        ("D4", "top_n.n 为 0", {"top_n": [{"field": "A", "n": 0}]}, ["不能小于"]),
        ("D5", "distributions.bins 含非数值", {"distributions": [{"dim": "A", "bins": ["低", "高"]}]},
         ["bins", "数值"]),
        ("D6", "distributions.bins 未排序", {"distributions": [{"dim": "A", "bins": [100, 50]}]},
         ["从小到大"]),
        ("D7", "analysis 段写成数组", [{"group_by": 1}], ["analysis", "对象"]),
        ("D8", "group_by 元素是字符串", {"group_by": ["DEPT"]}, ["group_by[0]", "对象"]),
        ("D9", "group_by.metrics 指向文本列", {"group_by": [{"dim": "DEPT", "metrics": ["NAME"]}]},
         ["键/维度列"]),
        ("D10", "outliers.method 拼错", {"outliers": [{"field": "A", "method": "iqrr"}]},
         ["method", "iqr"]),
        ("D11", "outliers.k 写成中文", {"outliers": [{"field": "A", "k": "大"}]}, ["k", "数值"]),
        ("D12", "rules.when 不是字符串", {"rules": [{"name": "r", "when": 1}]}, ["when"]),
        ("D13", "rules.when 引用未定义变量", {"rules": [{"name": "r", "when": "ZZZ>0"}]},
         ["未定义的变量"]),
        ("D14", "what_if.recompute 不是 derived 变量",
         {"what_if": [{"name": "w", "set": {"A": "A*1.1"}, "recompute": ["QQ"],
                       "targets": [{"name": "t", "expr": "A"}]}]}, ["recompute"]),
        ("D15", "what_if.targets 为空",
         {"what_if": [{"name": "w", "set": {"A": "A*1.1"}, "targets": []}]}, ["targets"]),
        ("D16", "what_if.set 写成数组",
         {"what_if": [{"name": "w", "set": ["A"], "targets": [{"name": "t", "expr": "A"}]}]},
         ["set", "对象"]),
        ("D17", "what_if.targets 同名",
         {"what_if": [{"name": "w", "set": {"A": "A*1.1"},
                       "targets": [{"name": "t", "expr": "A"}, {"name": "t", "expr": "B"}]}]},
         ["同名目标"]),
    ]
    for no, title, a, kws in cases:
        expect_clean_error(no, title, ["analyze", an(no, a), "-d", detail, "-o", outdir("a_" + no)], kws)

    # D18 正常的分析段必须跑通且数值对得上
    full = {"group_by": [{"dim": "DEPT", "metrics": ["G_x"], "count_as": "人数"}],
            "distributions": [{"dim": "DEPT"}, {"dim": "A", "bins": [5000, 15000, 25000]}],
            "top_n": [{"field": "G_x", "n": 5, "label": "NAME"}],
            "outliers": [{"field": "G_x", "method": "iqr", "k": 1.5, "label": "NAME"},
                         {"field": "A", "method": "zscore", "z": 2.5}],
            "rules": [{"name": "补贴为零", "when": "B2 == 0", "label": "NAME", "show": ["B2"]}],
            "what_if": [{"name": "基本工资上调10%", "set": {"A": "A*1.1"}, "recompute": ["B", "C"],
                         "targets": [{"name": "应发G", "expr": "A+B-C"}], "label": "NAME"}]}
    o = outdir("a_full")
    rc, out, err, sec = sh(["analyze", an("D18", full), "-d", detail, "-o", o])
    PERF.append(("analyze 1 万行", sec))
    rep = read(os.path.join(o, "分析报告.md")) if rc == 0 else ""
    ok = rc == 0 and "Traceback" not in err and rep.count("##") >= 7 and "合计" in rep
    record("D18", "完整 analysis 段 6 个小节全部产出", ok, "%.1fs" % sec)

    # D19 output
    o = outdir("a_out")
    rc, out, err, sec = sh(["output", an("D19", full), "-d", detail, "-o", o])
    PERF.append(("output 1 万行", sec))
    xlsx = os.path.join(o, "AI处理结果.xlsx")
    ok = rc == 0 and "Traceback" not in err and os.path.isfile(xlsx) and os.path.getsize(xlsx) > 10000
    record("D19", "1 万行 output 产出结果表", ok, "%.1fs" % sec)

    # D20 明细 CSV 与 spec 不匹配
    sp2 = dict(sp)
    sp2["fields"] = dict(sp["fields"])
    sp2["fields"]["不存在的字段"] = "基本工资A"
    expect_clean_error("D20", "明细 CSV 与 spec 不匹配 → 清晰报错",
                       ["analyze", spec("D20", sp2), "-d", detail, "-o", outdir("a_D20")],
                       ["缺少列", "重跑 verify"])

    # D21 明细 CSV 被手工塞了同名列
    bad = os.path.join(RUN, "bad_detail.csv")
    src = read(detail).splitlines()
    with open(bad, "w", encoding="utf-8-sig") as f:
        f.write(src[0] + ",A\n")
        for ln in src[1:6]:
            f.write(ln + ",1\n")
    expect_clean_error("D21", "明细 CSV 有同名列 → 拒绝分析",
                       ["analyze", spec("D21", sp), "-d", bad, "-o", outdir("a_D21")],
                       ["同名列"])


# =========================================================================
# 五、第二波：更刁钻的现实脏表
# =========================================================================
def group_wave2():
    print("\n【五】第二波：日期表头 / 全零行 / 数据区合并 / 除零 / 公式注入")

    # E1 两行表头且第二行是真日期（考勤表）——不能因为"日期算数据"就把表头当数据体
    o = outdir("w_日期表头")
    expect_ok("E1", "日期表头不崩", ["tomd", fx("A17_日期表头.xlsx"), "-o", o])
    md = read(os.path.join(o, "A17_日期表头", "日期表头.md"))
    ok = "推断表头行数：**2**" in md
    record("E1b", "整行全是日期 → 判为日期表头行而非数据体", ok,
           "" if ok else "表头行数推断成了别的值")

    # E2 数据体第一行全为 0（停薪留职）——不能被当成"横跨整表的大标题行"整行丢掉
    o = outdir("w_全零行")
    expect_ok("E2", "全零数据行不崩", ["tomd", fx("A18_全零行.xlsx"), "-o", o])
    md = read(os.path.join(o, "A18_全零行", "全零行.md"))
    ok = "推断表头行数：**2**" in md and "数据体行数：**5**" in md
    record("E2b", "全零行必须留在数据体（不得当成大标题行）", ok,
           "" if ok else "全零行被吞了")

    # E3 维度列在数据区被合并
    sp = base_spec("A19_数据区合并.xlsx", "数据区合并", keys={"ID": "工号"},
                   dimensions={"DEPT": "部门"}, fields={"M": "金额"},
                   checks=[{"name": "x", "target": "M", "expr": "M"}])
    o = outdir("w_合并")
    rc, out, err, _ = sh(["verify", spec("E3", sp), "-o", o])
    rep = read(os.path.join(o, "验证报告.md")) if rc == 0 else ""
    ok = rc == 0 and "合并单元格" in rep and "fill_merged" in rep
    record("E3", "维度列数据区合并 → 告警并给出 fill_merged 修法", ok,
           "" if ok else "静默产出一堆空维度")

    sp["fill_merged"] = ["DEPT"]
    o = outdir("w_合并2")
    rc, out, err, _ = sh(["verify", spec("E3b", sp), "-o", o])
    det = read(os.path.join(o, "verification_detail.csv")) if rc == 0 else ""
    ok = rc == 0 and det.count("研发部") == 3 and det.count("市场部") == 2
    record("E3b", "fill_merged 铺开后维度值完整", ok, "" if ok else "铺开无效")

    # E4 除零必须点名是哪一行哪个人
    sp = base_spec("A20_除零.xlsx", "除零", keys={"ID": "工号"},
                   fields={"N": "分子", "D": "分母", "R": "比率"},
                   checks=[{"name": "比率", "target": "R", "expr": "N/D"}])
    rc, out, err, _ = sh(["verify", spec("E4", sp), "-o", outdir("w_除零")])
    ok = rc != 0 and "Traceback" not in err and "第3行" in err and "E002" in err
    record("E4", "表达式除零报出 Excel 行号与键值", ok,
           "" if ok else "错误信息里没有行号/键：%s" % err.strip()[-100:])

    # E5 公式注入不得进交付物
    sp = base_spec("A16_公式注入.xlsx", "注入", keys={"ID": "工号", "NAME": "姓名"},
                   fields={"M": "金额"}, checks=[{"name": "x", "target": "M", "expr": "M"}])
    o = outdir("w_注入")
    rc, out, err, _ = sh(["verify", spec("E5", sp), "-o", o])
    det = os.path.join(o, "verification_detail.csv")
    raw = read(det) if rc == 0 else ""
    bad = [ln for ln in raw.splitlines()[1:]
           if any(f",{p}" in ln for p in ("=1+1", "@SUM", "+1+1", "-2+3+cmd"))]
    record("E5", "CSV 里的公式起手文本被加前导单引号", rc == 0 and not bad,
           "" if not bad else "未转义：%s" % bad[:1])

    o2 = outdir("w_注入o")
    rc2, _o, err2, _ = sh(["output", spec("E5", sp), "-d", det, "-o", o2])
    live = []
    if rc2 == 0:
        import openpyxl
        wb = openpyxl.load_workbook(os.path.join(o2, "AI处理结果.xlsx"))
        for wsname in wb.sheetnames:
            for row in wb[wsname].iter_rows():
                for c in row:
                    if c.data_type == "f":
                        live.append("%s!%s=%r" % (wsname, c.coordinate, c.value))
    record("E5b", "结果 xlsx 里没有任何活公式单元格", rc2 == 0 and not live,
           "" if not live else "活公式：%s" % live[:3])

    # E6 转义后的 CSV 读回来仍是原值（不污染分组键）
    o3 = outdir("w_注入a")
    sp2 = dict(sp)
    sp2["analysis"] = {"distributions": [{"dim": "NAME"}]}
    rc3, _o, err3, _ = sh(["analyze", spec("E6", sp2), "-d", det, "-o", o3])
    rep = read(os.path.join(o3, "分析报告.md")) if rc3 == 0 else ""
    ok = rc3 == 0 and "| =1+1 |" in rep and "'=1+1" not in rep
    record("E6", "转义只在 CSV 落盘层，读回仍是原值", ok, "" if ok else "分组键被污染了")


# =========================================================================
def main():
    if not os.path.isdir(FX):
        print("请先运行 python3 make_attack_fixtures.py")
        return 2
    os.makedirs(RUN, exist_ok=True)
    print("目标脚本：", SCRIPT)
    group_tomd()
    group_spec_poison()
    group_silent()
    group_analyze()
    group_wave2()

    print("\n" + "=" * 78)
    bad = [r for r in RESULTS if not r[2]]
    print("共 %d 项攻击，守住 %d 项，攻破 %d 项。" % (len(RESULTS), len(RESULTS) - len(bad), len(bad)))
    if PERF:
        print("性能：" + "；".join("%s %.1fs" % (n, s) for n, s in PERF))
    if bad:
        print("\n仍被攻破：")
        for no, title, _ok, note in bad:
            print("  - [%s] %s —— %s" % (no, title, note))
    print("=" * 78)
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
