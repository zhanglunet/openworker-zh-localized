#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
脏数据 / 边界攻击夹具生成器。

只往 attacks/dirty-data/fx/ 写文件，不碰 tests/ 与 fixtures/。
每个夹具都对应一类"现实里真会出现、但很容易把脚本弄崩"的表。
"""

import datetime as dt
import os
import random
import shutil
import sys

import openpyxl
from openpyxl.utils import get_column_letter

FX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fx")


def _new():
    wb = openpyxl.Workbook()
    return wb, wb.active


def _save(wb, name):
    if not os.path.isdir(FX):
        os.makedirs(FX)
    p = os.path.join(FX, name)
    wb.save(p)
    print("  夹具 →", p)
    return p


# ---------------------------------------------------------------- A1 空表
def a01_空表():
    """完全空的 Sheet（0 行 0 列）+ 一个只有一个空格的 Sheet。"""
    wb, ws = _new()
    ws.title = "全空"
    ws2 = wb.create_sheet("只有空格")
    ws2["A1"] = "   "
    ws3 = wb.create_sheet("正常表")
    ws3.append(["工号", "金额"])
    ws3.append(["E1", 100])
    return _save(wb, "A01_空表.xlsx")


# ------------------------------------------------------------ A2 单行表
def a02_单行表():
    """只有表头一行，没有任何数据体；另一个只有一行数据没表头。"""
    wb, ws = _new()
    ws.title = "只有表头"
    ws.append(["工号", "姓名", "基本工资", "应发"])
    ws2 = wb.create_sheet("只有一行数据")
    ws2.append(["工号", "姓名", "基本工资", "应发"])
    ws2.append(["E001", "张三", 100, 100])
    ws3 = wb.create_sheet("单列表")
    ws3.append(["唯一列"])
    for i in range(5):
        ws3.append([i])
    return _save(wb, "A02_单行表.xlsx")


# ------------------------------------------------------ A3 全空列 / 全空行
def a03_全空列():
    wb, ws = _new()
    ws.title = "含全空列"
    ws.append(["工号", "姓名", "备用列1", "基本工资", "备用列2", "应发"])
    for i in range(1, 6):
        ws.append(["E%03d" % i, "员工%d" % i, None, 1000 * i, None, 1000 * i])
    ws.append([None, None, None, None, None, None])          # 全空行夹在中间
    for i in range(6, 9):
        ws.append(["E%03d" % i, "员工%d" % i, None, 1000 * i, None, 1000 * i])
    return _save(wb, "A03_全空列.xlsx")


# ---------------------------------------------------------- A4 超长列名
def a04_超长列名():
    wb, ws = _new()
    ws.title = "超长列名"
    long1 = "本列用于记录该员工在本考核周期内因参与集团级重点专项工作而产生的额外绩效奖励金额" * 60  # ~2400 字
    long2 = "X" * 40000                                       # 超过 openpyxl 单元格上限会被拒，故 4 万字放列名第二行
    ws.append(["工号", long1, "基本工资"])
    ws.append(["", long2[:32000], ""])                        # 32767 是 xlsx 单元格上限
    for i in range(1, 4):
        ws.append(["E%03d" % i, 10 * i, 1000 * i])
    return _save(wb, "A04_超长列名.xlsx")


# ---------------------------------------------------------- A5 重复列名
def a05_重复列名():
    wb, ws = _new()
    ws.title = "重复列名"
    # 三个"金额"：两个完全同名，一个是包含关系
    ws.append(["工号", "金额", "金额", "金额小计", "1", "01"])
    for i in range(1, 6):
        ws.append(["E%03d" % i, 100 * i, 200 * i, 300 * i, 7, 8])
    return _save(wb, "A05_重复列名.xlsx")


# ------------------------------------------------------ A6 中文数字混排
def a06_中文数字混排():
    wb, ws = _new()
    ws.title = "中文数字混排"
    ws.append(["工号", "姓名", "绩效系数", "补贴", "应发"])
    rows = [
        ["E001", "张三", "1.2", "1,234.50", 1481.4],
        ["E002", "李四", "/", "－500", -500.0],          # 全角减号
        ["E003", "王五", "０.８", "１２３４", 987.2],      # 全角数字
        ["E004", "赵六", "不参加", "¥ 2,000.00", 0.0],
        ["E005", "钱七", "1.5", "(300)", -450.0],        # 会计式括号负数
        ["E006", "孙八", "80%", "1000元", 800.0],        # 百分号 + 中文单位
        ["E007", "周九", "−1", "3 000", -3000.0],        # U+2212 减号 + 空格千分位
    ]
    for r in rows:
        ws.append(r)
    return _save(wb, "A06_中文数字混排.xlsx")


# ------------------------------------------------------------ A7 日期列
def a07_日期列():
    wb, ws = _new()
    ws.title = "日期列"
    ws.append(["工号", "入职日期", "考勤日期", "打卡时间", "金额"])
    base = dt.date(2024, 1, 1)
    for i in range(1, 7):
        ws.append(["E%03d" % i,
                   dt.datetime(2020 + i % 3, (i % 12) + 1, (i % 28) + 1),
                   base + dt.timedelta(days=i),
                   dt.time(8, (i * 7) % 60),
                   1000 * i])
    ws.append(["E007", None, None, None, 7000])   # 日期列有空
    return _save(wb, "A07_日期列.xlsx")


# ------------------------------------------------------- A8 科学计数法
def a08_科学计数法():
    wb, ws = _new()
    ws.title = "科学计数法"
    ws.append(["工号", "大数", "小数", "文本科学计数", "金额"])
    ws.append(["E001", 1.23e18, 1.5e-12, "1.23E+05", 123000.0])
    ws.append(["E002", 9.99e15, 2.5e-9, "4.56e3", 4560.0])
    ws.append(["E003", -7.77e17, -1e-10, "-1.1E+2", -110.0])
    ws.append(["E004", 0.0, 0.0, "1E999", 0.0])       # 溢出成 inf
    return _save(wb, "A08_科学计数法.xlsx")


# --------------------------------------------------------------- A9 负数
def a09_负数():
    wb, ws = _new()
    ws.title = "负数"
    ws.append(["工号", "应发", "扣款", "实发"])
    rows = [["E001", 1000, 1200, -200],
            ["E002", -500, 0, -500],
            ["E003", 0, 0, 0],
            ["E004", 1000, -300, 1300]]     # 负扣款（合法但需确认）
    for r in rows:
        ws.append(r)
    return _save(wb, "A09_负数.xlsx")


# --------------------------------------------------------- A10 超大表 1 万行
def a10_超大表(nrows=10000):
    wb, ws = _new()
    ws.title = "万行明细"
    ws.append(["某公司超大工资表"] + [None] * 9)
    ws.append(["工号", "姓名", "部门", "基本工资A", "绩效B1", "补贴B2",
               "社保C1", "公积金C2", "应发G", "实发L"])
    random.seed(20240805)
    for i in range(1, nrows + 1):
        a = round(random.uniform(3000, 30000), 2)
        b1 = round(a * random.uniform(0.1, 0.5), 2)
        b2 = round(random.choice([0, 300, 500, 800]), 2)
        c1 = round(a * 0.105, 2)
        c2 = round(a * 0.12, 2)
        g = round(a + b1 + b2 - c1 - c2, 2)
        l = round(g * 0.97, 2)
        ws.append(["E%05d" % i, "员工%05d" % i, "部门%02d" % (i % 20),
                   a, b1, b2, c1, c2, g, l])
    ws.append([None, "合计", None, None, None, None, None, None, None, None])
    return _save(wb, "A10_超大表.xlsx")


# ------------------------------------------------------- A11 无缓存值 xlsx
def a11_无缓存值():
    """公式单元格只有公式、没有缓存结果（程序生成的 xlsx 就长这样）。"""
    wb, ws = _new()
    ws.title = "无缓存"
    ws.append(["工号", "基本工资", "补贴", "应发"])
    for i in range(1, 6):
        r = i + 1
        ws.cell(row=r, column=1, value="E%03d" % i)
        ws.cell(row=r, column=2, value=1000 * i)
        ws.cell(row=r, column=3, value=100 * i)
        ws.cell(row=r, column=4, value="=B%d+C%d" % (r, r))   # 无缓存值
    return _save(wb, "A11_无缓存值.xlsx")


# ------------------------------------------------------------- A12 .xls
def a12_xls旧格式():
    """真·旧格式无法用 openpyxl 造，这里造两种：
       (1) 真正的 BIFF8 头（伪造前 8 字节，pandas/xlrd 会认出是 xls）
       (2) 一个 xlsx 内容但扩展名叫 .xls（用户最常见的误操作）
    """
    p1 = os.path.join(FX, "A12_真旧格式.xls")
    with open(p1, "wb") as f:
        # OLE2 复合文档魔数，xlrd 会认它是 xls
        f.write(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512)
    print("  夹具 →", p1)

    wb, ws = _new()
    ws.title = "伪装"
    ws.append(["工号", "金额"])
    ws.append(["E001", 100])
    p2 = os.path.join(FX, "A12_其实是xlsx.xls")
    wb.save(p2)
    print("  夹具 →", p2)

    # (3) 完全不是表格的文件，但叫 .xlsx
    p3 = os.path.join(FX, "A12_不是表格.xlsx")
    with open(p3, "w", encoding="utf-8") as f:
        f.write("这不是一个 Excel 文件，只是一段中文文本。\n")
    print("  夹具 →", p3)
    return p1, p2, p3


# --------------------------------------------------- A13 汇总行未排除
def a13_汇总行():
    wb, ws = _new()
    ws.title = "带汇总"
    ws.append(["工号", "姓名", "基本工资", "补贴", "应发"])
    for i in range(1, 6):
        ws.append(["E%03d" % i, "员工%d" % i, 1000 * i, 100 * i, 1100 * i])
    ws.append([None, "合计", 15000, 1500, 16500])
    ws.append([None, "小计", 15000, 1500, 16500])
    return _save(wb, "A13_汇总行.xlsx")


# ---------------------------------------------- A14 Sheet 名会撞文件名
def a14_sheet名冲突():
    """safe_name 会 strip 掉首尾空格与点号，两个不同 Sheet 会撞成同一个 md 文件名。"""
    wb, ws = _new()
    ws.title = "报表."           # safe_name 后变成「报表」
    ws.append(["工号", "金额"])
    ws.append(["E001", 1])
    ws2 = wb.create_sheet("报表")
    ws2.append(["工号", "金额"])
    ws2.append(["E002", 2])
    ws3 = wb.create_sheet(" 报表 ")
    ws3.append(["工号", "金额"])
    ws3.append(["E003", 3])
    return _save(wb, "A14_Sheet名冲突.xlsx")


# ------------------------------------------------ A15 控制字符与特殊文本
def a15_特殊文本():
    wb, ws = _new()
    ws.title = "特殊文本"
    ws.append(["工号", "姓名", "备注", "金额"])
    ws.append(["E001", "张|三", "含竖线|与\n换行", 100])
    ws.append(["E002", "李\t四", "制表符", 200])
    ws.append(["E003", "=1+1", "看起来像公式的文本", 300])
    ws.append(["E004", "007", "前导零工号", 400])
    ws.append(["E005", "总计部", "部门名里带汇总词", 500])
    return _save(wb, "A15_特殊文本.xlsx")


def main():
    if os.path.isdir(FX):
        shutil.rmtree(FX)
    os.makedirs(FX)
    print("生成攻击夹具到", FX)
    a01_空表()
    a02_单行表()
    a03_全空列()
    a04_超长列名()
    a05_重复列名()
    a06_中文数字混排()
    a07_日期列()
    a08_科学计数法()
    a09_负数()
    a10_超大表()
    a11_无缓存值()
    a12_xls旧格式()
    a13_汇总行()
    a14_sheet名冲突()
    a15_特殊文本()
    print("完成。")


if __name__ == "__main__":
    sys.exit(main())


# ================= 第二波：更刁钻的现实脏表 =================

def a16_公式注入():
    """姓名/备注里藏 CSV/DDE 注入 payload —— 我们的交付物不能把它变成活公式。"""
    wb, ws = _new()
    ws.title = "注入"
    ws.append(["工号", "姓名", "备注", "金额"])
    rows = [["E001", "张三", "正常", 100],
            ["E002", "=1+1", "等号开头", 200],
            ["E003", "@SUM(A1:A9)", "at开头", 300],
            ["E004", "+1+1", "加号开头", 400],
            ["E005", "-2+3+cmd|'/c calc'!A1", "减号DDE", 500]]
    for r in rows:
        # openpyxl 会把 "=..." 自动当公式，这里用 data_type 强制成文本来模拟"原表里就是文本"
        for j, v in enumerate(r, start=1):
            c = ws.cell(row=ws.max_row + (1 if j == 1 else 0), column=j)
            c.value = v
            if isinstance(v, str) and v[:1] in "=+@-":
                c.data_type = "s"
    return _save(wb, "A16_公式注入.xlsx")


def a17_日期表头():
    """两行表头，第二行是真日期（考勤表最常见）—— 不能因为"日期算数据"就把表头当数据体。"""
    wb, ws = _new()
    ws.title = "日期表头"
    ws.append(["工号", "姓名", "1日", "2日", "3日"])
    ws.append([None, None, dt.date(2024, 5, 1), dt.date(2024, 5, 2), dt.date(2024, 5, 3)])
    for i in range(1, 5):
        ws.append(["E%03d" % i, "员工%d" % i, 8, 8, 7.5])
    return _save(wb, "A17_日期表头.xlsx")


def a18_全零行():
    """数据体第一行各项全为 0（停薪留职）—— 不能被当成"横跨整表的大标题行"。"""
    wb, ws = _new()
    ws.title = "全零行"
    ws.append(["XX公司2024年5月工资表"] + [None] * 7)
    ws.append(["工号", "姓名", "A", "B", "C", "D", "E", "应发"])
    ws.append([0, 0, 0, 0, 0, 0, 0, 0])
    for i in range(1, 5):
        ws.append(["E%03d" % i, "员工%d" % i, i, i, i, i, i, 5 * i])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    return _save(wb, "A18_全零行.xlsx")


def a19_数据区合并():
    """部门列在数据区被合并，只有第一行有值 —— 分组统计会全跑到"(空)"里。"""
    wb, ws = _new()
    ws.title = "数据区合并"
    ws.append(["工号", "部门", "金额"])
    ws.append(["E001", "研发部", 100])
    ws.append(["E002", None, 200])
    ws.append(["E003", None, 300])
    ws.append(["E004", "市场部", 400])
    ws.append(["E005", None, 500])
    ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
    ws.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)
    return _save(wb, "A19_数据区合并.xlsx")


def a20_除零():
    wb, ws = _new()
    ws.title = "除零"
    ws.append(["工号", "分子", "分母", "比率"])
    ws.append(["E001", 10, 2, 5])
    ws.append(["E002", 10, 0, 0])      # 第 3 行分母为 0
    ws.append(["E003", 10, 5, 2])
    return _save(wb, "A20_除零.xlsx")


def main2():
    a16_公式注入()
    a17_日期表头()
    a18_全零行()
    a19_数据区合并()
    a20_除零()
    print("第二波完成。")
