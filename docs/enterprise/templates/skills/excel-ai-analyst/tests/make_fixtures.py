#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel-ai-analyst 技能包 —— 「真实世界脏数据」Excel 夹具生成器
==============================================================

SKILL.md 的「通用踩坑清单」列了 11 条坑，本文件为其中 10 条各造一个 xlsx 夹具，
用来给 scripts/excel_ai.py 做独立回归测试（测试代码见同目录 test_excel_ai.py）。

夹具与坑的对应关系：

    01_多行表头合并单元格.xlsx  坑1 多行表头(4行) + 坑2 合并单元格 + 横跨整表的大标题行
    02_列名写公式.xlsx          坑… 公式不在单元格里而在列名里（应发工资G=A+B-C+D+E-F）
    03_单元格公式带缓存.xlsx    真实单元格公式 `=C2+D2-E2` 且带缓存值（tomd 抽公式）
    04_混合类型列.xlsx          坑5 数值列里混 `/`、`-`、"不适用"、空 → 类型告警 + 按 0.0 参与运算
    05_汇总行.xlsx              坑6 小计/合计/总计 混在数据体里 → skip_when 排掉否则金额翻倍
    06_check列.xlsx             坑4 原作者留下的断言列，白送的验证锚点
    07_跨表引用.xlsx            跨 Sheet 数据血缘（工号关联），且故意留 1 处不一致
    08_无缓存值.xlsx            坑11 公式列"全空"：程序生成的 xlsx 没有缓存值
    09_重名列.xlsx              列名模糊匹配歧义 → 必须报错并列候选；兼测"先精确后包含"
    10_有意算错.xlsx            数据体里有人为覆盖/算错的行 —— 验证环节最有价值的发现

设计原则：
  * 数据全部**确定性**（无随机），测试可以据此精确算出期望值；
  * 本模块同时导出原始数据常量（SALARY_ROWS / SALARY_ERRORS / ...），
    测试直接 import 后自行推导期望值，不把期望值硬编码两遍；
  * 只用 openpyxl 写，再对 xlsx 里的 sheet XML 做一次后处理来注入公式缓存值
    （openpyxl 本身没法同时写公式和缓存值）。

用法：
    python3 make_fixtures.py [输出目录]        # 默认 ./fixtures
"""

from __future__ import annotations

import json
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 夹具文件名 / Sheet 名（测试用这些常量引用，避免两处写死字符串）
# ---------------------------------------------------------------------------

F01 = "01_多行表头合并单元格.xlsx"
F02 = "02_列名写公式.xlsx"
F03 = "03_单元格公式带缓存.xlsx"
F04 = "04_混合类型列.xlsx"
F05 = "05_汇总行.xlsx"
F06 = "06_check列.xlsx"
F07 = "07_跨表引用.xlsx"
F08 = "08_无缓存值.xlsx"
F09 = "09_重名列.xlsx"
F10 = "10_有意算错.xlsx"

S01 = "工资明细表"
S02 = "结果列"
S03 = "明细"
S04 = "绩效"
S05 = "部门工资"
S06 = "带校验"
S07_MAIN = "工资表"
S07_REF = "绩效表"
S08 = "公式无缓存"
S09 = "重名"
S10 = "工资明细表"

ALL_FIXTURES = [F01, F02, F03, F04, F05, F06, F07, F08, F09, F10]

# ---------------------------------------------------------------------------
# 主数据：工资明细（F01 正确版 / F10 埋错版 共用同一套结构与数据）
# ---------------------------------------------------------------------------

# 4 行表头：第 1 行大标题（横跨整表，必须被 tomd 跳过）
SALARY_TITLE = "2026年6月员工薪酬明细表（全公司·公式勿动）"
# 第 2 行分组表头（合并单元格），(组名, 占几列)
SALARY_GROUPS = [
    ("基本信息", 3),
    ("固定项", 3),
    ("浮动项", 2),
    ("扣款项", 2),
    ("其他增减", 3),
    ("结果", 1),
]
# 第 3 行字段名
SALARY_FIELDS = [
    "工号", "姓名", "一级部门",
    "基本工资A", "岗位津贴A2", "工龄工资A3",
    "绩效工资B1", "加班费B2",
    "社保C1", "公积金C2",
    "补贴D", "奖金E", "罚款F",
    "应发工资G",
]
# 第 4 行单位行（末列顺带带上"公式勿动"，对应坑3）
SALARY_UNITS = ["", "", "", "元", "元", "元", "元", "元", "元", "元", "元", "元", "元", "元（公式勿动）"]

SALARY_HEADER_ROWS = 4  # 表头占 4 行，数据体从第 5 行开始

# 键: 工号；值: (姓名, 一级部门, A, A2, A3, B1, B2, C1, C2, D, E, F)
SALARY_ROWS = [
    ("E001", "陈静", "研发中心", 12340.00, 800.00, 300.00, 3200.50, 450.00, 1360.00, 1440.00, 200.00, 1500.00, 0.00),
    ("E002", "李强", "研发中心", 9800.00, 600.00, 200.00, 2100.25, 320.50, 1080.00, 1140.00, 150.00, 800.00, 120.00),
    ("E003", "赵敏", "研发中心", 15600.00, 900.00, 500.00, 4300.00, 0.00, 1720.00, 1820.00, 300.00, 2000.00, 0.00),
    ("E004", "孙涛", "市场部", 8600.00, 500.00, 100.00, 1800.75, 260.00, 950.00, 1000.00, 100.00, 600.00, 50.00),
    ("E005", "周杰", "市场部", 11200.00, 700.00, 400.00, 2650.00, 480.25, 1240.00, 1300.00, 250.00, 1200.00, 0.00),
    ("E006", "吴磊", "市场部", 7400.00, 400.00, 0.00, 1200.00, 180.00, 820.00, 860.00, 80.00, 400.00, 200.00),
    ("E007", "郑爽", "财务部", 10500.00, 650.00, 300.00, 2400.40, 0.00, 1160.00, 1220.00, 180.00, 900.00, 0.00),
    ("E008", "王芳", "财务部", 13800.00, 850.00, 600.00, 3600.00, 520.00, 1520.00, 1600.00, 220.00, 1800.00, 0.00),
    ("E009", "刘洋", "财务部", 8900.00, 550.00, 100.00, 1900.00, 240.00, 980.00, 1030.00, 120.00, 500.00, 80.00),
    # 高薪离群点：给 analyze 的 IQR / Z-score 离群检测一个必须被逮到的目标
    ("E010", "高峰", "研发中心", 38000.00, 2000.00, 1200.00, 12000.00, 0.00, 4180.00, 4560.00, 800.00, 6000.00, 0.00),
]

# F10 里被"人为覆盖"的应发工资G：工号 -> 相对正确值的偏差
#   +100.00 : 典型的人为覆盖，必须被逮出来（最有价值的发现）
#   +0.005  : 舍入噪声，金额容差 0.01 之内，**不**该被判为不匹配
#   +0.05   : 超出 0.01 容差、但在 0.10 容差之内 —— 用来测容差边界
SALARY_ERRORS = {"E003": 100.00, "E007": 0.005, "E009": 0.05}


def salary_g(row) -> float:
    """应发工资G = A + A2 + A3 + B1 + B2 - C1 - C2 + D + E - F（保留 2 位）。"""
    (_id, _name, _dept, a, a2, a3, b1, b2, c1, c2, d, e, f) = row
    return round(a + a2 + a3 + b1 + b2 - c1 - c2 + d + e - f, 2)


# ---------------------------------------------------------------------------
# 底层小工具
# ---------------------------------------------------------------------------

_HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
_TITLE_FONT = Font(bold=True, size=14)
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")


def _put_row(ws, row_idx: int, values, *, bold=False, fill=False) -> None:
    """把一行值写进 ws 的第 row_idx 行（1 基）。"""
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        if bold:
            cell.font = _BOLD
        if fill:
            cell.fill = _HDR_FILL
            cell.alignment = _CENTER


def _sheet_xml_map(path: Path) -> dict:
    """返回 {Sheet 名: zip 内 worksheet xml 路径}。"""
    with zipfile.ZipFile(path) as zf:
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    rid2target = {}
    for m in re.finditer(r"<Relationship\b[^>]*/>", rels_xml):
        tag = m.group(0)
        if "/worksheet" not in tag:
            continue
        rid = re.search(r'Id="([^"]+)"', tag).group(1)
        target = re.search(r'Target="([^"]+)"', tag).group(1).lstrip("/")
        if not target.startswith("xl/"):
            target = "xl/" + target
        rid2target[rid] = target

    mapping = {}
    for m in re.finditer(r"<sheet\b[^>]*/>", wb_xml):
        tag = m.group(0)
        name = re.search(r'name="([^"]+)"', tag).group(1)
        rid = re.search(r'r:id="([^"]+)"', tag).group(1)
        mapping[name] = rid2target[rid]
    return mapping


def _inject_cached_values(path: Path, cached: dict) -> None:
    """
    给公式单元格补上缓存值（`<f>…</f><v>值</v>`）。

    openpyxl 只能写公式、写不了缓存值（存出来是 `<f>…</f><v/>`），
    于是真实业务表里"公式列有值"的样子必须靠改 sheet XML 造出来。
    cached 形如 {"Sheet名": {"F2": 123.45, ...}}。
    """
    sheet_map = _sheet_xml_map(path)
    with zipfile.ZipFile(path) as zf:
        blobs = {name: zf.read(name) for name in zf.namelist()}

    for sheet_name, cells in cached.items():
        xml_path = sheet_map[sheet_name]
        xml = blobs[xml_path].decode("utf-8")
        for ref, value in cells.items():
            pattern = re.compile(
                r'(<c r="%s"(?:\s[^>]*)?>)(\s*<f>.*?</f>)\s*(?:<v\s*/>|<v>.*?</v>)?\s*(</c>)' % re.escape(ref),
                re.S,
            )
            xml, n = pattern.subn(
                lambda m, v=value: "%s%s<v>%s</v>%s" % (m.group(1), m.group(2), v, m.group(3)), xml
            )
            if n != 1:
                raise RuntimeError("注入缓存值失败：%s!%s 命中 %d 次" % (sheet_name, ref, n))
        blobs[xml_path] = xml.encode("utf-8")

    # 顺手去掉 fullCalcOnLoad，让文件更像"人手工另存过一次"的真实工作簿
    wb_key = "xl/workbook.xml"
    if wb_key in blobs:
        blobs[wb_key] = blobs[wb_key].decode("utf-8").replace(' fullCalcOnLoad="1"', "").encode("utf-8")

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, blob in blobs.items():
            zf.writestr(name, blob)


# ---------------------------------------------------------------------------
# 夹具 01 / 10：4 行表头 + 合并单元格（10 埋了 3 处差异）
# ---------------------------------------------------------------------------

def _build_salary(path: Path, sheet_name: str, errors: dict) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ncols = len(SALARY_FIELDS)

    # 第 1 行：横跨整表的大标题（tomd 必须跳过它，不能拼进列名）
    ws.cell(row=1, column=1, value=SALARY_TITLE).font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).alignment = _CENTER

    # 第 2 行：分组表头，全部是合并单元格（拼列名前必须先把值铺开）
    col = 1
    for group_name, span in SALARY_GROUPS:
        cell = ws.cell(row=2, column=col, value=group_name)
        cell.font = _BOLD
        cell.fill = _HDR_FILL
        cell.alignment = _CENTER
        if span > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + span - 1)
            for extra in range(1, span):
                ws.cell(row=2, column=col + extra).fill = _HDR_FILL
        col += span
    assert col - 1 == ncols, "分组表头列数与字段数对不上"

    # 第 3 行字段名 / 第 4 行单位行
    _put_row(ws, 3, SALARY_FIELDS, bold=True, fill=True)
    _put_row(ws, 4, SALARY_UNITS, fill=True)

    # 第 5 行起是数据体
    for i, row in enumerate(SALARY_ROWS):
        r = SALARY_HEADER_ROWS + 1 + i
        g = salary_g(row)
        g_written = round(g + errors.get(row[0], 0.0), 3) if row[0] in errors else g
        _put_row(ws, r, list(row) + [g_written])

    ws.freeze_panes = "A5"
    wb.save(path)
    return path


def build_01(outdir: Path) -> Path:
    """坑1 多行表头(4行) + 坑2 合并单元格 + 横跨整表的大标题行；数据全部算得对。"""
    return _build_salary(outdir / F01, S01, errors={})


def build_10(outdir: Path) -> Path:
    """与 01 结构完全一致，但 3 行的应发工资G 被"人为覆盖"过（+100 / +0.005 / +0.05）。"""
    return _build_salary(outdir / F10, S10, errors=SALARY_ERRORS)


# ---------------------------------------------------------------------------
# 夹具 02：公式写在列名里
# ---------------------------------------------------------------------------

F02_RESULT_COL = "应发工资G=A+B-C+D+E-F"
F02_COLS = ["工号", "姓名", "基本工资A", "浮动合计B", "扣款合计C", "补贴D", "奖金E", "罚款F", F02_RESULT_COL]
# (工号, 姓名, A, B, C, D, E, F)
F02_ROWS = [
    ("E001", "陈静", 12340.00, 3650.50, 2800.00, 200.00, 1500.00, 0.00),
    ("E002", "李强", 9800.00, 2420.75, 2220.00, 150.00, 800.00, 120.00),
    ("E003", "赵敏", 15600.00, 4300.00, 3540.00, 300.00, 2000.00, 0.00),
    ("E004", "孙涛", 8600.00, 2060.75, 1950.00, 100.00, 600.00, 50.00),
    ("E005", "周杰", 11200.00, 3130.25, 2540.00, 250.00, 1200.00, 0.00),
    ("E006", "吴磊", 7400.00, 1380.00, 1680.00, 80.00, 400.00, 200.00),
]


def f02_g(row) -> float:
    _id, _name, a, b, c, d, e, f = row
    return round(a + b - c + d + e - f, 2)


def build_02(outdir: Path) -> Path:
    """公式不在单元格里、而在列名里（SKILL Step3：中文列名 + 字母编号混排，照单全收）。"""
    path = outdir / F02
    wb = Workbook()
    ws = wb.active
    ws.title = S02
    _put_row(ws, 1, F02_COLS, bold=True, fill=True)
    for i, row in enumerate(F02_ROWS):
        _put_row(ws, 2 + i, list(row) + [f02_g(row)])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 夹具 03：单元格里真有公式，且带缓存值
# ---------------------------------------------------------------------------

F03_COLS = ["工号", "姓名", "基本工资A", "浮动合计B", "扣款合计C", "应发工资G"]
F03_ROWS = [
    ("E001", "陈静", 12340.00, 3650.50, 2800.00),
    ("E002", "李强", 9800.00, 2420.75, 2220.00),
    ("E003", "赵敏", 15600.00, 4300.00, 3540.00),
    ("E004", "孙涛", 8600.00, 2060.75, 1950.00),
    ("E005", "周杰", 11200.00, 3130.25, 2540.00),
]
# 第 2 行起，F 列写 =C{r}+D{r}-E{r}
F03_FORMULA_TEMPLATE = "=C{r}+D{r}-E{r}"
F03_SAMPLE_FORMULA = "C2+D2-E2"  # 报告里通常带不带 "=" 都可能，测试用不带 = 的核心部分


def f03_g(row) -> float:
    _id, _name, a, b, c = row
    return round(a + b - c, 2)


def build_03(outdir: Path) -> Path:
    """真实单元格公式 + 缓存值：tomd 必须用 data_only=False 抽出公式本身。"""
    path = outdir / F03
    wb = Workbook()
    ws = wb.active
    ws.title = S03
    _put_row(ws, 1, F03_COLS, bold=True, fill=True)
    cached = {}
    for i, row in enumerate(F03_ROWS):
        r = 2 + i
        _put_row(ws, r, list(row))
        ws.cell(row=r, column=6, value=F03_FORMULA_TEMPLATE.format(r=r))
        cached["F%d" % r] = f03_g(row)
    wb.save(path)
    _inject_cached_values(path, {S03: cached})
    return path


# ---------------------------------------------------------------------------
# 夹具 04：混合类型列（`/`、`-`、"不适用"、空）
# ---------------------------------------------------------------------------

F04_COLS = ["工号", "姓名", "一级部门", "绩效系数COEF", "绩效基数BASE", "绩效工资B1"]
# 12 行里：8 个数值 + 3 个非数值文本 + 1 个空 →
# 数值占比 8/12=66.7%（含空）或 8/11=72.7%（不含空），两种口径都落在
# [50%, 95%) 区间内，必须触发"数值为主 ⚠️含非数值"告警而不是判成纯数值列。
F04_ROWS = [
    ("E001", "陈静", "研发中心", 1.0, 3000.00),
    ("E002", "李强", "研发中心", 0.9, 2800.00),
    ("E003", "赵敏", "研发中心", "/", 3200.00),        # 不参加考核
    ("E004", "孙涛", "市场部", 1.2, 2500.00),
    ("E005", "周杰", "市场部", "-", 2600.00),          # 无此项
    ("E006", "吴磊", "市场部", 0.8, 2200.00),
    ("E007", "郑爽", "财务部", "不适用", 2400.00),      # 中文说明
    ("E008", "王芳", "财务部", 1.1, 3100.00),
    ("E009", "刘洋", "财务部", None, 2000.00),          # 空
    ("E010", "高峰", "研发中心", 1.0, 5000.00),
    ("E011", "钱多", "市场部", 0.95, 2700.00),
    ("E012", "孙俪", "财务部", 1.05, 2900.00),
]


def f04_coef_num(coef) -> float:
    """非数值（/ - 不适用 空）一律按 0.0 参与运算 —— SKILL 明确要求的口径。"""
    try:
        return float(coef)
    except (TypeError, ValueError):
        return 0.0


def f04_b1(row) -> float:
    _id, _name, _dept, coef, base = row
    return round(f04_coef_num(coef) * base, 2)


def build_04(outdir: Path) -> Path:
    """坑5：数值列里混 `/`、`-`、"不适用"、空 —— 类型告警 + 非数值按 0.0 运算。"""
    path = outdir / F04
    wb = Workbook()
    ws = wb.active
    ws.title = S04
    _put_row(ws, 1, F04_COLS, bold=True, fill=True)
    for i, row in enumerate(F04_ROWS):
        _put_row(ws, 2 + i, list(row) + [f04_b1(row)])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 夹具 05：汇总行混在数据体里
# ---------------------------------------------------------------------------

F05_COLS = ["工号", "姓名", "一级部门", "基本工资A", "浮动合计B", "扣款合计C", "应发工资G"]
# (工号, 姓名, 部门, A, B, C)；工号为 None 表示这是一行汇总行
F05_EMPLOYEES = [
    ("E001", "王芳", "销售一部", 9000.00, 2100.50, 1800.00),
    ("E002", "刘洋", "销售一部", 10500.00, 2450.25, 2050.00),
    ("E003", "赵敏", "销售二部", 8800.00, 1900.00, 1700.00),
    ("E004", "孙涛", "销售二部", 12000.00, 3300.75, 2400.00),
    ("E005", "周杰", "销售二部", 7600.00, 1500.00, 1450.00),
]


def f05_g(row) -> float:
    _id, _name, _dept, a, b, c = row
    return round(a + b - c, 2)


F05_EMPLOYEE_TOTAL = round(sum(f05_g(r) for r in F05_EMPLOYEES), 2)


def build_05(outdir: Path) -> Path:
    """坑6：小计/合计/总计 夹在数据体中间，不排掉的话金额直接翻倍。"""
    path = outdir / F05
    wb = Workbook()
    ws = wb.active
    ws.title = S05
    _put_row(ws, 1, F05_COLS, bold=True, fill=True)

    r = 2
    for dept in ("销售一部", "销售二部"):
        members = [x for x in F05_EMPLOYEES if x[2] == dept]
        for m in members:
            _put_row(ws, r, list(m) + [f05_g(m)])
            r += 1
        # 小计行：工号为空（skip_when.empty 抓它），姓名写"小计"（label_in 也抓它）
        _put_row(ws, r, [
            None, "小计", dept,
            round(sum(m[3] for m in members), 2),
            round(sum(m[4] for m in members), 2),
            round(sum(m[5] for m in members), 2),
            round(sum(f05_g(m) for m in members), 2),
        ], bold=True)
        r += 1
    # 总计行：工号这次**不为空**而是写着"合计"，只有 label_in 能抓到它
    _put_row(ws, r, [
        "合计", "合计", None,
        round(sum(m[3] for m in F05_EMPLOYEES), 2),
        round(sum(m[4] for m in F05_EMPLOYEES), 2),
        round(sum(m[5] for m in F05_EMPLOYEES), 2),
        F05_EMPLOYEE_TOTAL,
    ], bold=True)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 夹具 06：check 列（原作者的断言）
# ---------------------------------------------------------------------------

F06_COLS = ["工号", "姓名", "基本工资A", "浮动合计B", "扣款合计C", "应发工资G", "核对check"]
F06_ROWS = [
    ("E001", "陈静", 12340.00, 3650.50, 2800.00),
    ("E002", "李强", 9800.00, 2420.75, 2220.00),
    ("E003", "赵敏", 15600.00, 4300.00, 3540.00),
    ("E004", "孙涛", 8600.00, 2060.75, 1950.00),
    ("E005", "周杰", 11200.00, 3130.25, 2540.00),
    ("E006", "吴磊", 7400.00, 1380.00, 1680.00),
]


def f06_g(row) -> float:
    _id, _name, a, b, c = row
    return round(a + b - c, 2)


def build_06(outdir: Path) -> Path:
    """坑4：check 列是原作者留下的单元测试，恒为 0，拿来当验证锚点。"""
    path = outdir / F06
    wb = Workbook()
    ws = wb.active
    ws.title = S06
    _put_row(ws, 1, F06_COLS, bold=True, fill=True)
    for i, row in enumerate(F06_ROWS):
        g = f06_g(row)
        check = round(g - (row[2] + row[3] - row[4]), 2)  # 恒等于 0.0
        _put_row(ws, 2 + i, list(row) + [g, check])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 夹具 07：跨表引用（工号关联），故意留 1 处不一致
# ---------------------------------------------------------------------------

F07_MAIN_COLS = ["工号", "姓名", "绩效系数COEF", "绩效基数", "绩效工资B1"]
F07_REF_COLS = ["工号", "姓名", "考核月份", "绩效系数"]
# (工号, 姓名, 工资表里的系数, 绩效基数)
F07_ROWS = [
    ("E001", "陈静", 1.00, 3000.00),
    ("E002", "李强", 0.90, 2800.00),
    ("E003", "赵敏", 1.20, 3200.00),
    ("E004", "孙涛", 1.00, 2500.00),   # ← 绩效表里是 1.30，跨表传递不一致
    ("E005", "周杰", 0.80, 2600.00),
    ("E006", "吴磊", 1.10, 2200.00),
]
F07_INCONSISTENT_ID = "E004"
F07_REF_COEF_OVERRIDE = {"E004": 1.30}


def f07_b1(row) -> float:
    _id, _name, coef, base = row
    return round(coef * base, 2)


def build_07(outdir: Path) -> Path:
    """跨 Sheet 数据血缘：绩效表.绩效系数 → 工资表.绩效系数COEF，连接键=工号。"""
    path = outdir / F07
    wb = Workbook()
    ws = wb.active
    ws.title = S07_MAIN
    _put_row(ws, 1, F07_MAIN_COLS, bold=True, fill=True)
    for i, row in enumerate(F07_ROWS):
        _put_row(ws, 2 + i, list(row) + [f07_b1(row)])

    ws2 = wb.create_sheet(S07_REF)
    _put_row(ws2, 1, F07_REF_COLS, bold=True, fill=True)
    for i, row in enumerate(F07_ROWS):
        coef = F07_REF_COEF_OVERRIDE.get(row[0], row[2])
        _put_row(ws2, 2 + i, [row[0], row[1], "2026-06", coef])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 夹具 08：公式列没有缓存值（程序生成的 xlsx）
# ---------------------------------------------------------------------------

F08_COLS = ["工号", "姓名", "数量", "单价", "金额小计"]
F08_ROWS = [
    ("E001", "陈静", 12, 35.50),
    ("E002", "李强", 8, 42.00),
    ("E003", "赵敏", 20, 18.75),
    ("E004", "孙涛", 5, 99.90),
    ("E005", "周杰", 33, 7.20),
]
F08_FORMULA_TEMPLATE = "=C{r}*D{r}"
F08_SAMPLE_FORMULA = "C2*D2"


def build_08(outdir: Path) -> Path:
    """坑11：公式列"全空" —— 这个 xlsx 没有缓存值，tomd 仍应把公式本身抽出来并且不崩。"""
    path = outdir / F08
    wb = Workbook()
    ws = wb.active
    ws.title = S08
    _put_row(ws, 1, F08_COLS, bold=True, fill=True)
    for i, row in enumerate(F08_ROWS):
        r = 2 + i
        _put_row(ws, r, list(row))
        ws.cell(row=r, column=5, value=F08_FORMULA_TEMPLATE.format(r=r))
    wb.save(path)  # 故意不注入缓存值
    return path


# ---------------------------------------------------------------------------
# 夹具 09：重名列 + 精确/包含匹配优先级
# ---------------------------------------------------------------------------

# 第 5、6 两列都叫"金额" → 按列名引用必须报错并列候选；
# "工资" 与 "基本工资"：按 "工资" 引用时精确匹配应当胜出（否则会歧义）。
F09_COLS = ["工号", "姓名", "工资", "基本工资", "金额", "金额", "结果"]
F09_DUP_NAME = "金额"
F09_EXACT_NAME = "工资"
F09_AMOUNT_1 = 7.00      # 第 5 列（1 基）
F09_AMOUNT_2 = 9.00      # 第 6 列（1 基）
F09_WAGE = 100.00        # "工资" 列
F09_BASE_WAGE = 1000.00  # "基本工资" 列（若匹配错列，结果对不上）
F09_ROWS = [
    ("E001", "陈静"),
    ("E002", "李强"),
    ("E003", "赵敏"),
    ("E004", "孙涛"),
]


def build_09(outdir: Path) -> Path:
    """列名模糊匹配的两个反面教材：真歧义要报错，假歧义（精确优先）不许报错。"""
    path = outdir / F09
    wb = Workbook()
    ws = wb.active
    ws.title = S09
    _put_row(ws, 1, F09_COLS, bold=True, fill=True)
    for i, row in enumerate(F09_ROWS):
        # 结果列 == "工资"列 的值：匹配到"基本工资"就会立刻露馅
        _put_row(ws, 2 + i, [row[0], row[1], F09_WAGE, F09_BASE_WAGE,
                             F09_AMOUNT_1, F09_AMOUNT_2, F09_WAGE])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------

_BUILDERS = [
    ("坑1+坑2 多行表头与合并单元格", F01, build_01),
    ("公式写在列名里", F02, build_02),
    ("真实单元格公式（带缓存值）", F03, build_03),
    ("坑5 混合类型列", F04, build_04),
    ("坑6 汇总行混在数据体", F05, build_05),
    ("坑4 check 断言列", F06, build_06),
    ("跨表引用（含 1 处不一致）", F07, build_07),
    ("坑11 公式列无缓存值", F08, build_08),
    ("重名列 / 精确匹配优先", F09, build_09),
    ("有意算错的行（最有价值的发现）", F10, build_10),
]


def build_all(outdir) -> dict:
    """生成全部夹具，返回 {文件名: Path}。"""
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    made = {}
    manifest = []
    for desc, filename, builder in _BUILDERS:
        path = builder(outdir)
        made[filename] = path
        manifest.append({"文件": filename, "针对的坑": desc})
    (outdir / "00_夹具清单.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return made


def main(argv=None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    outdir = Path(argv[0]) if argv else Path.cwd() / "fixtures"
    made = build_all(outdir)
    print("夹具输出目录：%s" % outdir)
    for desc, filename, _ in _BUILDERS:
        print("  %-28s  %s" % (filename, desc))
    print("共 %d 个夹具。" % len(made))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
