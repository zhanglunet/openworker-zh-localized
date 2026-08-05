#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/excel_ai.py 的独立回归测试
==================================

本文件是 **独立验证方**：只依据 SKILL.md（唯一事实来源）里「工具脚本从哪来」一节
对 tomd / verify / output / analyze 四个子命令的行为规格来写断言，
全部通过 `subprocess` 走命令行调用被测脚本，不 import、不读它的实现。

夹具由同目录 make_fixtures.py 生成（每个夹具对应「通用踩坑清单」里的一条坑）。
期望值一律由夹具的原始数据现算，不在两处硬编码。

跑法：
    python3 -m pytest tests/test_excel_ai.py -v
    EXCEL_AI=/path/to/excel_ai.py python3 -m pytest tests/test_excel_ai.py -v
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent))
import make_fixtures as mf  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SCRIPT = Path(os.environ.get("EXCEL_AI", str(ROOT / "scripts" / "excel_ai.py")))

# SKILL.md 规定的产物文件名
REPORT_MD = "验证报告.md"
DETAIL_CSV = "verification_detail.csv"
MISMATCH_CSV = "mismatches.csv"
OUTPUT_XLSX = "AI处理结果.xlsx"
ANALYSIS_MD = "分析报告.md"
INDEX_MD = "00-索引.md"

# SKILL.md「output」一节规定的三种配色
COLOR_INPUT = "DAEEF3"   # 蓝 = 原始输入
COLOR_AI = "E2EFDA"      # 绿 = AI 计算
COLOR_DIFF = "FCE4D6"    # 橙 = 差异非零

EXTRA_SHEETS = ["校验汇总", "数据血缘", "字段本体"]


# ===========================================================================
# 基础设施
# ===========================================================================

@pytest.fixture(autouse=True)
def _require_script(request):
    """被测脚本还没写出来时，除 test_00 外全部跳过，避免刷屏。"""
    if request.node.name.startswith("test_00"):
        return
    if not SCRIPT.exists():
        pytest.skip("被测脚本不存在：%s" % SCRIPT)


@pytest.fixture(scope="session")
def fx(tmp_path_factory) -> Path:
    """一次性生成全部「脏数据」夹具（session 级，10 个 xlsx）。"""
    d = tmp_path_factory.mktemp("夹具")
    mf.build_all(d)
    return d


def run(*args, cwd=None) -> subprocess.CompletedProcess:
    """调 CLI。返回值上挂一个 .out 属性 = stdout + stderr，方便断言。"""
    env = dict(os.environ)
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    cp = subprocess.run(
        [sys.executable, str(SCRIPT)] + [str(a) for a in args],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        cwd=str(cwd) if cwd else None, env=env, timeout=300,
    )
    cp.out = (cp.stdout or "") + "\n" + (cp.stderr or "")
    return cp


def run_ok(*args, cwd=None) -> subprocess.CompletedProcess:
    """调 CLI 并要求退出码为 0。"""
    cp = run(*args, cwd=cwd)
    assert cp.returncode == 0, (
        "命令应当成功退出（退出码 0），实际 %d。\n命令: %s\n输出:\n%s"
        % (cp.returncode, " ".join(str(a) for a in args), cp.out[-4000:])
    )
    return cp


def read_text(path) -> str:
    return Path(path).read_text(encoding="utf-8", errors="replace")


def write_spec(dirpath: Path, spec: dict, name: str = "spec.json") -> Path:
    dirpath.mkdir(parents=True, exist_ok=True)
    p = dirpath / name
    p.write_text(json.dumps(spec, ensure_ascii=False, indent=2), encoding="utf-8")
    return p


def read_csv_rows(path) -> list:
    """按 utf-8-sig 读 CSV，返回 list[dict]。"""
    text = Path(path).read_text(encoding="utf-8-sig", errors="replace")
    return list(csv.DictReader(io.StringIO(text)))


_NUM_RE = re.compile(r"-?\d+(?:\.\d+)?")


def numbers_in(text: str) -> list:
    """抠出文本里的所有数字（先去掉千分位逗号）。"""
    plain = re.sub(r"(?<=\d),(?=\d{3}\b)", "", text)
    out = []
    for tok in _NUM_RE.findall(plain):
        try:
            out.append(float(tok))
        except ValueError:
            pass
    return out


def has_number(text: str, expected: float, tol: float = 0.6) -> bool:
    """报告里是否出现了某个数（容忍千分位与 0~2 位小数的不同格式）。"""
    return any(abs(n - expected) <= tol for n in numbers_in(text))


def pick_file(outdir: Path, exact_name: str, suffix: str) -> Path:
    """优先按 SKILL.md 规定的文件名取；取不到就退化成同后缀唯一文件（便于定位真实问题）。"""
    outdir = Path(outdir)
    p = outdir / exact_name
    if p.exists():
        return p
    cands = sorted(x for x in outdir.rglob("*" + suffix) if x.is_file())
    assert cands, "输出目录 %s 里既没有 %s，也没有任何 %s 文件" % (outdir, exact_name, suffix)
    return cands[0]


def verify_artifacts(outdir: Path):
    """返回 (验证报告.md, verification_detail.csv, mismatches.csv 或 None)。"""
    report = pick_file(outdir, REPORT_MD, ".md")
    detail = Path(outdir) / DETAIL_CSV
    if not detail.exists():
        cands = [p for p in Path(outdir).rglob("*.csv") if "mismatch" not in p.name.lower()]
        assert cands, "verify 必须产出逐行明细 CSV（SKILL: verification_detail.csv），%s 里没有" % outdir
        detail = cands[0]
    mism = Path(outdir) / MISMATCH_CSV
    return report, detail, (mism if mism.exists() else None)


def find_sheet_md(outdir: Path, workbook_filename: str, sheet: str) -> Path:
    """在 tomd 输出目录里定位某个 Sheet 的 MD（SKILL: 01_raw_md/<文件名>/<Sheet>.md）。"""
    outdir = Path(outdir)
    mds = [p for p in outdir.rglob("*.md") if p.is_file()]
    assert mds, "tomd 没有在 %s 下产出任何 MD 文件" % outdir
    stem = Path(workbook_filename).stem
    body = [p for p in mds if "索引" not in p.name]
    by_name = [p for p in body if sheet in p.name]
    narrowed = [p for p in by_name if stem in str(p.parent) or stem in p.name]
    pool = narrowed or by_name or [p for p in body if sheet in read_text(p)]
    assert pool, (
        "找不到工作簿《%s》Sheet《%s》对应的 MD。SKILL Step1 规定产出 "
        "01_raw_md/<文件名>/<Sheet>.md。实际产出：%s" % (workbook_filename, sheet, [p.name for p in mds])
    )
    return pool[0]


def lines_with(text: str, *tokens) -> list:
    """返回同时包含全部 token 的行。"""
    return [ln for ln in text.splitlines() if all(t in ln for t in tokens)]


def find_col(fieldnames, *tokens, exclude=()) -> str:
    """在 CSV 表头里找同时含全部 token 的列名。"""
    for name in fieldnames or []:
        if all(t in name for t in tokens) and not any(e in name for e in exclude):
            return name
    raise AssertionError(
        "CSV 表头里找不到同时含 %s 的列。实际表头：%s" % (list(tokens), list(fieldnames or []))
    )


def key_col(fieldnames) -> str:
    """明细 CSV 里的主键列（spec 的 keys.ID → 列名 ID 或 工号）。"""
    for cand in ("ID", "工号"):
        for name in fieldnames or []:
            if name == cand:
                return name
    for name in fieldnames or []:
        if "ID" in name or "工号" in name:
            return name
    raise AssertionError("明细 CSV 里找不到主键列（spec.keys 里的 ID/工号）。表头：%s" % list(fieldnames or []))


def sha256(path: Path) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


# ===========================================================================
# spec 构造
# ===========================================================================

def salary_spec(fx: Path, filename: str, sheet: str, tolerance: float = 0.01,
                layered: bool = False, skip: bool = True) -> dict:
    """F01 / F10 共用的工资明细 spec（4 行表头、列名靠包含匹配命中）。"""
    checks = [{
        "name": "应发工资G",
        "target": "G_x",
        # SKILL Step4: checks 尽量展开到输入层，避免某层错了连锁污染
        "expr": "round(A + A2 + A3 + B1 + B2 - C1 - C2 + D + E - F, 2)",
        "tolerance": tolerance,
    }]
    if layered:
        # 用 derived 的中间量再算一遍：验证「先算 derived 再算 checks」且 derived 按声明顺序求值
        checks.append({
            "name": "应发工资G_分层",
            "target": "G_x",
            "expr": "round(AT + NET + D + E - F, 2)",
            "tolerance": tolerance,
        })
    spec = {
        "workbook": str(fx / filename),
        "sheet": sheet,
        "header_rows": mf.SALARY_HEADER_ROWS,
        "keys": {"ID": "工号", "NAME": "姓名"},
        "dimensions": {"DEPT": "一级部门"},
        "fields": {
            "A": "基本工资A", "A2": "岗位津贴A2", "A3": "工龄工资A3",
            "B1": "绩效工资B1", "B2": "加班费B2",
            "C1": "社保C1", "C2": "公积金C2",
            "D": "补贴D", "E": "奖金E", "F": "罚款F",
            "G_x": "应发工资G",
        },
        # 声明顺序有意义：NET 依赖前面声明的 B / C
        "derived": {"AT": "A + A2 + A3", "B": "B1 + B2", "C": "C1 + C2", "NET": "B - C"},
        "checks": checks,
    }
    if skip:
        spec["skip_when"] = {"empty": ["ID"], "label_in": ["汇总", "合计", "总计", "小计"]}
    return spec


def salary_expected_g() -> dict:
    return {row[0]: mf.salary_g(row) for row in mf.SALARY_ROWS}


# ===========================================================================
# 0. 前置
# ===========================================================================

def test_00_被测脚本必须存在():
    assert SCRIPT.exists(), (
        "SKILL.md「工具脚本从哪来」要求配套脚本落在 scripts/excel_ai.py（企业版随技能分发）。"
        "当前找不到：%s（可用环境变量 EXCEL_AI 指定路径）" % SCRIPT
    )


def test_00_夹具生成器可用(fx):
    missing = [name for name in mf.ALL_FIXTURES if not (fx / name).exists()]
    assert not missing, "夹具没生成全：%s" % missing


# ===========================================================================
# 1. tomd
# ===========================================================================

@pytest.fixture(scope="session")
def tomd_f01(fx, tmp_path_factory) -> Path:
    out = tmp_path_factory.mktemp("tomd_f01")
    run_ok("tomd", fx / mf.F01, "-o", out)
    return out


def test_tomd_多行表头逐层拼接且合并单元格被铺开(fx, tomd_f01):
    """SKILL: 用 range_boundaries 把合并单元格的值铺满覆盖区再拼多行列名。

    第 2 行 D2:F2 合并成"固定项"，第 3 行是"基本工资A"，第 4 行是单位"元"。
    列名必须拼成含这三段的一串，否则说明合并区没铺开或多行表头没拼。
    """
    md = read_text(find_sheet_md(tomd_f01, mf.F01, mf.S01))
    hit = lines_with(md, "固定项", "基本工资A")
    assert hit, (
        "第 4 列的拼接列名里应同时出现合并单元格分组名『固定项』与字段名『基本工资A』，"
        "MD 里找不到这样一行 —— 合并单元格没铺开或多行表头没拼接。\nMD 片段:\n%s" % md[:2500]
    )
    assert any("元" in ln for ln in hit), (
        "第 4 行的单位行『元』也属于表头（header_rows=4），必须一起拼进列名。实际行：%s" % hit
    )
    # 分组表头对其覆盖的每一列都要生效，不能只铺到第一列
    assert lines_with(md, "固定项", "工龄工资A3"), (
        "合并区 D2:F2 只铺到了首列：『工龄工资A3』（第 6 列）的列名里没有『固定项』"
    )


def test_tomd_跳过横跨整表的大标题行(fx, tomd_f01):
    """SKILL: 跳过"整行只有一个取值且横跨大半张表"的大标题行。"""
    md = read_text(find_sheet_md(tomd_f01, mf.F01, mf.S01))
    hit = lines_with(md, "固定项", "基本工资A")
    assert hit, "先要能拼出列名（见上一条用例）"
    for ln in hit:
        assert mf.SALARY_TITLE not in ln, (
            "第 1 行是横跨 A1:N1 的大标题，必须被跳过、不能拼进列名，实际列名行：%s" % ln
        )


def test_tomd_给出数据体起始行(fx, tomd_f01):
    """SKILL Step1: 脚本要启发式定位并给出数据体起始行（本夹具是第 5 行）。"""
    md = read_text(find_sheet_md(tomd_f01, mf.F01, mf.S01))
    hit = [ln for ln in md.splitlines()
           if ("起始" in ln or "开始" in ln or "起始行" in ln) and "5" in ln]
    assert hit, (
        "MD 里应给出数据体起始行的推断值（本夹具 4 行表头 → 数据体从第 5 行开始），"
        "没找到含『起始/开始』且含 5 的行。\nMD 片段:\n%s" % md[:2500]
    )


def test_tomd_抽出单元格里的真实公式(fx, tmp_path):
    """SKILL: openpyxl data_only=False 抽真实公式，并输出真实公式清单。"""
    out = tmp_path / "out"
    run_ok("tomd", fx / mf.F03, "-o", out)
    md = read_text(find_sheet_md(out, mf.F03, mf.S03))
    assert mf.F03_SAMPLE_FORMULA in md.replace(" ", ""), (
        "单元格 F2 里的真实公式 =%s 必须被抽进 MD（真实公式清单），"
        "MD 里没有。\nMD 片段:\n%s" % (mf.F03_SAMPLE_FORMULA, md[:2500])
    )


def test_tomd_混合类型列给出类型告警(fx, tmp_path):
    """SKILL: 列类型按 to_numeric 成功率判定，≥95% 判数值，≥50% 标"含非数值"告警。

    绩效系数COEF：12 行里 8 个数值 + 3 个文本 + 1 个空（66.7% / 72.7%），必须告警；
    绩效基数BASE：100% 数值，不许告警。
    """
    out = tmp_path / "out"
    run_ok("tomd", fx / mf.F04, "-o", out)
    md = read_text(find_sheet_md(out, mf.F04, mf.S04))

    coef_lines = lines_with(md, "绩效系数COEF")
    assert coef_lines, "MD 的列字段定义表里没有『绩效系数COEF』这一列"
    assert any(("含非数值" in ln or "⚠" in ln) for ln in coef_lines), (
        "绩效系数COEF 混了 `/`、`-`、`不适用`、空（数值占比约 67%%，落在 [50%%,95%%) 区间），"
        "必须标出『⚠️含非数值』告警。实际：%s" % coef_lines
    )

    base_lines = [ln for ln in lines_with(md, "绩效基数BASE") if "COEF" not in ln]
    assert base_lines, "MD 里没有『绩效基数BASE』这一列"
    assert not any("含非数值" in ln for ln in base_lines), (
        "绩效基数BASE 是 100%% 数值列，不该出现『含非数值』告警（告警不能一刀切）。实际：%s" % base_lines
    )


def test_tomd_列名里的公式原样保留(fx, tmp_path):
    """SKILL Step3: 公式经常不在单元格里而在列名里（应发工资G=A+B-C+D+E-F），照单全收。"""
    out = tmp_path / "out"
    run_ok("tomd", fx / mf.F02, "-o", out)
    md = read_text(find_sheet_md(out, mf.F02, mf.S02))
    assert mf.F02_RESULT_COL in md.replace(" ", ""), (
        "写在列名里的公式『%s』必须原样出现在 MD 里，实际没有。\nMD 片段:\n%s"
        % (mf.F02_RESULT_COL, md[:2500])
    )


def test_tomd_公式列无缓存值也不崩且照样抽公式(fx, tmp_path):
    """坑11：程序生成的 xlsx 没有缓存值，公式列显示"全空"。

    tomd 必须照常跑完（退出码 0）并仍然抽出公式本身（data_only=False 那一路）。
    """
    out = tmp_path / "out"
    cp = run_ok("tomd", fx / mf.F08, "-o", out)
    md = read_text(find_sheet_md(out, mf.F08, mf.S08))
    assert mf.F08_SAMPLE_FORMULA in md.replace(" ", ""), (
        "工作簿没有公式缓存值时，data_only=False 仍能读到公式 =%s，MD 必须把它列出来。\nMD 片段:\n%s"
        % (mf.F08_SAMPLE_FORMULA, md[:2000])
    )
    assert "Traceback" not in cp.out, "无缓存值的工作簿不该让 tomd 抛异常：\n%s" % cp.out[-2000:]


def test_tomd_多工作簿与索引文件(fx, tmp_path):
    """SKILL Step1: 支持一次转多个工作簿，并输出 00-索引.md。"""
    out = tmp_path / "out"
    run_ok("tomd", fx / mf.F01, fx / mf.F07, "-o", out)
    idx = [p for p in out.rglob("*.md") if p.name == INDEX_MD] or \
          [p for p in out.rglob("*.md") if "索引" in p.name]
    assert idx, "tomd 必须输出索引文件 %s，实际产出：%s" % (INDEX_MD, [p.name for p in out.rglob("*.md")])
    text = read_text(idx[0])
    for sheet in (mf.S01, mf.S07_MAIN, mf.S07_REF):
        assert sheet in text, "索引文件里应列出 Sheet《%s》，实际内容:\n%s" % (sheet, text[:2000])
    # 跨表工作簿的两个 Sheet 都要各出一份 MD
    find_sheet_md(out, mf.F07, mf.S07_MAIN)
    find_sheet_md(out, mf.F07, mf.S07_REF)


def test_tomd_可选参数被接受且sheets过滤生效(fx, tmp_path):
    """SKILL Step1 文档化的可选参数：--sheets / --header-scan / --preview-rows。"""
    out = tmp_path / "out"
    run_ok("tomd", fx / mf.F07, "--sheets", mf.S07_REF, "--header-scan", "8", "--preview-rows", "3", "-o", out)
    body = [p for p in out.rglob("*.md") if "索引" not in p.name]
    assert body, "指定 --sheets 后仍应产出被选 Sheet 的 MD"
    assert all(mf.S07_MAIN not in p.name for p in body), (
        "--sheets 只选了《%s》，却把《%s》也转了：%s" % (mf.S07_REF, mf.S07_MAIN, [p.name for p in body])
    )


def test_tomd_文件不存在应非零退出(tmp_path):
    cp = run("tomd", tmp_path / "根本没有这个文件.xlsx", "-o", tmp_path / "out")
    assert cp.returncode != 0, "输入文件不存在时 tomd 必须以非零退出码失败，实际退出码 0"


# ===========================================================================
# 2. verify —— 方法论的核心
# ===========================================================================

@pytest.fixture(scope="session")
def verify_f01(fx, tmp_path_factory):
    """F01（数据全对）跑一次 verify，session 复用。"""
    d = tmp_path_factory.mktemp("verify_f01")
    spec = write_spec(d, salary_spec(fx, mf.F01, mf.S01, layered=True))
    out = d / "03_verify"
    cp = run_ok("verify", spec, "-o", out)
    return spec, out, cp


@pytest.fixture(scope="session")
def verify_f10(fx, tmp_path_factory):
    """F10（3 行被人为覆盖）跑一次 verify，容差 0.01。"""
    d = tmp_path_factory.mktemp("verify_f10")
    spec = write_spec(d, salary_spec(fx, mf.F10, mf.S10, tolerance=0.01))
    out = d / "03_verify"
    cp = run_ok("verify", spec, "-o", out)
    return spec, out, cp


def test_verify_产物三件套齐全(verify_f10):
    """SKILL: verify 产出 验证报告.md + verification_detail.csv + mismatches.csv。"""
    _spec, out, _cp = verify_f10
    for name in (REPORT_MD, DETAIL_CSV, MISMATCH_CSV):
        assert (out / name).exists(), (
            "SKILL 规定 verify 要产出 %s，实际输出目录内容：%s" % (name, [p.name for p in out.iterdir()])
        )


def test_verify_正确的spec通过率100(verify_f01):
    """SKILL Step4: 通过率 100% 才算 AI 学会了这套业务逻辑。"""
    _spec, out, _cp = verify_f01
    report, detail, mism = verify_artifacts(out)
    text = read_text(report)
    assert re.search(r"100(?:\.0+)?\s*%", text), (
        "F01 的数据全部算得对，验证报告里必须出现 100%% 通过率。报告:\n%s" % text[:3000]
    )
    if mism is not None:
        rows = read_csv_rows(mism)
        assert not rows, "F01 没有任何错误行，mismatches.csv 却有 %d 条：%s" % (len(rows), rows[:3])
    rows = read_csv_rows(detail)
    assert len(rows) == len(mf.SALARY_ROWS), (
        "明细 CSV 应逐行覆盖全部 %d 名员工（全量数据验证），实际 %d 行"
        % (len(mf.SALARY_ROWS), len(rows))
    )


def test_verify_明细CSV的三列命名约定(verify_f01):
    """SKILL: verification_detail.csv 每个校验项三列 `名·AI`、`名·Excel`、`名·差异`。"""
    _spec, out, _cp = verify_f01
    _r, detail, _m = verify_artifacts(out)
    rows = read_csv_rows(detail)
    assert rows, "明细 CSV 是空的"
    names = list(rows[0].keys())
    for suffix in ("·AI", "·Excel", "·差异"):
        col = "应发工资G" + suffix
        assert col in names, (
            "校验项『应发工资G』应当出现三列 名·AI / 名·Excel / 名·差异，缺少 %s。实际表头：%s"
            % (col, names)
        )


def test_verify_derived按声明顺序求值(verify_f01):
    """SKILL: 先算 derived 再算 checks，derived 按声明顺序求值。

    NET = B - C 依赖前面声明的 B、C；用 NET 拼出来的第二个校验项必须同样 100% 通过。
    """
    _spec, out, _cp = verify_f01
    _r, detail, _m = verify_artifacts(out)
    rows = read_csv_rows(detail)
    diff_col = find_col(rows[0].keys(), "应发工资G_分层", "差异")
    bad = [r for r in rows if abs(float(r[diff_col] or 0)) > 0.01]
    assert not bad, (
        "用 derived 中间量（AT / B / C / NET，NET 依赖 B、C）拼出的校验项出现差异，"
        "说明 derived 没有按声明顺序求值、或 checks 先于 derived 求值。样例：%s" % bad[:2]
    )


def test_verify_逮出被人为覆盖的那几行(verify_f10):
    """SKILL Step4:「只有零星几行差 → 多半是 Excel 里的人为覆盖或错误，这恰恰是最有价值的发现」。

    F10 里 E003 的应发工资G 被人为 +100.00，必须被逮出来；
    E007 只差 0.005（在 0.01 金额容差内）不许误报；E009 差 0.05（超容差）也要逮出。
    """
    _spec, out, _cp = verify_f10
    _r, _d, mism = verify_artifacts(out)
    assert mism is not None, "F10 有算错的行，必须产出 mismatches.csv"
    text = read_text(mism)
    rows = read_csv_rows(mism)

    assert "E003" in text, "E003 的应发工资G 被人为 +100.00，必须出现在不匹配清单里。实际:\n%s" % text[:2000]
    assert "E009" in text, "E009 差 0.05 已超出 0.01 金额容差，必须记为不匹配。实际:\n%s" % text[:2000]
    assert "E007" not in text, (
        "E007 只差 0.005，在 0.01 金额容差之内，不该被判为不匹配（比较用容差不用 ==）。实际:\n%s" % text[:2000]
    )
    assert "E001" not in text, "E001 数据是对的，不该进不匹配清单。实际:\n%s" % text[:2000]
    assert len(rows) == 2, (
        "本 spec 只有 1 个校验项、只有 E003/E009 两行超容差，mismatches.csv 应为 2 条，实际 %d 条：%s"
        % (len(rows), rows[:4])
    )


def test_verify_报告里点名了算错的行(verify_f10):
    _spec, out, _cp = verify_f10
    report, _d, _m = verify_artifacts(out)
    text = read_text(report)
    assert "E003" in text or "赵敏" in text, (
        "SKILL 要求验证报告里给出「不匹配明细」，报告里应点名 E003/赵敏。报告:\n%s" % text[:3000]
    )


def test_verify_差异列的数值就是真实差额(verify_f10):
    """差异列要能一眼看出「差一个固定值」这种模式（SKILL: 看差异模式）。"""
    _spec, out, _cp = verify_f10
    _r, detail, _m = verify_artifacts(out)
    rows = read_csv_rows(detail)
    kcol = key_col(rows[0].keys())
    dcol = find_col(rows[0].keys(), "应发工资G", "差异", exclude=("分层",))
    diffs = {r[kcol]: float(r[dcol] or 0) for r in rows}
    assert abs(abs(diffs["E003"]) - 100.00) < 0.02, (
        "E003 的 Excel 值比正确值高 100.00，差异列应当是 ±100.00，实际 %s" % diffs.get("E003")
    )
    assert abs(diffs["E001"]) < 0.005, "E001 算得对，差异列应当≈0，实际 %s" % diffs.get("E001")
    assert abs(abs(diffs["E009"]) - 0.05) < 0.005, "E009 差异应≈0.05，实际 %s" % diffs.get("E009")


def test_verify_容差边界随spec生效(fx, tmp_path):
    """SKILL: 金额容差 0.01；这里把容差放大到 0.10，E009(差0.05) 应当被容差吃掉，E003(差100) 不行。"""
    spec = write_spec(tmp_path, salary_spec(fx, mf.F10, mf.S10, tolerance=0.10))
    out = tmp_path / "out"
    run_ok("verify", spec, "-o", out)
    _r, _d, mism = verify_artifacts(out)
    assert mism is not None, "仍有 E003 不匹配，必须产出 mismatches.csv"
    text = read_text(mism)
    rows = read_csv_rows(mism)
    assert "E003" in text, "容差 0.10 也吃不掉 E003 的 100.00 差额"
    assert "E009" not in text, (
        "E009 只差 0.05，在 spec 声明的 0.10 容差之内，不该记为不匹配 —— 容差没有按 spec 生效。实际:\n%s" % text[:1500]
    )
    assert len(rows) == 1, "容差 0.10 时只剩 E003 一条不匹配，实际 %d 条" % len(rows)


def test_verify_重名列必须报错并列出候选(fx, tmp_path):
    """SKILL Step4:「列名匹配到多列会直接报错并列出候选，这时改用列号，别去猜」。"""
    spec = write_spec(tmp_path, {
        "workbook": str(fx / mf.F09),
        "sheet": mf.S09,
        "header_rows": 1,
        "keys": {"ID": "工号"},
        "fields": {"AMT": mf.F09_DUP_NAME, "R": "结果"},
        "checks": [{"name": "金额", "target": "R", "expr": "AMT", "tolerance": 0.001}],
    })
    out = tmp_path / "out"
    cp = run("verify", spec, "-o", out)
    assert cp.returncode != 0, (
        "第 5、6 两列都叫『%s』，按列名引用属于歧义，必须直接报错退出（非零退出码），"
        "而不是随便挑一列继续算。实际退出码 0，输出:\n%s" % (mf.F09_DUP_NAME, cp.out[-2000:])
    )
    assert mf.F09_DUP_NAME in cp.out, "报错信息里应指明是哪个列名歧义（%s）。实际:\n%s" % (mf.F09_DUP_NAME, cp.out[-2000:])
    assert any(w in cp.out for w in ("候选", "candidates", "列号", "多列", "歧义")), (
        "报错信息必须列出候选列（SKILL 原文：报错并列出候选），实际输出:\n%s" % cp.out[-2000:]
    )


def test_verify_精确匹配优先于包含匹配(fx, tmp_path):
    """SKILL: 列引用「先精确后包含匹配」。

    夹具里同时有『工资』和『基本工资』两列：按『工资』引用时精确匹配应当胜出，
    既不能报歧义，也不能匹配到『基本工资』（值 1000，一对就露馅）。
    """
    spec = write_spec(tmp_path, {
        "workbook": str(fx / mf.F09),
        "sheet": mf.S09,
        "header_rows": 1,
        "keys": {"ID": "工号"},
        "fields": {"W": mf.F09_EXACT_NAME, "R": "结果"},
        "checks": [{"name": "结果应等于工资列", "target": "R", "expr": "W", "tolerance": 0.001}],
    })
    out = tmp_path / "out"
    cp = run("verify", spec, "-o", out)
    assert cp.returncode == 0, (
        "『工资』在表里有精确同名列，应当精确命中而不是按包含匹配判成歧义。输出:\n%s" % cp.out[-2000:]
    )
    _r, detail, mism = verify_artifacts(out)
    rows = read_csv_rows(detail)
    assert len(rows) == len(mf.F09_ROWS)
    dcol = find_col(rows[0].keys(), "结果应等于工资列", "差异")
    bad = [r for r in rows if abs(float(r[dcol] or 0)) > 0.001]
    assert not bad, (
        "『结果』列的值等于『工资』列（%.2f）而不是『基本工资』（%.2f）；出现差异说明包含匹配抢在了精确匹配前面。样例：%s"
        % (mf.F09_WAGE, mf.F09_BASE_WAGE, bad[:2])
    )


def test_verify_支持用列号引用列(fx, tmp_path):
    """SKILL: 列引用支持列号或列名；重名时改用列号。

    第 5/6 两列都是『金额』（7.00 / 9.00）。不论列号按 0 基还是 1 基解释，
    引用 5 都应落在这两列之一 —— 关键是不许再报歧义、也不许落到别的列上。
    """
    spec = write_spec(tmp_path, {
        "workbook": str(fx / mf.F09),
        "sheet": mf.S09,
        "header_rows": 1,
        "keys": {"ID": "工号"},
        "fields": {"AMT": 5},
        "checks": [{
            "name": "列号解析",
            "target": "AMT",
            # 三元 + 比较，命中任一『金额』列都返回它自己；落到别的列则返回 -1 立刻暴露
            "expr": "%.1f if AMT == %.1f else (%.1f if AMT == %.1f else -1.0)"
                    % (mf.F09_AMOUNT_1, mf.F09_AMOUNT_1, mf.F09_AMOUNT_2, mf.F09_AMOUNT_2),
            "tolerance": 0.001,
        }],
    })
    out = tmp_path / "out"
    cp = run("verify", spec, "-o", out)
    assert cp.returncode == 0, "用列号引用重名列应当能正常跑通。输出:\n%s" % cp.out[-2000:]
    _r, detail, _m = verify_artifacts(out)
    rows = read_csv_rows(detail)
    dcol = find_col(rows[0].keys(), "列号解析", "差异")
    bad = [r for r in rows if abs(float(r[dcol] or 0)) > 0.001]
    assert not bad, (
        "列号 5 没有解析到任何一列『金额』（应当是 %.2f 或 %.2f）。样例：%s"
        % (mf.F09_AMOUNT_1, mf.F09_AMOUNT_2, bad[:2])
    )


def test_verify_skip_when排掉汇总行金额才不翻倍(fx, tmp_path):
    """坑6：小计/合计/总计 混在数据体里，不排掉金额直接翻倍。"""
    spec = write_spec(tmp_path, {
        "workbook": str(fx / mf.F05),
        "sheet": mf.S05,
        "header_rows": 1,
        "keys": {"ID": "工号", "NAME": "姓名"},
        "dimensions": {"DEPT": "一级部门"},
        "fields": {"A": "基本工资A", "B": "浮动合计B", "C": "扣款合计C", "G_x": "应发工资G"},
        "checks": [{"name": "应发工资G", "target": "G_x", "expr": "round(A + B - C, 2)", "tolerance": 0.01}],
        "skip_when": {"empty": ["ID"], "label_in": ["汇总", "合计", "总计", "小计"]},
    })
    out = tmp_path / "out"
    run_ok("verify", spec, "-o", out)
    report, detail, _m = verify_artifacts(out)
    rows = read_csv_rows(detail)
    assert rows, "明细 CSV 是空的：skip_when 把所有行都排掉了？"
    assert len(rows) == len(mf.F05_EMPLOYEES), (
        "数据体里有 2 行小计 + 1 行合计（其中合计行的工号写的是『合计』而非空），"
        "skip_when 之后应当只剩 %d 名员工，实际 %d 行：%s"
        % (len(mf.F05_EMPLOYEES), len(rows), [r.get(key_col(rows[0].keys())) for r in rows])
    )
    acol = find_col(rows[0].keys(), "应发工资G", "AI")
    total = round(sum(float(r[acol] or 0) for r in rows), 2)
    assert abs(total - mf.F05_EMPLOYEE_TOTAL) < 0.05, (
        "汇总行被算进去会让金额翻倍：应发工资G 合计应为 %.2f，实际 %.2f"
        % (mf.F05_EMPLOYEE_TOTAL, total)
    )
    text = read_text(report)
    assert has_number(text, mf.F05_EMPLOYEE_TOTAL), (
        "验证报告的「汇总额对比」里应出现 %.2f（排掉汇总行后的真实合计）。报告:\n%s"
        % (mf.F05_EMPLOYEE_TOTAL, text[:3000])
    )


def test_verify_非数值单元格按0参与运算(fx, tmp_path):
    """SKILL: 非数值单元格按 0.0 参与运算（`/`、`-`、"不适用"、空 表示不参加）。"""
    spec = write_spec(tmp_path, {
        "workbook": str(fx / mf.F04),
        "sheet": mf.S04,
        "header_rows": 1,
        "keys": {"ID": "工号", "NAME": "姓名"},
        "fields": {"COEF": "绩效系数COEF", "BASE": "绩效基数BASE", "B1": "绩效工资B1"},
        "checks": [{"name": "绩效工资B1", "target": "B1", "expr": "round(COEF * BASE, 2)", "tolerance": 0.01}],
    })
    out = tmp_path / "out"
    run_ok("verify", spec, "-o", out)
    _r, detail, mism = verify_artifacts(out)
    rows = read_csv_rows(detail)
    assert len(rows) == len(mf.F04_ROWS)
    if mism is not None:
        bad = read_csv_rows(mism)
        assert not bad, (
            "`/`、`-`、`不适用`、空 必须按 0.0 参与运算（对应 Excel 里绩效工资=0），"
            "出现不匹配说明非数值单元格没按 0.0 处理：%s" % bad[:3]
        )


def test_verify_check列可直接当验证锚点(fx, tmp_path):
    """SKILL 坑4：check 列是原作者留下的断言，优先拿来做验证锚点。"""
    spec = write_spec(tmp_path, {
        "workbook": str(fx / mf.F06),
        "sheet": mf.S06,
        "header_rows": 1,
        "keys": {"ID": "工号", "NAME": "姓名"},
        "fields": {"A": "基本工资A", "B": "浮动合计B", "C": "扣款合计C",
                   "G_x": "应发工资G", "CHK": "核对check"},
        "checks": [
            {"name": "作者check列应恒为0", "target": "CHK", "expr": "0", "tolerance": 0.001},
            {"name": "应发工资G", "target": "G_x", "expr": "round(A + B - C, 2)", "tolerance": 0.01},
        ],
    })
    out = tmp_path / "out"
    run_ok("verify", spec, "-o", out)
    report, _d, mism = verify_artifacts(out)
    if mism is not None:
        assert not read_csv_rows(mism), "check 列恒为 0、G 也算得对，不该有不匹配"
    assert "作者check列应恒为0" in read_text(report), (
        "SKILL 要求验证报告里有「校验项定义」一节，应当列出每个校验项的名字"
    )


def test_verify_跨表一致性检查(fx, tmp_path):
    """SKILL Step4 cross_checks：跨表传递一致性。E004 在两张表里的绩效系数不一致，必须被点名。"""
    spec = write_spec(tmp_path, {
        "workbook": str(fx / mf.F07),
        "sheet": mf.S07_MAIN,
        "header_rows": 1,
        "keys": {"ID": "工号", "NAME": "姓名"},
        "fields": {"COEF": "绩效系数COEF", "BASE": "绩效基数", "B1": "绩效工资B1"},
        "checks": [{"name": "绩效工资B1", "target": "B1", "expr": "round(COEF * BASE, 2)", "tolerance": 0.01}],
        "cross_checks": [{
            "name": "绩效系数传递",
            "workbook": str(fx / mf.F07),
            "sheet": mf.S07_REF,
            "header_rows": 1,
            "key": "工号",
            "left_key": "ID",
            "compare": [{"left": "COEF", "right": "绩效系数", "tolerance": 0.001}],
        }],
    })
    out = tmp_path / "out"
    run_ok("verify", spec, "-o", out)
    report, _d, _m = verify_artifacts(out)
    text = read_text(report)
    assert "绩效系数传递" in text, (
        "SKILL 规定验证报告里要有「跨表一致性」一节，应当出现 cross_check 的名字。报告:\n%s" % text[:3000]
    )
    assert mf.F07_INCONSISTENT_ID in text, (
        "%s 在工资表里系数 1.00、绩效表里 1.30，跨表一致性检查必须点名它。报告:\n%s"
        % (mf.F07_INCONSISTENT_ID, text[:3000])
    )


# --- 表达式求值器：白名单语法必须支持 -------------------------------------

@pytest.mark.parametrize("label,expr", [
    ("加减乘除", "(A + B - C) * 1 / 1 - (A + B - C)"),
    ("幂与取模", "0 ** 2 + 4 % 2"),
    ("round/abs", "round(abs(0.0), 2)"),
    ("min/max", "min(0, 1) + max(-1, 0)"),
    ("int/float", "int(0.9) + float(0)"),
    ("floor/ceil", "floor(0.5) + ceil(0.0)"),
    ("三元表达式", "0 if A > 0 else 1"),
    ("比较与布尔", "0 if (A > 0 and B >= 0 or not (C < 0)) else 1"),
])
def test_verify_白名单表达式语法(fx, tmp_path, label, expr):
    """SKILL: 表达式语法 `+ - * / ** %`、比较与 and/or/not、三元、白名单函数
    abs round min max int float sum floor ceil。上述写法都必须能求值且结果为 0。"""
    spec = write_spec(tmp_path, {
        "workbook": str(fx / mf.F06),
        "sheet": mf.S06,
        "header_rows": 1,
        "keys": {"ID": "工号"},
        "fields": {"A": "基本工资A", "B": "浮动合计B", "C": "扣款合计C", "CHK": "核对check"},
        "checks": [{"name": "语法_" + label.replace("/", "或"), "target": "CHK",
                    "expr": expr, "tolerance": 0.001}],
    }, name="spec_%s.json" % label.replace("/", "_"))
    out = tmp_path / "out"
    cp = run("verify", spec, "-o", out)
    assert cp.returncode == 0, (
        "白名单语法「%s」(%s) 必须被求值器支持，实际报错:\n%s" % (label, expr, cp.out[-1500:])
    )
    _r, detail, mism = verify_artifacts(out)
    if mism is not None:
        assert not read_csv_rows(mism), "「%s」(%s) 的结果应恒为 0（check 列也是 0），却出现了不匹配" % (label, expr)


# --- 表达式求值器：安全性 ---------------------------------------------------

def _danger_spec(fx: Path, tmp_path: Path, expr: str) -> Path:
    return write_spec(tmp_path, {
        "workbook": str(fx / mf.F06),
        "sheet": mf.S06,
        "header_rows": 1,
        "keys": {"ID": "工号"},
        "fields": {"A": "基本工资A", "CHK": "核对check"},
        "checks": [{"name": "恶意表达式", "target": "CHK", "expr": expr, "tolerance": 0.001}],
    })


@pytest.mark.parametrize("label,expr_tpl", [
    ("导入并执行系统命令", "__import__('os').system('touch {sentinel}')"),
    ("open写文件", "open('{sentinel}', 'w')"),
    ("open读系统文件", "open('/etc/passwd')"),
    ("属性访问", "A.__class__"),
    ("dunder链逃逸", "().__class__.__bases__"),
    ("下标访问", "A[0]"),
    ("eval嵌套", "eval('1+1')"),
    ("非白名单函数", "len('abc')"),
    ("exec", "exec('x=1')"),
])
def test_verify_表达式求值器拒绝危险写法(fx, tmp_path, label, expr_tpl):
    """SKILL: 表达式用 ast 白名单求值 —— 禁止属性访问、下标、导入、其他调用。

    这条是企业版分发的底线：spec.json 可能来自任何人，求值器绝不能变成任意代码执行入口。
    """
    sentinel = tmp_path / "pwned.txt"
    expr = expr_tpl.format(sentinel=str(sentinel).replace("\\", "/"))
    spec = _danger_spec(fx, tmp_path, expr)
    out = tmp_path / "out"
    cp = run("verify", spec, "-o", out)

    assert not sentinel.exists(), (
        "严重安全缺陷：表达式「%s」被真的执行了（%s 已被创建）。"
        "SKILL 要求 ast 白名单求值，禁止导入/调用/属性访问。" % (expr, sentinel)
    )
    assert cp.returncode != 0, (
        "表达式「%s」（%s）属于 SKILL 明令禁止的写法，verify 必须拒绝并以非零退出码失败，"
        "实际退出码 0。输出:\n%s" % (expr, label, cp.out[-2000:])
    )


def test_verify_危险表达式不产生可用产物(fx, tmp_path):
    """被拒绝的 spec 不应留下"看上去验证通过了"的明细 CSV，避免下游 output/analyze 误用。"""
    spec = _danger_spec(fx, tmp_path, "A.__class__")
    out = tmp_path / "out"
    run("verify", spec, "-o", out)
    detail = out / DETAIL_CSV
    if detail.exists():
        rows = read_csv_rows(detail)
        assert not rows, "求值被拒绝后不该产出有内容的 %s（实际 %d 行）" % (DETAIL_CSV, len(rows))


# --- verify 的通用行为 ------------------------------------------------------

def test_verify_有不匹配时仍应正常退出0(verify_f10):
    """不匹配是业务发现（可能是 Excel 自己算错），不是工具错误；
    而且 Step5 的 output/analyze 要靠这次产出的明细 CSV 继续跑，所以必须正常退出。"""
    _spec, _out, cp = verify_f10
    assert cp.returncode == 0, "verify 逮到不匹配属于正常产出，应退出码 0，实际 %d" % cp.returncode


def test_verify_CSV一律utf8sig(verify_f10):
    """坑9：CSV 一律 utf-8-sig，否则用户用 Excel 打开是乱码。"""
    _spec, out, _cp = verify_f10
    _r, detail, mism = verify_artifacts(out)
    for path in [p for p in (detail, mism) if p is not None]:
        raw = Path(path).read_bytes()
        assert raw.startswith(b"\xef\xbb\xbf"), (
            "%s 必须用 utf-8-sig 编码（带 BOM），实际开头字节：%r" % (Path(path).name, raw[:8])
        )
        text = raw.decode("utf-8-sig")
        assert "应发工资G" in text, "CSV 里的中文列名解码后应可读，实际表头：%s" % text.splitlines()[:1]


def test_verify_spec不存在应非零退出(tmp_path):
    cp = run("verify", tmp_path / "没有这个spec.json", "-o", tmp_path / "out")
    assert cp.returncode != 0, "spec 文件不存在时 verify 必须非零退出"


def test_verify_spec格式错误应非零退出(tmp_path):
    bad = tmp_path / "bad.json"
    bad.write_text("{ 这不是合法的 JSON ", encoding="utf-8")
    cp = run("verify", bad, "-o", tmp_path / "out")
    assert cp.returncode != 0, "spec 不是合法 JSON 时 verify 必须非零退出"


def test_verify_header_rows决定数据体起点(fx, tmp_path):
    """SKILL:「header_rows 表头占几行，数错一行全废」——不带 skip_when 时，
    数据体应严格从第 header_rows+1 行开始，4 行表头不能被当成数据行。"""
    spec = write_spec(tmp_path, salary_spec(fx, mf.F01, mf.S01, skip=False))
    out = tmp_path / "out"
    run_ok("verify", spec, "-o", out)
    _r, detail, _m = verify_artifacts(out)
    rows = read_csv_rows(detail)
    assert len(rows) == len(mf.SALARY_ROWS), (
        "header_rows=4 时数据体从第 5 行开始、共 %d 行；实际验证了 %d 行 —— "
        "表头行被当成数据行（或反过来漏了数据行）。" % (len(mf.SALARY_ROWS), len(rows))
    )


def test_verify_明细CSV带上keys维度与字段列(verify_f01):
    """Step5 的 analyze 是拿 detail CSV 当输入的（group_by/top_n/rules/what_if 都按
    spec 里的字段名引用），所以 keys / dimensions / fields 必须原样落进明细 CSV。"""
    _spec, out, _cp = verify_f01
    _r, detail, _m = verify_artifacts(out)
    rows = read_csv_rows(detail)
    assert rows, "明细 CSV 是空的"
    names = list(rows[0].keys())
    wanted = {"ID": "工号", "NAME": "姓名", "DEPT": "一级部门", "A": "基本工资A", "G_x": "应发工资G"}
    for alias, cn in wanted.items():
        ok = any(n == alias or cn in n for n in names)
        assert ok, (
            "明细 CSV 里找不到 %s（spec 别名 %s / 原列名 %s）；"
            "analyze 的 group_by / top_n / rules / what_if 都要靠这些列。实际表头：%s"
            % (alias, alias, cn, names)
        )


# ===========================================================================
# 3. output
# ===========================================================================

LINEAGE_TOKEN = "血缘锚点TOKEN"
ONTOLOGY_TOKEN = "本体锚点TOKEN"


@pytest.fixture(scope="session")
def output_run(fx, tmp_path_factory):
    """用 F10（有差异）跑 verify → output，保证三种配色都有机会出现。"""
    d = tmp_path_factory.mktemp("output")
    spec_dict = salary_spec(fx, mf.F10, mf.S10)
    spec_dict["lineage"] = [
        {"源": "绩效表.绩效系数", "目标": "工资明细表.绩效工资B1", "连接键": "工号",
         "说明": "月考月发绩效系数×绩效基数 " + LINEAGE_TOKEN},
        {"源": "调薪表.基本工资", "目标": "工资明细表.基本工资A", "连接键": "工号", "说明": "调薪生效月起"},
    ]
    spec_dict["ontology"] = [
        {"字段": "应发工资G", "业务含义": "税前应发合计 " + ONTOLOGY_TOKEN,
         "类型": "金额", "计算关系": "A+A2+A3+B-C+D+E-F", "所属表": "工资明细表", "角色": "最终输出"},
        {"字段": "基本工资A", "业务含义": "月度基本工资", "类型": "金额",
         "计算关系": "人工输入", "所属表": "工资明细表", "角色": "人工输入"},
    ]
    spec = write_spec(d, spec_dict)
    vout = d / "03_verify"
    run_ok("verify", spec, "-o", vout)
    _r, detail, _m = verify_artifacts(vout)
    oout = d / "04_output"
    cp = run_ok("output", spec, "-d", detail, "-o", oout)
    return spec, detail, oout, cp


def _load_result_wb(outdir: Path):
    from openpyxl import load_workbook
    path = pick_file(outdir, OUTPUT_XLSX, ".xlsx")
    return path, load_workbook(path)


def test_output_生成结果xlsx(output_run):
    _spec, _detail, oout, _cp = output_run
    assert (oout / OUTPUT_XLSX).exists(), (
        "SKILL Step5 规定 output 产出 %s，实际输出目录：%s" % (OUTPUT_XLSX, [p.name for p in oout.iterdir()])
    )


def test_output_三个附加Sheet都在(output_run):
    """SKILL: 另附「校验汇总」「数据血缘」「字段本体」三个 Sheet。"""
    _spec, _detail, oout, _cp = output_run
    _path, wb = _load_result_wb(oout)
    for name in EXTRA_SHEETS:
        assert any(name in s for s in wb.sheetnames), (
            "结果 xlsx 里缺少「%s」Sheet，实际 Sheet：%s" % (name, wb.sheetnames)
        )


def test_output_三种配色都出现(output_run):
    """SKILL: 蓝 DAEEF3=原始输入、绿 E2EFDA=AI 计算、橙 FCE4D6=差异非零。"""
    _spec, _detail, oout, _cp = output_run
    _path, wb = _load_result_wb(oout)
    seen = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                fill = cell.fill
                for color in (getattr(fill, "start_color", None), getattr(fill, "fgColor", None)):
                    rgb = getattr(color, "rgb", None)
                    if isinstance(rgb, str) and len(rgb) >= 6:
                        seen.add(rgb[-6:].upper())
    for code, meaning in ((COLOR_INPUT, "蓝=原始输入"), (COLOR_AI, "绿=AI计算"), (COLOR_DIFF, "橙=差异非零")):
        assert code in seen, (
            "结果表里没出现配色 %s（%s）。SKILL Step5 明确规定了这三种颜色；"
            "本次数据里 E003/E009 存在差异，橙色必须出现。实际出现的填充色：%s"
            % (code, meaning, sorted(seen))
        )


def test_output_血缘与本体内容写进对应Sheet(output_run):
    """SKILL: 三个附加 Sheet 的内容取自 spec 的 lineage / ontology 段。"""
    _spec, _detail, oout, _cp = output_run
    _path, wb = _load_result_wb(oout)
    blob = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    blob.append(str(v))
    text = "\n".join(blob)
    assert LINEAGE_TOKEN in text, "spec.lineage 的内容没写进「数据血缘」Sheet（找不到锚点 %s）" % LINEAGE_TOKEN
    assert ONTOLOGY_TOKEN in text, "spec.ontology 的内容没写进「字段本体」Sheet（找不到锚点 %s）" % ONTOLOGY_TOKEN
    assert "应发工资G" in text, "公式说明区/校验汇总里应出现校验项『应发工资G』"


def test_output_明细数据落进结果表(output_run):
    _spec, _detail, oout, _cp = output_run
    _path, wb = _load_result_wb(oout)
    text = "\n".join(
        str(v) for ws in wb.worksheets for row in ws.iter_rows(values_only=True) for v in row if v is not None
    )
    for emp in ("E001", "E003", "高峰"):
        assert emp in text, "结果表里应包含逐行明细（缺 %s）" % emp


def test_output_明细CSV不存在应非零退出(fx, tmp_path):
    spec = write_spec(tmp_path, salary_spec(fx, mf.F01, mf.S01))
    cp = run("output", spec, "-d", tmp_path / "没有.csv", "-o", tmp_path / "out")
    assert cp.returncode != 0, "-d 指定的明细 CSV 不存在时 output 必须非零退出"


# ===========================================================================
# 4. analyze
# ===========================================================================

WHATIF_NAME = "基本工资上调10%"
RULE_NAME = "有罚款F的人"


def _analysis_section() -> dict:
    return {
        "group_by": [{"dim": "DEPT", "metrics": ["A", "G_x"], "count_as": "人数"}],
        "distributions": [{"dim": "DEPT"}],
        "top_n": [{"field": "G_x", "n": 3, "label": "NAME"}],
        "outliers": [{"field": "G_x", "method": "iqr", "k": 1.5, "label": "NAME"},
                     {"field": "G_x", "method": "zscore", "z": 2.0, "label": "NAME"}],
        "rules": [{"name": RULE_NAME, "when": "F > 0", "label": "NAME"}],
        "what_if": [{
            "name": WHATIF_NAME,
            "set": {"A": "round(A * 1.1, 2)"},
            "recompute": ["AT", "B", "C", "NET"],
            "targets": [{"name": "应发G", "expr": "AT + NET + D + E - F"}],
            "label": "NAME",
        }],
    }


@pytest.fixture(scope="session")
def analyze_run(fx, tmp_path_factory):
    d = tmp_path_factory.mktemp("analyze")
    spec_dict = salary_spec(fx, mf.F01, mf.S01)
    spec_dict["analysis"] = _analysis_section()
    spec = write_spec(d, spec_dict)
    vout = d / "03_verify"
    run_ok("verify", spec, "-o", vout)
    _r, detail, _m = verify_artifacts(vout)
    aout = d / "04_output"
    cp = run_ok("analyze", spec, "-d", detail, "-o", aout)
    return spec, detail, aout, cp


def analysis_text(aout: Path) -> str:
    return read_text(pick_file(aout, ANALYSIS_MD, ".md"))


def test_analyze_生成分析报告(analyze_run):
    _spec, _detail, aout, _cp = analyze_run
    assert (aout / ANALYSIS_MD).exists(), (
        "SKILL Step5 规定 analyze 产出 %s，实际：%s" % (ANALYSIS_MD, [p.name for p in aout.iterdir()])
    )


def test_analyze_分组汇总数值正确(analyze_run):
    """SKILL: 分组汇总（数量/总额/人均/占比）。"""
    text = analysis_text(analyze_run[2])
    assert "人数" in text, "group_by 声明了 count_as=人数，报告里应出现该表头"
    from collections import defaultdict
    by_dept = defaultdict(list)
    for row in mf.SALARY_ROWS:
        by_dept[row[2]].append(row)
    for dept, rows in by_dept.items():
        assert dept in text, "分组汇总里缺部门『%s』" % dept
        sum_a = round(sum(r[3] for r in rows), 2)
        sum_g = round(sum(mf.salary_g(r) for r in rows), 2)
        assert has_number(text, sum_a), "部门『%s』的基本工资A 合计应为 %.2f，报告里没有这个数" % (dept, sum_a)
        assert has_number(text, sum_g), "部门『%s』的应发工资G 合计应为 %.2f，报告里没有这个数" % (dept, sum_g)


def test_analyze_取值分布段存在(analyze_run):
    text = analysis_text(analyze_run[2])
    assert "分布" in text, "SKILL 要求分析报告含「取值分布」一段，报告里找不到『分布』"


def test_analyze_TopN点名正确(analyze_run):
    """SKILL: Top/Bottom。按应发工资G 取前 3。"""
    text = analysis_text(analyze_run[2])
    ranked = sorted(mf.SALARY_ROWS, key=lambda r: mf.salary_g(r), reverse=True)
    top3 = [r[1] for r in ranked[:3]]
    assert "Top" in text or "top" in text or "前" in text, "报告里应有 Top/Bottom 段落"
    for name in top3:
        assert name in text, "应发工资G 前三名是 %s，报告里缺『%s』" % (top3, name)


def test_analyze_离群点被逮出(analyze_run):
    """SKILL: 异常检测（IQR / Z-score 离群）。高峰的应发工资G 是其他人的 2~6 倍，
    在任何常见四分位口径与 z>2.0 下都是离群点。"""
    text = analysis_text(analyze_run[2])
    assert ("离群" in text or "异常" in text), "报告里应有离群/异常检测一段"
    outlier_name = max(mf.SALARY_ROWS, key=lambda r: mf.salary_g(r))[1]
    assert outlier_name in text, "IQR(k=1.5) 与 Z-score(z=2.0) 都应逮出『%s』" % outlier_name


def test_analyze_自定义规则命中(analyze_run):
    """SKILL: 自定义规则（表达式，复用同一个安全求值器）。"""
    text = analysis_text(analyze_run[2])
    assert RULE_NAME in text, "规则名『%s』应出现在异常检测一节" % RULE_NAME
    hit = [r[1] for r in mf.SALARY_ROWS if r[12] > 0]
    for name in hit:
        assert name in text, "规则 `F > 0` 命中 %s，报告里缺『%s』" % (hit, name)


def test_analyze_WhatIf合计行数值正确(analyze_run):
    """SKILL: What-If —— set 改动 → recompute 重算中间量 → 对 targets 出「现状/情景/变化」三组列 + 合计行。

    情景：基本工资A 上调 10%%。应发G = AT + NET + D + E - F，只有 AT 里的 A 变了，
    所以合计变化就是 Σ round(A*1.1,2) - Σ A。
    """
    text = analysis_text(analyze_run[2])
    assert WHATIF_NAME in text, "报告里应出现 What-If 情景名『%s』" % WHATIF_NAME
    for word in ("现状", "情景", "变化"):
        assert word in text, "What-If 必须给出「现状/情景/变化」三组列，缺『%s』" % word

    base_total = round(sum(mf.salary_g(r) for r in mf.SALARY_ROWS), 2)
    scen_total = 0.0
    for r in mf.SALARY_ROWS:
        (_id, _n, _d, a, a2, a3, b1, b2, c1, c2, dd, e, f) = r
        a_new = round(a * 1.1, 2)
        scen_total += round(a_new + a2 + a3 + b1 + b2 - c1 - c2 + dd + e - f, 2)
    scen_total = round(scen_total, 2)
    delta = round(scen_total - base_total, 2)

    assert has_number(text, base_total), "What-If 合计行的『现状』应为 %.2f，报告里没有" % base_total
    assert has_number(text, scen_total), "What-If 合计行的『情景』应为 %.2f，报告里没有" % scen_total
    assert has_number(text, delta), "What-If 合计行的『变化』应为 %.2f，报告里没有" % delta


def test_analyze_规则表达式同样受安全求值器约束(fx, tmp_path):
    """SKILL: 规则异常「复用同一个安全求值器」——那就必须同样拒绝危险写法。"""
    spec_dict = salary_spec(fx, mf.F01, mf.S01)
    spec = write_spec(tmp_path, spec_dict, name="spec_v.json")
    vout = tmp_path / "03_verify"
    run_ok("verify", spec, "-o", vout)
    _r, detail, _m = verify_artifacts(vout)

    sentinel = tmp_path / "pwned_rule.txt"
    spec_dict["analysis"] = {
        "rules": [{"name": "注入", "when": "__import__('os').system('touch %s') == 0"
                                          % str(sentinel).replace("\\", "/"), "label": "NAME"}]
    }
    bad_spec = write_spec(tmp_path, spec_dict, name="spec_bad.json")
    cp = run("analyze", bad_spec, "-d", detail, "-o", tmp_path / "04_output")
    assert not sentinel.exists(), (
        "严重安全缺陷：analyze 的规则表达式被真的执行了（%s 已创建）。规则必须复用 ast 白名单求值器。" % sentinel
    )
    assert cp.returncode != 0, (
        "规则里的危险表达式必须被拒绝并非零退出。输出:\n%s" % cp.out[-2000:]
    )


# ===========================================================================
# 5. 通用纪律
# ===========================================================================

def test_全流程绝不修改用户原表(fx, tmp_path):
    """坑10：不要改用户的原表 —— 所有产物写到独立输出目录。

    跑完 tomd/verify/output/analyze 之后，逐个比对原文件的 mtime 与 sha256。
    """
    work = tmp_path / "原表"
    work.mkdir()
    for name in mf.ALL_FIXTURES:
        shutil.copy2(fx / name, work / name)

    before = {p.name: (p.stat().st_mtime_ns, p.stat().st_size, sha256(p))
              for p in sorted(work.glob("*.xlsx"))}

    run_ok("tomd", *[work / n for n in mf.ALL_FIXTURES], "-o", tmp_path / "01_raw_md")

    spec_dict = salary_spec(work, mf.F10, mf.S10)
    spec_dict["analysis"] = _analysis_section()
    spec = write_spec(tmp_path, spec_dict)
    run_ok("verify", spec, "-o", tmp_path / "03_verify")
    _r, detail, _m = verify_artifacts(tmp_path / "03_verify")
    run_ok("output", spec, "-d", detail, "-o", tmp_path / "04_output")
    run_ok("analyze", spec, "-d", detail, "-o", tmp_path / "04_output")

    after = {p.name: (p.stat().st_mtime_ns, p.stat().st_size, sha256(p))
             for p in sorted(work.glob("*.xlsx"))}

    assert set(before) == set(after), "原表目录里的文件被增删了：%s → %s" % (sorted(before), sorted(after))
    changed = [n for n in before if before[n] != after[n]]
    assert not changed, (
        "以下原表被改动了（mtime/大小/哈希 变了）：%s。SKILL 坑10：所有产物必须写到独立输出目录。" % changed
    )


def test_产物全部落在指定输出目录(fx, tmp_path):
    """-o 指定的目录之外不该冒出产物（尤其不该写到工作簿旁边）。"""
    work = tmp_path / "原表"
    work.mkdir()
    shutil.copy2(fx / mf.F01, work / mf.F01)
    before = {p.name for p in work.iterdir()}
    run_ok("tomd", work / mf.F01, "-o", tmp_path / "md")
    after = {p.name for p in work.iterdir()}
    assert before == after, "tomd 在原表目录里多写了文件：%s" % (after - before)
    assert (tmp_path / "md").exists(), "-o 指定的输出目录没被创建"


def test_无子命令时应给出用法并非零退出():
    cp = run()
    assert cp.returncode != 0, "不带子命令时应打印用法并非零退出"
    assert any(w in cp.out for w in ("tomd", "verify", "output", "analyze", "usage", "用法")), (
        "帮助信息里应列出四个子命令。实际输出:\n%s" % cp.out[-1500:]
    )
